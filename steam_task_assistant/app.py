from __future__ import annotations

import csv
import json
import os
import queue
import re
import sqlite3
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import BOTH, BOTTOM, END, LEFT, RIGHT, TOP, X, Y, BooleanVar, IntVar, StringVar, Tk, filedialog, messagebox, ttk

import openpyxl


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
SCREENSHOT_DIR = Path(os.environ.get("STEAM_ASSISTANT_SCREENSHOT_DIR", APP_DIR / "screenshots"))
DB_PATH = Path(os.environ.get("STEAM_ASSISTANT_DB_PATH", DATA_DIR / "tasks.sqlite3"))
AUTOMATION_LOG_PATH = APP_DIR / "自动化错误.log"

STEAM_LOGIN_URL = "https://store.steampowered.com/login/"
STEAM_REDEEM_URL = "https://store.steampowered.com/account/redeemwalletcode"
STEAM_ACTIVATE_URL = "https://store.steampowered.com/account/registerkey"
STEAM_HISTORY_URL = "https://store.steampowered.com/account/history/"
STEAM_LICENSES_URL = "https://store.steampowered.com/account/licenses/"
STEAM_CART_URL = "https://store.steampowered.com/cart/"
STEAM_SEARCH_URL = "https://store.steampowered.com/search/?term="
STEAM_FRIENDS_ADD_URL = "https://steamcommunity.com/my/friends/add"
STEAM_PENDING_GIFTS_URL = "https://steamcommunity.com/my/inventory/#pending_gifts"

MODES = ("兑换码兑换", "激活码激活", "游戏购买", "游玩时间检测", "好友码提货")
STATUSES = ("未处理", "处理中", "需要人工处理", "成功", "部分成功", "失败", "跳过")
REQUIRED_COLUMNS = {
    "序号",
    "处理状态",
    "区域",
    "省份",
    "城市",
    "店编",
    "店名",
    "Steam账号",
    "Steam密码",
    "购买游戏",
    "游戏价格",
    "兑换码列表",
    "激活码列表",
    "模式",
    "备注",
}

GAME_CATALOG = {
    "红色沙漠": {
        "official_name": "红色沙漠",
        "license_names": ["红色沙漠", "Crimson Desert"],
        "app_id": 3321460,
        "package_id": 1176092,
        "price": 268.0,
    },
    "极限竞速6": {
        "official_name": "极限竞速：地平线 6",
        "license_names": ["极限竞速：地平线 6", "Forza Horizon 6"],
        "app_id": 2483190,
        "package_id": 950527,
        "price": 298.0,
    },
    "极限竞速地平线6": {
        "official_name": "极限竞速：地平线 6",
        "license_names": ["极限竞速：地平线 6", "Forza Horizon 6"],
        "app_id": 2483190,
        "package_id": 950527,
        "price": 298.0,
    },
}


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)


def clean_filename_part(value: object, fallback: str = "空") -> str:
    text = str(value or fallback).strip()
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", " ", text)
    return text[:80] or fallback


def split_codes(value: object) -> list[str]:
    if value is None:
        return []
    parts = re.split(r"[\n\r;；,，]+", str(value))
    return [part.strip() for part in parts if part.strip()]


def normalize_identity_text(value: object) -> str:
    text = str(value or "").lower()
    text = re.sub(r"alienware|外星人", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)
    return text


def location_token(value: object) -> str:
    return re.sub(
        r"(特别行政区|壮族自治区|回族自治区|维吾尔自治区|自治区|省|市)$",
        "",
        str(value or ""),
    ).strip()


def store_identity_core(store_name: object, province: object, city: object) -> str:
    core = normalize_identity_text(store_name)
    removable = [
        normalize_identity_text(province),
        normalize_identity_text(city),
        normalize_identity_text(location_token(province)),
        normalize_identity_text(location_token(city)),
        "专卖店",
        "旗舰店",
        "体验店",
        "门店",
        "店",
    ]
    for item in sorted({item for item in removable if item}, key=len, reverse=True):
        core = core.replace(item, "")
    return core


def normalize_game_name(value: object) -> str:
    return re.sub(r"[\s:：·_\-]+", "", str(value or "").strip()).lower()


def resolve_game_product(game_name: object) -> dict | None:
    normalized = normalize_game_name(game_name)
    for alias, product in GAME_CATALOG.items():
        if normalize_game_name(alias) == normalized:
            return product
    return None


def product_in_text(product: dict, text: str) -> bool:
    normalized_text = normalize_game_name(text)
    names = product.get("license_names") or [product["official_name"]]
    return any(normalize_game_name(name) in normalized_text for name in names)


def copy_to_clipboard(root: Tk, value: str) -> None:
    root.clipboard_clear()
    root.clipboard_append(value)
    root.update()


@dataclass
class Task:
    id: int
    row_no: int
    status: str
    region: str
    province: str
    city: str
    store_code: str
    store_name: str
    steam_account: str
    steam_password: str
    game_name: str
    game_price: str
    voucher_codes: list[str]
    activation_codes: list[str]
    mode: str
    note: str
    screenshot_dir: str
    profile_link: str
    last_played_game: str
    last_played_time: str
    last_played_days: str
    last_played_source: str
    pickup_code: str
    pickup_url: str
    friend_link_1: str
    friend_link_2: str
    pickup_status: str


class TaskStore:
    def __init__(self, db_path: Path) -> None:
        self.conn = sqlite3.connect(db_path, timeout=30)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def _init_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                row_no INTEGER,
                status TEXT NOT NULL DEFAULT '未处理',
                region TEXT,
                province TEXT,
                city TEXT,
                store_code TEXT,
                store_name TEXT,
                steam_account TEXT,
                steam_password TEXT,
                game_name TEXT,
                game_price TEXT,
                voucher_codes TEXT,
                activation_codes TEXT,
                mode TEXT,
                note TEXT,
                screenshot_dir TEXT,
                pickup_code TEXT,
                pickup_url TEXT,
                friend_link_1 TEXT,
                friend_link_2 TEXT,
                pickup_status TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER,
                event_time TEXT DEFAULT CURRENT_TIMESTAMP,
                event_type TEXT,
                message TEXT,
                screenshot_path TEXT
            );
            CREATE TABLE IF NOT EXISTS voucher_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                code_index INTEGER NOT NULL,
                code TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '待兑换',
                message TEXT,
                screenshot_path TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(task_id, code_index)
            );
            """
        )
        existing_columns = {
            row["name"] for row in self.conn.execute("PRAGMA table_info(tasks)").fetchall()
        }
        for column_name in (
            "profile_link",
            "last_played_game",
            "last_played_time",
            "last_played_days",
            "last_played_source",
            "pickup_code",
            "pickup_url",
            "friend_link_1",
            "friend_link_2",
            "pickup_status",
        ):
            if column_name not in existing_columns:
                self.conn.execute(f"ALTER TABLE tasks ADD COLUMN {column_name} TEXT")
        self.conn.commit()

    def clear_tasks(self) -> None:
        self.conn.execute("DELETE FROM voucher_results")
        self.conn.execute("DELETE FROM events")
        self.conn.execute("DELETE FROM tasks")
        self.conn.commit()

    def _clear_mode_tasks(self, mode: str) -> None:
        ids = [
            row["id"]
            for row in self.conn.execute("SELECT id FROM tasks WHERE mode = ?", (mode,)).fetchall()
        ]
        if not ids:
            return
        placeholders = ",".join("?" for _ in ids)
        self.conn.execute(f"DELETE FROM voucher_results WHERE task_id IN ({placeholders})", ids)
        self.conn.execute(f"DELETE FROM events WHERE task_id IN ({placeholders})", ids)
        self.conn.execute(f"DELETE FROM tasks WHERE id IN ({placeholders})", ids)

    def import_excel(self, path: Path, replace: bool = True) -> int:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["任务数据库"] if "任务数据库" in wb.sheetnames else wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError("缺少标准列：" + "、".join(sorted(missing)))
        index = {name: headers.index(name) for name in headers if name}

        def optional_value(row: tuple[object, ...], column_name: str) -> object:
            column_index = index.get(column_name)
            if column_index is None or column_index >= len(row):
                return ""
            return row[column_index]

        if replace:
            self.clear_tasks()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            mode = str(row[index["模式"]] or "兑换码兑换").strip()
            if mode not in MODES:
                mode = "兑换码兑换"
            status = str(row[index["处理状态"]] or "未处理").strip()
            if status not in STATUSES:
                status = "未处理"
            values = {
                "row_no": row[index["序号"]] or count + 1,
                "status": status,
                "region": row[index["区域"]],
                "province": row[index["省份"]],
                "city": row[index["城市"]],
                "store_code": row[index["店编"]],
                "store_name": row[index["店名"]],
                "steam_account": row[index["Steam账号"]],
                "steam_password": row[index["Steam密码"]],
                "game_name": row[index["购买游戏"]],
                "game_price": row[index["游戏价格"]],
                "voucher_codes": "\n".join(split_codes(row[index["兑换码列表"]])),
                "activation_codes": "\n".join(split_codes(row[index["激活码列表"]])),
                "mode": mode,
                "note": row[index["备注"]],
                "screenshot_dir": "",
                "pickup_code": optional_value(row, "提货码"),
                "pickup_url": optional_value(row, "提货网址"),
                "friend_link_1": optional_value(row, "好友链接1"),
                "friend_link_2": optional_value(row, "好友链接2"),
                "pickup_status": optional_value(row, "提货状态"),
            }
            self.conn.execute(
                """
                INSERT INTO tasks (
                    row_no, status, region, province, city, store_code, store_name,
                    steam_account, steam_password, game_name, game_price, voucher_codes,
                    activation_codes, mode, note, screenshot_dir, pickup_code, pickup_url,
                    friend_link_1, friend_link_2, pickup_status
                ) VALUES (
                    :row_no, :status, :region, :province, :city, :store_code, :store_name,
                    :steam_account, :steam_password, :game_name, :game_price, :voucher_codes,
                    :activation_codes, :mode, :note, :screenshot_dir, :pickup_code, :pickup_url,
                    :friend_link_1, :friend_link_2, :pickup_status
                )
                """,
                values,
            )
            count += 1
        self.conn.commit()
        return count

    def import_playtime_tasks(self, workspace_root: Path, replace: bool = True) -> int:
        v38_candidates = sorted(workspace_root.glob("*V3.8*.xlsx"))
        if not v38_candidates:
            raise FileNotFoundError("未找到 V3.8 账号清单，无法匹配账号链接。")
        v38_rows = self._read_v38_account_rows(v38_candidates[0])
        account_rows = [row for row in v38_rows if row.get("sonic_code")]
        if not account_rows:
            raise ValueError("V3.8 账号清单的《索尼克赛车交叉世界》列下没有可用激活码，无法生成检测任务。")
        by_account = {
            str(row.get("steam_account") or "").strip().lower(): row
            for row in v38_rows
            if row.get("steam_account")
        }
        by_store = {
            str(row.get("store_code") or "").strip().lower(): row
            for row in v38_rows
            if row.get("store_code")
        }

        if replace:
            ids = [
                row["id"]
                for row in self.conn.execute("SELECT id FROM tasks WHERE mode = '游玩时间检测'").fetchall()
            ]
            if ids:
                placeholders = ",".join("?" for _ in ids)
                self.conn.execute(f"DELETE FROM voucher_results WHERE task_id IN ({placeholders})", ids)
                self.conn.execute(f"DELETE FROM events WHERE task_id IN ({placeholders})", ids)
                self.conn.execute(f"DELETE FROM tasks WHERE id IN ({placeholders})", ids)

        count = 0
        by_sonic_code = {
            str(row.get("sonic_code") or "").strip().lower(): row
            for row in v38_rows
            if row.get("sonic_code")
        }

        for index, row in enumerate(account_rows, start=1):
            account = str(row.get("steam_account") or "").strip()
            if not account:
                continue
            store_code = str(row.get("store_code") or "").strip()
            sonic_code = str(row.get("sonic_code") or "").strip()
            v38_row = (
                by_account.get(account.lower())
                or by_store.get(store_code.lower())
                or by_sonic_code.get(sonic_code.lower())
                or {}
            )
            values = {
                "row_no": row.get("row_no") or index,
                "status": "未处理",
                "region": row.get("region") or v38_row.get("region"),
                "province": row.get("province") or v38_row.get("province"),
                "city": row.get("city") or v38_row.get("city"),
                "store_code": store_code or v38_row.get("store_code"),
                "store_name": row.get("store_name") or v38_row.get("store_name"),
                "steam_account": account,
                "steam_password": row.get("steam_password") or v38_row.get("steam_password"),
                "game_name": "",
                "game_price": "",
                "voucher_codes": "",
                "activation_codes": sonic_code,
                "mode": "游玩时间检测",
                "note": "V3.8索尼克赛车交叉世界激活码账号范围；检测账号上一次游玩的游戏与时间",
                "screenshot_dir": "",
                "profile_link": v38_row.get("profile_link") or "",
                "last_played_game": "",
                "last_played_time": "",
                "last_played_days": "",
                "last_played_source": "",
            }
            self.conn.execute(
                """
                INSERT INTO tasks (
                    row_no, status, region, province, city, store_code, store_name,
                    steam_account, steam_password, game_name, game_price, voucher_codes,
                    activation_codes, mode, note, screenshot_dir, profile_link,
                    last_played_game, last_played_time, last_played_days, last_played_source
                ) VALUES (
                    :row_no, :status, :region, :province, :city, :store_code, :store_name,
                    :steam_account, :steam_password, :game_name, :game_price, :voucher_codes,
                    :activation_codes, :mode, :note, :screenshot_dir, :profile_link,
                    :last_played_game, :last_played_time, :last_played_days, :last_played_source
                )
                """,
                values,
            )
            count += 1
        self.conn.commit()
        return count

    def import_friend_claim_tasks(self, workspace_root: Path, replace: bool = True) -> int:
        v38_candidates = sorted(workspace_root.glob("*V3.8*.xlsx"))
        if not v38_candidates:
            raise FileNotFoundError("未找到 V3.8 账号清单，无法匹配58家账号的登录信息。")
        report_candidates = list(
            workspace_root.glob(
                "outputs/**/ALIENWARE零售渠道Steam账号清单 《红色沙漠》部署.xlsx"
            )
        )
        if not report_candidates:
            raise FileNotFoundError("未找到《红色沙漠》部署报告，无法确认58家账号范围。")

        report_path = max(report_candidates, key=lambda path: path.stat().st_mtime)
        target_accounts = self._read_red_desert_deployment_accounts(report_path)
        if len(target_accounts) != 58 or len({account.lower() for account in target_accounts}) != 58:
            raise ValueError(
                f"《红色沙漠》部署报告应包含58个唯一账号，当前读取到{len(target_accounts)}个。"
            )

        v38_rows = self._read_v38_account_rows(v38_candidates[0])
        by_account = {
            str(row.get("steam_account") or "").strip().lower(): row
            for row in v38_rows
            if row.get("steam_account") and row.get("steam_password")
        }
        missing_accounts = [
            account for account in target_accounts if account.lower() not in by_account
        ]
        if missing_accounts:
            raise ValueError(
                f"有{len(missing_accounts)}个部署账号未能在V3.8清单中匹配账号和密码。"
            )

        if replace:
            self._clear_mode_tasks("好友码提货")

        for index, account in enumerate(target_accounts, start=1):
            row = by_account[account.lower()]
            values = {
                "row_no": index,
                "status": "未处理",
                "region": row.get("region") or "",
                "province": row.get("province") or "",
                "city": row.get("city") or "",
                "store_code": row.get("store_code") or "",
                "store_name": row.get("store_name") or "",
                "steam_account": row.get("steam_account") or "",
                "steam_password": row.get("steam_password") or "",
                "game_name": "红色沙漠",
                "game_price": "",
                "voucher_codes": "",
                "activation_codes": "",
                "mode": "好友码提货",
                "note": "红色沙漠部署58家门店：批量采集两条Steam快速邀请链接",
                "screenshot_dir": "",
                "profile_link": row.get("profile_link") or "",
                "pickup_code": "",
                "pickup_url": "",
                "friend_link_1": "",
                "friend_link_2": "",
                "pickup_status": "待采集",
            }
            self.conn.execute(
                """
                INSERT INTO tasks (
                    row_no, status, region, province, city, store_code, store_name,
                    steam_account, steam_password, game_name, game_price, voucher_codes,
                    activation_codes, mode, note, screenshot_dir, profile_link, pickup_code,
                    pickup_url, friend_link_1, friend_link_2, pickup_status
                ) VALUES (
                    :row_no, :status, :region, :province, :city, :store_code, :store_name,
                    :steam_account, :steam_password, :game_name, :game_price, :voucher_codes,
                    :activation_codes, :mode, :note, :screenshot_dir, :profile_link, :pickup_code,
                    :pickup_url, :friend_link_1, :friend_link_2, :pickup_status
                )
                """,
                values,
            )
        self.conn.commit()
        return len(target_accounts)

    def _read_red_desert_deployment_accounts(self, path: Path) -> list[str]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "《红色沙漠》序列号" not in wb.sheetnames:
            raise ValueError("《红色沙漠》部署报告中缺少《红色沙漠》序列号工作表。")
        ws = wb["《红色沙漠》序列号"]
        account_column = None
        for cell in ws[1]:
            if str(cell.value or "").strip() == "账号":
                account_column = cell.column
                break
        if account_column is None:
            raise ValueError("《红色沙漠》序列号工作表中缺少“账号”列。")
        accounts = []
        for row_index in range(2, ws.max_row + 1):
            account = str(ws.cell(row_index, account_column).value or "").strip()
            if account:
                accounts.append(account)
        return accounts

    def _read_sonic_racing_rows(self, path: Path) -> list[dict[str, str]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        header_row = 3
        sonic_col = None
        for cell in ws[header_row]:
            text = str(cell.value or "")
            if "索尼克赛车交叉世界" in text:
                sonic_col = cell.column
                break
        if sonic_col is None:
            raise ValueError("Q1 FY27 游戏部署表中未找到《索尼克赛车交叉世界》列。")

        rows = []
        for row_index in range(header_row + 1, ws.max_row + 1):
            code = str(ws.cell(row_index, sonic_col).value or "").strip()
            if not code or code == "已部署":
                continue
            rows.append(
                {
                    "row_no": ws.cell(row_index, 1).value,
                    "region": str(ws.cell(row_index, 2).value or "").strip(),
                    "province": str(ws.cell(row_index, 3).value or "").strip(),
                    "city": str(ws.cell(row_index, 4).value or "").strip(),
                    "store_code": str(ws.cell(row_index, 6).value or "").strip(),
                    "valid_state": str(ws.cell(row_index, 7).value or "").strip(),
                    "store_name": str(ws.cell(row_index, 8).value or "").strip(),
                    "steam_account": str(ws.cell(row_index, 9).value or "").strip(),
                    "steam_password": str(ws.cell(row_index, 10).value or "").strip(),
                    "sonic_code": code,
                }
            )
        return rows

    def _read_v38_account_rows(self, path: Path) -> list[dict[str, str]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row_index in range(4, ws.max_row + 1):
            account = ws.cell(row_index, 15).value
            if not account:
                continue
            rows.append(
                {
                    "row_no": ws.cell(row_index, 1).value,
                    "region": str(ws.cell(row_index, 2).value or "").strip(),
                    "province": str(ws.cell(row_index, 3).value or "").strip(),
                    "city": str(ws.cell(row_index, 4).value or "").strip(),
                    "store_code": str(ws.cell(row_index, 6).value or "").strip(),
                    "valid_state": str(ws.cell(row_index, 7).value or "").strip(),
                    "store_name": str(ws.cell(row_index, 8).value or "").strip(),
                    "profile_link": str(ws.cell(row_index, 12).value or "").strip(),
                    "steam_account": str(ws.cell(row_index, 15).value or "").strip(),
                    "steam_password": str(ws.cell(row_index, 16).value or "").strip(),
                    "sonic_code": str(ws.cell(row_index, 22).value or "").strip(),
                }
            )
        return rows

    def filtered_tasks(self, mode_filter: str) -> list[Task]:
        if mode_filter == "全部模式":
            rows = self.conn.execute("SELECT * FROM tasks ORDER BY row_no, id").fetchall()
        else:
            rows = self.conn.execute("SELECT * FROM tasks WHERE mode = ? ORDER BY row_no, id", (mode_filter,)).fetchall()
        return [self._row_to_task(row) for row in rows]

    def get_task(self, task_id: int | None) -> Task | None:
        if task_id is None:
            return None
        row = self.conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
        return self._row_to_task(row) if row else None

    def next_task_after(self, task_id: int | None, mode_filter: str) -> Task | None:
        mode_clause = "" if mode_filter == "全部模式" else " AND mode = ?"
        mode_params = () if mode_filter == "全部模式" else (mode_filter,)
        if task_id is None:
            row = self.conn.execute(
                f"SELECT * FROM tasks WHERE status IN ('未处理','处理中','需要人工处理'){mode_clause} ORDER BY row_no, id LIMIT 1",
                mode_params,
            ).fetchone()
        else:
            current = self.conn.execute("SELECT row_no, id FROM tasks WHERE id = ?", (task_id,)).fetchone()
            if not current:
                return self.next_task_after(None, mode_filter)
            row = self.conn.execute(
                f"""
                SELECT * FROM tasks
                WHERE (row_no > ? OR (row_no = ? AND id > ?))
                  AND status IN ('未处理','处理中','需要人工处理')
                  {mode_clause}
                ORDER BY row_no, id LIMIT 1
                """,
                (current["row_no"], current["row_no"], current["id"], *mode_params),
            ).fetchone()
        return self._row_to_task(row) if row else None

    def pending_purchase_tasks(
        self,
        start_task_id: int | None,
        limit: int,
    ) -> list[Task]:
        tasks = [
            task
            for task in self.filtered_tasks("游戏购买")
            if task.status in {"未处理", "处理中", "需要人工处理"}
        ]
        if start_task_id is not None:
            start_index = next(
                (index for index, task in enumerate(tasks) if task.id == start_task_id),
                0,
            )
            tasks = tasks[start_index:]
        return tasks if limit == 0 else tasks[:limit]

    def update_status(self, task_id: int, status: str) -> None:
        self.conn.execute("UPDATE tasks SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (status, task_id))
        self.conn.commit()

    def update_screenshot_dir(self, task_id: int, screenshot_dir: Path) -> None:
        self.conn.execute(
            "UPDATE tasks SET screenshot_dir = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (str(screenshot_dir), task_id),
        )
        self.conn.commit()

    def upsert_voucher_result(
        self,
        task_id: int,
        code_index: int,
        code: str,
        status: str,
        message: str = "",
        screenshot_path: str = "",
    ) -> None:
        self.conn.execute(
            """
            INSERT INTO voucher_results (
                task_id, code_index, code, status, message, screenshot_path, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(task_id, code_index) DO UPDATE SET
                code = excluded.code,
                status = excluded.status,
                message = excluded.message,
                screenshot_path = excluded.screenshot_path,
                updated_at = CURRENT_TIMESTAMP
            """,
            (task_id, code_index, code, status, message, screenshot_path),
        )
        self.conn.commit()

    def voucher_results(self, task_id: int) -> dict[int, sqlite3.Row]:
        rows = self.conn.execute(
            "SELECT * FROM voucher_results WHERE task_id = ? ORDER BY code_index",
            (task_id,),
        ).fetchall()
        return {int(row["code_index"]): row for row in rows}

    def pending_redeem_tasks(self, start_task_id: int | None = None, limit: int = 0) -> list[Task]:
        rows = self.conn.execute(
            """
            SELECT * FROM tasks
            WHERE mode = '兑换码兑换'
              AND status IN ('未处理','处理中','需要人工处理')
              AND COALESCE(voucher_codes, '') <> ''
            ORDER BY row_no, id
            """
        ).fetchall()
        tasks = [self._row_to_task(row) for row in rows]
        if start_task_id is not None:
            start_index = next((i for i, task in enumerate(tasks) if task.id == start_task_id), 0)
            tasks = tasks[start_index:]
        if limit > 0:
            tasks = tasks[:limit]
        return tasks

    def pending_activation_tasks(self, start_task_id: int | None = None, limit: int = 0) -> list[Task]:
        rows = self.conn.execute(
            """
            SELECT * FROM tasks
            WHERE mode = '激活码激活'
              AND status IN ('未处理','处理中','需要人工处理')
              AND COALESCE(activation_codes, '') <> ''
            ORDER BY row_no, id
            """
        ).fetchall()
        tasks = [self._row_to_task(row) for row in rows]
        if start_task_id is not None:
            start_index = next((i for i, task in enumerate(tasks) if task.id == start_task_id), 0)
            tasks = tasks[start_index:]
        if limit > 0:
            tasks = tasks[:limit]
        return tasks

    def pending_playtime_tasks(self, start_task_id: int | None = None, limit: int = 0) -> list[Task]:
        rows = self.conn.execute(
            """
            SELECT * FROM tasks
            WHERE mode = '游玩时间检测'
              AND status IN ('未处理','处理中','需要人工处理')
            ORDER BY row_no, id
            """
        ).fetchall()
        tasks = [self._row_to_task(row) for row in rows]
        if start_task_id is not None:
            start_index = next((i for i, task in enumerate(tasks) if task.id == start_task_id), 0)
            tasks = tasks[start_index:]
        if limit > 0:
            tasks = tasks[:limit]
        return tasks

    def pending_friend_claim_tasks(
        self,
        start_task_id: int | None = None,
        limit: int = 0,
    ) -> list[Task]:
        rows = self.conn.execute(
            """
            SELECT * FROM tasks
            WHERE mode = '好友码提货'
              AND status IN ('未处理','处理中','需要人工处理')
              AND (
                    COALESCE(friend_link_1, '') = ''
                    OR COALESCE(friend_link_2, '') = ''
                  )
            ORDER BY row_no, id
            """
        ).fetchall()
        tasks = [self._row_to_task(row) for row in rows]
        if start_task_id is not None:
            start_index = next((i for i, task in enumerate(tasks) if task.id == start_task_id), 0)
            tasks = tasks[start_index:]
        if limit > 0:
            tasks = tasks[:limit]
        return tasks

    def pending_friend_game_claim_tasks(
        self,
        start_task_id: int | None = None,
        limit: int = 0,
    ) -> list[Task]:
        rows = self.conn.execute(
            """
            SELECT * FROM tasks
            WHERE mode = '好友码提货'
              AND COALESCE(game_name, '') <> ''
              AND status IN ('未处理','处理中','需要人工处理','成功','部分成功')
              AND COALESCE(pickup_status, '') NOT LIKE '自动游戏领取完成%'
            ORDER BY row_no, id
            """
        ).fetchall()
        tasks = [self._row_to_task(row) for row in rows]
        if start_task_id is not None:
            start_index = next((i for i, task in enumerate(tasks) if task.id == start_task_id), 0)
            tasks = tasks[start_index:]
        if limit > 0:
            tasks = tasks[:limit]
        return tasks

    def update_playtime_result(
        self,
        task_id: int,
        status: str,
        game: str,
        played_time: str,
        days: str,
        source: str,
        note: str = "",
        screenshot_path: str = "",
    ) -> None:
        self.conn.execute(
            """
            UPDATE tasks
            SET status = ?,
                last_played_game = ?,
                last_played_time = ?,
                last_played_days = ?,
                last_played_source = ?,
                note = CASE WHEN ? <> '' THEN ? ELSE note END,
                screenshot_dir = CASE WHEN ? <> '' THEN ? ELSE screenshot_dir END,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                status,
                game,
                played_time,
                days,
                source,
                note,
                note,
                str(Path(screenshot_path).parent) if screenshot_path else "",
                str(Path(screenshot_path).parent) if screenshot_path else "",
                task_id,
            ),
        )
        self.conn.commit()

    def update_friend_claim_links(
        self,
        task_id: int,
        friend_link_1: str,
        friend_link_2: str,
    ) -> None:
        self.conn.execute(
            """
            UPDATE tasks
            SET friend_link_1 = ?,
                friend_link_2 = ?,
                pickup_status = '好友链接已采集',
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (friend_link_1, friend_link_2, task_id),
        )
        self.conn.commit()

    def update_friend_claim_link(
        self,
        task_id: int,
        link_number: int,
        friend_link: str,
    ) -> None:
        """Persist each invite URL immediately so a later step cannot lose it."""
        if link_number == 1:
            self.conn.execute(
                """
                UPDATE tasks
                SET friend_link_1 = ?,
                    pickup_status = '已采集好友链接1，待生成好友链接2',
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (friend_link, task_id),
            )
        elif link_number == 2:
            self.conn.execute(
                """
                UPDATE tasks
                SET friend_link_2 = ?,
                    pickup_status = '已采集两条好友链接',
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (friend_link, task_id),
            )
        else:
            raise ValueError(f"不支持的好友链接编号：{link_number}")
        self.conn.commit()

    def update_friend_claim_state(
        self,
        task_id: int,
        status: str,
        pickup_status: str,
        screenshot_path: str = "",
    ) -> None:
        self.conn.execute(
            """
            UPDATE tasks
            SET status = ?,
                pickup_status = ?,
                screenshot_dir = CASE WHEN ? <> '' THEN ? ELSE screenshot_dir END,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                status,
                pickup_status,
                str(Path(screenshot_path).parent) if screenshot_path else "",
                str(Path(screenshot_path).parent) if screenshot_path else "",
                task_id,
            ),
        )
        self.conn.commit()

    def add_event(self, task_id: int, event_type: str, message: str, screenshot_path: str = "") -> None:
        self.conn.execute(
            "INSERT INTO events (task_id, event_type, message, screenshot_path) VALUES (?, ?, ?, ?)",
            (task_id, event_type, message, screenshot_path),
        )
        self.conn.commit()

    def events(self, task_id: int | None, limit: int = 250) -> list[sqlite3.Row]:
        if task_id is None:
            return []
        return self.conn.execute(
            "SELECT * FROM events WHERE task_id = ? ORDER BY id DESC LIMIT ?",
            (task_id, limit),
        ).fetchall()

    def export_csv(self, path: Path) -> None:
        rows = self.conn.execute("SELECT * FROM tasks ORDER BY row_no, id").fetchall()
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "序号",
                    "处理状态",
                    "区域",
                    "省份",
                    "城市",
                    "店编",
                    "店名",
                    "Steam账号",
                    "Steam密码",
                    "购买游戏",
                    "游戏价格",
                    "兑换码列表",
                    "激活码列表",
                    "模式",
                    "备注",
                    "截图文件夹",
                    "账号链接",
                    "上次游玩游戏",
                    "上次游玩时间",
                    "距离今天",
                    "检测来源",
                    "提货码",
                    "提货网址",
                    "好友链接1",
                    "好友链接2",
                    "提货状态",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row["row_no"],
                        row["status"],
                        row["region"],
                        row["province"],
                        row["city"],
                        row["store_code"],
                        row["store_name"],
                        row["steam_account"],
                        row["steam_password"],
                        row["game_name"],
                        row["game_price"],
                        row["voucher_codes"],
                        row["activation_codes"],
                        row["mode"],
                        row["note"],
                        row["screenshot_dir"],
                        row["profile_link"],
                        row["last_played_game"],
                        row["last_played_time"],
                        row["last_played_days"],
                        row["last_played_source"],
                        row["pickup_code"],
                        row["pickup_url"],
                        row["friend_link_1"],
                        row["friend_link_2"],
                        row["pickup_status"],
                    ]
                )

    def _row_to_task(self, row: sqlite3.Row) -> Task:
        keys = set(row.keys())
        return Task(
            id=row["id"],
            row_no=row["row_no"],
            status=row["status"] or "未处理",
            region=row["region"] or "",
            province=row["province"] or "",
            city=row["city"] or "",
            store_code=row["store_code"] or "",
            store_name=row["store_name"] or "",
            steam_account=row["steam_account"] or "",
            steam_password=row["steam_password"] or "",
            game_name=row["game_name"] or "",
            game_price=str(row["game_price"] or ""),
            voucher_codes=split_codes(row["voucher_codes"]),
            activation_codes=split_codes(row["activation_codes"]),
            mode=row["mode"] or "兑换码兑换",
            note=row["note"] or "",
            screenshot_dir=row["screenshot_dir"] or "",
            profile_link=row["profile_link"] or "" if "profile_link" in keys else "",
            last_played_game=row["last_played_game"] or "" if "last_played_game" in keys else "",
            last_played_time=row["last_played_time"] or "" if "last_played_time" in keys else "",
            last_played_days=row["last_played_days"] or "" if "last_played_days" in keys else "",
            last_played_source=row["last_played_source"] or "" if "last_played_source" in keys else "",
            pickup_code=row["pickup_code"] or "" if "pickup_code" in keys else "",
            pickup_url=row["pickup_url"] or "" if "pickup_url" in keys else "",
            friend_link_1=row["friend_link_1"] or "" if "friend_link_1" in keys else "",
            friend_link_2=row["friend_link_2"] or "" if "friend_link_2" in keys else "",
            pickup_status=row["pickup_status"] or "" if "pickup_status" in keys else "",
        )


class BrowserWorker:
    def __init__(self, app: "SteamTaskAssistant") -> None:
        self.app = app
        self.commands: queue.Queue[tuple[str, tuple]] = queue.Queue()
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def submit(self, command: str, *args) -> None:
        self.commands.put((command, args))

    def shutdown(self) -> None:
        self.commands.put(("__shutdown__", ()))

    def _run(self) -> None:
        controller = None
        while True:
            command, args = self.commands.get()
            try:
                if command == "__shutdown__":
                    if controller is not None:
                        controller.close()
                    controller = None
                    return
                if controller is None:
                    controller = SteamBrowserController(self.app)
                getattr(controller, command)(*args)
            except Exception as exc:
                try:
                    if controller is not None:
                        controller.close()
                except Exception:
                    pass
                controller = None
                self.app.ui_error(f"浏览器自动化失败：{exc}")
            finally:
                self.commands.task_done()


class SteamBrowserController:
    def __init__(self, app: "SteamTaskAssistant") -> None:
        from playwright.sync_api import sync_playwright

        self.app = app
        self.playwright = sync_playwright().start()
        browser_args = ["--start-maximized"]
        if os.environ.get("STEAM_ASSISTANT_USE_SYSTEM_PROXY", "").lower() not in {"1", "true", "yes"}:
            browser_args.extend(["--no-proxy-server", "--proxy-server=direct://", "--proxy-bypass-list=*"])
        self.context = self.playwright.chromium.launch_persistent_context(
            user_data_dir=str(APP_DIR / "browser_profile"),
            headless=False,
            viewport={"width": 1366, "height": 900},
            args=browser_args,
        )
        self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
        self.page.set_default_timeout(20000)

    def close(self) -> None:
        try:
            self.context.close()
        finally:
            self.playwright.stop()

    def _ensure_page(self) -> None:
        if self.context.pages:
            for page in self.context.pages:
                if not page.is_closed():
                    self.page = page
                    return
        self.page = self.context.new_page()

    def _log(self, message: str) -> None:
        self.app.ui_log("浏览器", message)

    def _stopped(self) -> bool:
        return self.app.emergency_stopped or self.app.auto_stop_event.is_set()

    def _wait_auto_gate(self) -> bool:
        while self.app.auto_pause_event.is_set() and not self._stopped():
            time.sleep(0.2)
        return not self._stopped()

    def _goto(self, url: str) -> None:
        if self._stopped():
            return
        last_error = None
        for attempt in range(3):
            self._ensure_page()
            try:
                self.page.goto(url, wait_until="domcontentloaded")
                break
            except Exception as exc:
                last_error = exc
                message = str(exc)
                recoverable = (
                    "ERR_ABORTED" in message
                    or "is interrupted by another navigation" in message
                )
                if not recoverable or attempt == 2:
                    raise
                self._log("检测到页面仍在跳转，等待后重试")
                time.sleep(1.5)
        else:
            raise RuntimeError(f"无法打开页面：{last_error}")
        self._log(f"打开页面：{url}")

    def _goto_friend_claim_page(self, url: str, page_name: str) -> bool:
        """Open a Steam page without treating slow secondary assets as a navigation failure."""
        if self._stopped():
            return False
        last_error: Exception | None = None
        for attempt in range(2):
            self._ensure_page()
            try:
                self.page.goto(url, wait_until="commit", timeout=45000)
                self._log(f"打开{page_name}：{url}")
                return True
            except Exception as exc:
                last_error = exc
                if self._stopped():
                    return False
                message = str(exc)
                retryable = any(
                    marker in message
                    for marker in (
                        "Timeout",
                        "ERR_CONNECTION",
                        "ERR_ABORTED",
                        "is interrupted by another navigation",
                    )
                )
                if not retryable or attempt == 1:
                    break
                self._log(f"{page_name}加载缓慢，等待后重试")
                time.sleep(2)
        self._log(f"无法打开{page_name}：{last_error}")
        return False

    def _first_visible(self, selectors: list[str], timeout_ms: int = 12000):
        self._ensure_page()
        deadline = time.time() + timeout_ms / 1000
        while time.time() < deadline and not self._stopped():
            for selector in selectors:
                locator = self.page.locator(selector).first
                try:
                    if locator.count() > 0 and locator.is_visible(timeout=500):
                        return locator
                except Exception:
                    continue
            time.sleep(0.2)
        return None

    def _fill(self, selectors: list[str], value: str, label: str) -> bool:
        locator = self._first_visible(selectors)
        if locator is None:
            self._log(f"没有找到{label}输入框")
            return False
        locator.click()
        locator.fill(value)
        self._log(f"已填入{label}")
        return True

    def _click(self, selectors: list[str], label: str) -> bool:
        locator = self._first_visible(selectors)
        if locator is None:
            self._log(f"没有找到{label}")
            return False
        try:
            locator.click()
        except Exception:
            try:
                locator.click(force=True)
            except Exception:
                try:
                    locator.evaluate("element => element.click()")
                except Exception as exc:
                    self._log(f"点击{label}失败：{exc}")
                    return False
        self._log(f"已点击{label}")
        return True

    def _fill_login_credentials(self, account: str, password: str) -> bool:
        self._ensure_page()
        try:
            password_locator = self.page.locator("input[type='password']").first
            password_locator.wait_for(state="visible", timeout=30000)
        except Exception:
            self._log("登录页密码框没有加载出来，已停止填写，避免误填到搜索框")
            return False

        password_box = password_locator.bounding_box()
        if not password_box:
            self._log("无法读取密码框位置，已停止填写")
            return False

        account_locator = None
        text_inputs = self.page.locator("input[type='text'], input:not([type]), input[type='email']")
        best_distance = 999999
        for index in range(text_inputs.count()):
            candidate = text_inputs.nth(index)
            try:
                if not candidate.is_visible(timeout=500):
                    continue
                box = candidate.bounding_box()
                if not box:
                    continue
                vertically_near = password_box["y"] - 160 <= box["y"] <= password_box["y"] + 10
                horizontally_near = abs(box["x"] - password_box["x"]) < 260
                if not vertically_near or not horizontally_near:
                    continue
                distance = abs((password_box["y"] - box["y"]) - 58) + abs(password_box["x"] - box["x"])
                if distance < best_distance:
                    best_distance = distance
                    account_locator = candidate
            except Exception:
                continue

        if account_locator is None:
            account_locator = self._first_visible(
                [
                    "div[class*='newlogindialog'] input[type='text']",
                    "form input[type='text']",
                    "input[autocomplete='username']",
                    "input[name='username']",
                ],
                timeout_ms=5000,
            )
        if account_locator is None:
            self._log("没有找到登录账号框，已停止填写，避免误填到搜索框")
            return False

        account_locator.click()
        account_locator.fill(account)
        password_locator.click()
        password_locator.fill(password)
        self._log("已在登录表单中填入账号和密码")
        return True

    def _submit_login(self) -> None:
        self._ensure_page()
        try:
            password_box = self.page.locator("input[type='password']").first.bounding_box()
        except Exception:
            password_box = None

        if password_box:
            candidates = self.page.locator(
                "button, div[role='button'], a[role='button'], "
                "[class*='Submit'], [class*='submit'], [class*='Button'], [class*='button']"
            )
            best = None
            best_score = 999999
            for index in range(candidates.count()):
                candidate = candidates.nth(index)
                try:
                    if not candidate.is_visible(timeout=300):
                        continue
                    box = candidate.bounding_box()
                    if not box:
                        continue
                    below_password = password_box["y"] + 20 <= box["y"] <= password_box["y"] + 220
                    horizontally_close = abs((box["x"] + box["width"] / 2) - (password_box["x"] + password_box["width"] / 2)) < 360
                    if not below_password or not horizontally_close:
                        continue
                    text = (candidate.inner_text(timeout=300) or "").strip().lower()
                    score = abs(box["y"] - (password_box["y"] + 95))
                    if "sign" in text or "登录" in text or "log" in text:
                        score -= 100
                    if score < best_score:
                        best_score = score
                        best = candidate
                except Exception:
                    continue
            if best is not None:
                try:
                    best.click(force=True)
                except Exception:
                    best.evaluate("el => el.click()")
                self._log("已点击密码框下方的登录按钮")
                return

        clicked = self._click(
            [
                "[class*='SubmitButton']",
                "[class*='submit']",
                "button[type='submit']",
                "button:has-text('Sign in')",
                "button:has-text('登录')",
                "div[role='button']:has-text('Sign in')",
                "div[role='button']:has-text('登录')",
            ],
            "登录按钮",
        )
        if clicked:
            return
        self.page.locator("input[type='password']").first.press("Enter")
        self._log("未找到登录按钮，已在密码框内回车提交")

    def _wait_logged_in(self) -> bool:
        deadline = time.time() + 60
        while time.time() < deadline and not self._stopped():
            try:
                if self.page.locator("#account_pulldown").count() > 0:
                    self._log("已确认登录成功：检测到右上角账号菜单")
                    return True
                if self.page.locator("a[href*='logout']").count() > 0:
                    self._log("已确认登录成功：检测到退出链接")
                    return True
            except Exception:
                pass
            try:
                if self.page.locator("input[type='password']").count() > 0:
                    self._log("仍在登录页，继续等待登录完成")
            except Exception:
                pass
            time.sleep(2)
        self._log("60秒内未检测到登录成功，已停止后续跳转，请检查是否需要手动点击登录或页面有提示")
        return False
    def _is_logged_in(self) -> bool:
        self._ensure_page()
        try:
            return (
                self.page.locator("#account_pulldown").count() > 0
                or self.page.locator("a[href*='logout']").count() > 0
            )
        except Exception:
            return False

    def ensure_logged_out_before_new_account(self) -> bool:
        self._goto("https://store.steampowered.com/")
        if self._stopped():
            return False
        time.sleep(1)
        if self._is_logged_in():
            self._log("检测到浏览器中已有登录账号，先退出当前账号")
            return self.logout_by_menu()
        else:
            self._log("未检测到已登录账号，继续登录当前任务账号")
            return True

    def start_redeem_account(self, task: Task) -> None:
        if not self.ensure_logged_out_before_new_account():
            return
        self._goto(STEAM_LOGIN_URL)
        if self._stopped():
            return
        if not self._fill_login_credentials(task.steam_account, task.steam_password):
            return
        self._submit_login()
        if not self._wait_logged_in():
            return
        self.capture_history_then_redeem(task)

    def capture_history_then_redeem(self, task: Task) -> None:
        self._goto(STEAM_HISTORY_URL)
        if self._stopped():
            return
        time.sleep(2)
        path = self.app.make_screenshot_path(task, "消费记录")
        self.page.screenshot(path=str(path), full_page=True)
        self.app.ui_screenshot_saved(task.id, path, "消费记录")
        time.sleep(1)
        self._goto(STEAM_REDEEM_URL)

    def fill_redeem_code(self, code: str) -> None:
        self._goto(STEAM_REDEEM_URL)
        self._fill(
            [
                "#wallet_code",
                "input[name='wallet_code']",
                "input[type='text']",
            ],
            code,
            "兑换码",
        )

    def confirm_redeem_code(self) -> bool:
        clicked = self._click(
            [
                "#validate_btn",
                "button[type='submit']",
                "button:has-text('Continue')",
                "button:has-text('继续')",
                "button:has-text('兑换')",
            ],
            "确认兑换按钮",
        )
        if not clicked:
            self._log("确认按钮未定位到，请人工检查页面")
        return clicked

    def _capture_page(self, task: Task, status_text: str) -> Path:
        path = self.app.make_screenshot_path(task, status_text)
        self.page.screenshot(path=str(path), full_page=True)
        self.app.ui_screenshot_saved(task.id, path, status_text)
        return path

    def _capture_history(self, task: Task, status_text: str) -> Path | None:
        self._goto(STEAM_HISTORY_URL)
        if self._stopped():
            return None
        time.sleep(2)
        return self._capture_page(task, status_text)

    def _playtime_url(self, task: Task, use_login: bool) -> str | None:
        if use_login:
            return "https://steamcommunity.com/my/games/?tab=recent"
        link = (task.profile_link or "").strip()
        if not link:
            return None
        link = link.replace("http://", "https://").rstrip("/")
        return f"{link}/games/?tab=recent"

    def _decode_js_string(self, value: str) -> str:
        try:
            return json.loads(f'"{value}"')
        except Exception:
            return value.replace("\\/", "/").replace('\\"', '"').replace("\\'", "'")

    def _parse_playtime_from_page(self) -> tuple[str, str, str, str]:
        self._ensure_page()
        try:
            self.page.wait_for_load_state("networkidle", timeout=12000)
        except Exception:
            pass
        time.sleep(2)
        content = self.page.content()
        candidates: list[tuple[int, str]] = []
        patterns = [
            r'"name"\s*:\s*"((?:\\.|[^"\\])*)".{0,1200}?"last_played"\s*:\s*(\d+)',
            r'"last_played"\s*:\s*(\d+).{0,1200}?"name"\s*:\s*"((?:\\.|[^"\\])*)"',
            r'data-name="([^"]+)".{0,1200}?data-last_played="(\d+)"',
            r'data-last_played="(\d+)".{0,1200}?data-name="([^"]+)"',
        ]
        for pattern in patterns:
            for match in re.finditer(pattern, content, flags=re.IGNORECASE | re.DOTALL):
                groups = match.groups()
                if groups[0].isdigit():
                    timestamp = int(groups[0])
                    game_name = self._decode_js_string(groups[1])
                else:
                    game_name = self._decode_js_string(groups[0])
                    timestamp = int(groups[1])
                if timestamp > 0 and game_name:
                    candidates.append((timestamp, game_name))
        if candidates:
            timestamp, game_name = max(candidates, key=lambda item: item[0])
            played_dt = datetime.fromtimestamp(timestamp)
            days = (datetime.now().date() - played_dt.date()).days
            return game_name, played_dt.strftime("%Y-%m-%d %H:%M:%S"), f"{days}天前", "Steam游戏列表"

        try:
            body_text = (self.page.locator("body").inner_text(timeout=3000) or "").strip()
        except Exception:
            body_text = ""
        lowered = body_text.lower()
        if "sign in" in lowered or "login" in lowered or "登录" in body_text:
            return "", "", "", "需要登录"
        if "private" in lowered or "私密" in body_text:
            return "", "", "", "资料私密"
        if "no games" in lowered or "没有游戏" in body_text:
            return "", "", "", "未显示游戏"
        return "", "", "", "页面未解析到last_played"

    def _finish_playtime_detection(self, task: Task, source_label: str) -> bool:
        path = self._capture_page(task, f"游玩时间检测-{source_label}")
        game, played_time, days, source = self._parse_playtime_from_page()
        if game and played_time:
            self.app.ui_playtime_result(
                task.id,
                "成功",
                game,
                played_time,
                days,
                source_label,
                f"检测成功：{game}，{played_time}，{days}",
                path,
            )
            return True
        message = f"未能确认最近游玩记录：{source}"
        self.app.ui_playtime_result(
            task.id,
            "需要人工处理",
            "",
            "",
            "",
            source_label,
            message,
            path,
        )
        return False

    def detect_playtime_public(self, task: Task) -> bool:
        url = self._playtime_url(task, use_login=False)
        if not url:
            self.app.ui_playtime_result(
                task.id,
                "需要人工处理",
                "",
                "",
                "",
                "公开页",
                "未匹配账号链接，无法公开检测",
            )
            return False
        self._goto(url)
        if self._stopped():
            return False
        return self._finish_playtime_detection(task, "公开页")

    def detect_playtime_login(self, task: Task) -> bool:
        if not self.ensure_logged_out_before_new_account():
            self.app.ui_playtime_result(task.id, "需要人工处理", "", "", "", "登录检测", "开始新账号前无法确认登出")
            return False
        self._goto(STEAM_LOGIN_URL)
        if not self._fill_login_credentials(task.steam_account, task.steam_password):
            self.app.ui_playtime_result(task.id, "需要人工处理", "", "", "", "登录检测", "登录表单填写失败")
            return False
        self._submit_login()
        if not self._wait_logged_in():
            self.app.ui_playtime_result(task.id, "需要人工处理", "", "", "", "登录检测", "登录结果未确认")
            return False
        if not self._verify_task_identity(task):
            self.app.ui_playtime_result(task.id, "需要人工处理", "", "", "", "登录检测", "账号省市/店名核心身份核对失败")
            return False
        self._goto(self._playtime_url(task, use_login=True) or "https://steamcommunity.com/my/games/?tab=recent")
        if self._stopped():
            return False
        ok = self._finish_playtime_detection(task, "登录检测")
        if not self.logout_by_menu():
            self.app.ui_playtime_result(task.id, "需要人工处理", "", "", "", "登录检测", "检测完成后账号登出未确认")
            return False
        return ok

    def full_auto_playtime(self, tasks: list[Task], use_login: bool) -> None:
        self.app.ui_playtime_batch_started(len(tasks), use_login)
        stopped = False
        for task_number, task in enumerate(tasks, start=1):
            if not self._wait_auto_gate():
                stopped = True
                break
            self.app.ui_playtime_task_started(task, task_number, len(tasks))
            if use_login:
                self.detect_playtime_login(task)
            else:
                self.detect_playtime_public(task)
            if self._stopped():
                stopped = True
                break
        self.app.ui_playtime_batch_finished(stopped)

    def _verify_task_identity(self, task: Task) -> bool:
        locator = self._first_visible(["#account_pulldown", ".pulldown.global_action_link"], timeout_ms=5000)
        if locator is None:
            self._log("无法读取右上角账号名称，不能执行全自动身份核对")
            return False
        try:
            display_name = (locator.inner_text(timeout=1000) or "").strip()
        except Exception:
            return False

        display_normalized = normalize_identity_text(display_name)
        expected_locations = [
            normalize_identity_text(location_token(task.city)),
            normalize_identity_text(location_token(task.province)),
        ]
        expected_locations = [item for item in expected_locations if len(item) >= 2]
        if any(item in display_normalized for item in expected_locations):
            self._log(f"账号身份核对通过（省市）：{display_name}")
            return True

        store_core = store_identity_core(task.store_name, task.province, task.city)
        if len(store_core) >= 4 and (
            store_core in display_normalized or display_normalized in store_core
        ):
            self._log(f"账号身份核对通过（店名核心“{store_core}”）：{display_name}")
            return True

        self._log(
            f"账号身份核对失败：页面名称“{display_name}”既未包含省市，"
            f"也未匹配店名核心“{store_core or '空'}”"
        )
        return False

    def _redeem_result_text(self) -> str:
        selectors = [
            "[role='dialog']",
            "[aria-modal='true']",
            "[class*='modal']",
            "[class*='Modal']",
            "[class*='popup']",
            "[class*='Popup']",
            "[class*='success']",
            "[class*='Success']",
            "#error_display",
            ".error_display",
            "#purchase_result_details",
            ".checkout_error",
            ".newmodal_content",
            "#wallet_code_form",
        ]
        parts = []
        for selector in selectors:
            locator = self.page.locator(selector).first
            try:
                if locator.count() > 0 and locator.is_visible(timeout=300):
                    text = (locator.inner_text(timeout=500) or "").strip()
                    if text:
                        parts.append(text)
            except Exception:
                continue
        if parts:
            return "\n".join(parts)
        try:
            return (self.page.locator("body").inner_text(timeout=1000) or "").strip()
        except Exception:
            return ""

    def _read_wallet_balance(self) -> float | None:
        selectors = [
            "#header_wallet_balance",
            "a[href*='/account/history']",
            ".accountData.price",
            "[class*='wallet_balance']",
            "[class*='WalletBalance']",
        ]
        balances: list[float] = []
        for selector in selectors:
            locators = self.page.locator(selector)
            for index in range(min(locators.count(), 8)):
                locator = locators.nth(index)
                try:
                    if not locator.is_visible(timeout=200):
                        continue
                    text = (locator.inner_text(timeout=300) or "").replace(",", "")
                    match = re.search(r"[¥￥]\s*(\d+(?:\.\d{1,2})?)", text)
                    if match:
                        balances.append(float(match.group(1)))
                except Exception:
                    continue
        return max(balances) if balances else None

    def _classify_redeem_result(self, text: str) -> tuple[str, str]:
        normalized = re.sub(r"\s+", " ", text).strip()
        lowered = normalized.lower()
        phrase_groups = [
            (
                "兑换成功",
                [
                    "successfully redeemed",
                    "successfully added",
                    "兑换成功",
                    "成功兑换",
                    "成功!",
                    "成功！",
                    "已添加至您的 steam 钱包",
                    "已添加到您的 steam 钱包",
                ],
            ),
            ("已使用", ["already been redeemed", "already redeemed", "previously redeemed", "already owns", "已经兑换", "已被兑换", "已使用", "已经在您的帐户中", "已经在你的账户中"]),
            ("无效", ["invalid code", "not valid", "无效", "不正确", "无法识别"]),
            (
                "地区受限",
                [
                    "not available in your region",
                    "cannot be redeemed in your country",
                    "currency mismatch",
                    "无法在您所在的地区兑换",
                    "不能在您的国家/地区兑换",
                    "货币不匹配",
                ],
            ),
        ]
        for status, phrases in phrase_groups:
            if any(phrase in lowered for phrase in phrases):
                return status, normalized[:500]
        return "结果不明确", normalized[:500]

    def _wait_redeem_result(
        self,
        balance_before: float | None,
        timeout_seconds: int = 25,
    ) -> tuple[str, str]:
        deadline = time.time() + timeout_seconds
        last_text = ""
        while time.time() < deadline and not self._stopped():
            time.sleep(0.15)
            last_text = self._redeem_result_text()
            status, message = self._classify_redeem_result(last_text)
            if status != "结果不明确":
                return status, message
            balance_after = self._read_wallet_balance()
            if (
                balance_before is not None
                and balance_after is not None
                and balance_after > balance_before + 0.01
            ):
                return (
                    "兑换成功",
                    f"钱包余额由¥{balance_before:.2f}增加至¥{balance_after:.2f}",
                )
        return "结果不明确", re.sub(r"\s+", " ", last_text).strip()[:500]

    def full_auto_redeem(self, tasks: list[Task], skip_map: dict[int, set[int]] | None = None) -> None:
        skip_map = skip_map or {}
        self.app.ui_auto_batch_started(len(tasks))
        for task_number, task in enumerate(tasks, start=1):
            if not self._wait_auto_gate():
                break
            self.app.ui_auto_task_started(task, task_number, len(tasks))
            if not self.ensure_logged_out_before_new_account():
                self.app.ui_auto_paused(task.id, "开始新账号前无法确认登出")
                break

            self._goto(STEAM_LOGIN_URL)
            if not self._fill_login_credentials(task.steam_account, task.steam_password):
                self.app.ui_auto_paused(task.id, "登录表单填写失败")
                break
            self._submit_login()
            if not self._wait_logged_in():
                self.app.ui_auto_paused(task.id, "登录结果未确认")
                break
            if not self._verify_task_identity(task):
                self.app.ui_auto_paused(task.id, "账号省市/店名核心身份核对失败")
                break

            self._capture_history(task, "全自动-初始消费记录")
            self._goto(STEAM_REDEEM_URL)
            task_has_exception = False
            for code_index, code in enumerate(task.voucher_codes):
                if code_index in skip_map.get(task.id, set()):
                    self._log(f"第{code_index + 1}个兑换码已有终态结果，断点续跑时跳过")
                    continue
                if not self._wait_auto_gate():
                    break
                self.app.ui_auto_code_started(task.id, code_index, code, len(task.voucher_codes))
                if not self._fill(
                    ["#wallet_code", "input[name='wallet_code']", "input[type='text']"],
                    code,
                    "兑换码",
                ):
                    self.app.ui_auto_paused(task.id, f"第{code_index + 1}个兑换码输入框未找到")
                    task_has_exception = True
                    break
                balance_before = self._read_wallet_balance()
                if not self.confirm_redeem_code():
                    self.app.ui_auto_paused(task.id, f"第{code_index + 1}个兑换码确认按钮未找到")
                    task_has_exception = True
                    break

                status, message = self._wait_redeem_result(balance_before)
                result_path = self._capture_page(
                    task,
                    f"全自动-{status}-码{code_index + 1}-{code[-5:]}",
                )
                self.app.ui_auto_code_result(
                    task.id,
                    code_index,
                    code,
                    status,
                    message,
                    result_path,
                )
                if status == "结果不明确":
                    task_has_exception = True
                    self.app.ui_auto_paused(task.id, f"第{code_index + 1}个兑换结果无法识别")
                    if not self._wait_auto_gate():
                        break
                elif status in {"无效", "地区受限"}:
                    task_has_exception = True

                self._capture_history(task, f"全自动-码{code_index + 1}-{status}-消费记录")
                self._goto(STEAM_REDEEM_URL)

            if self._stopped():
                break
            if not self.logout_by_menu():
                self.app.ui_auto_paused(task.id, "当前账号登出未确认")
                break
            final_status = "部分成功" if task_has_exception else "成功"
            self.app.ui_auto_task_finished(task.id, final_status)
            time.sleep(1)

        self.app.ui_auto_batch_finished(self._stopped())

    def full_auto_purchase(self, tasks: list[Task]) -> None:
        self.app.ui_purchase_batch_started(len(tasks))
        stopped = False
        for task_number, task in enumerate(tasks, start=1):
            if not self._wait_auto_gate():
                stopped = True
                break
            self.app.ui_purchase_batch_task_started(
                task,
                task_number,
                len(tasks),
            )
            if not self.start_purchase_account(task):
                self.app.ui_purchase_batch_paused(task.id, "账号登录或身份核对失败")
                stopped = True
                break
            if not self.clear_purchase_cart(task, require_confirmation=True):
                self.app.ui_purchase_batch_paused(task.id, "购物车清空未完成")
                stopped = True
                break

            if not self.check_purchase_license(task):
                if not self.logout_by_menu():
                    self.app.ui_purchase_batch_paused(
                        task.id,
                        "资产已存在，但账号登出未确认",
                    )
                    stopped = True
                    break
                self.app.ui_purchase_batch_already_deployed(task.id)
                continue

            if not self.add_standard_game_to_cart(task):
                self.app.ui_purchase_batch_paused(task.id, "加入购物车失败")
                stopped = True
                break
            if not self.prepare_purchase_checkout(task):
                self.app.ui_purchase_batch_paused(task.id, "进入结算页失败")
                stopped = True
                break
            if not self.app.request_purchase_confirmation(task):
                self.app.ui_purchase_batch_paused(task.id, "用户未确认本次购买")
                stopped = True
                break
            if not self.confirm_purchase_checkout(task):
                self.app.ui_purchase_batch_paused(task.id, "购买结果未确认")
                stopped = True
                break
            if not self.verify_purchase_asset_and_logout(task):
                self.app.ui_purchase_batch_paused(task.id, "资产复核或登出失败")
                stopped = True
                break
            time.sleep(1)

        self.app.ui_purchase_batch_finished(stopped or self._stopped())

    def start_activation_account(self, task: Task) -> bool:
        if not self.ensure_logged_out_before_new_account():
            return False
        self._goto(STEAM_LOGIN_URL)
        if not self._fill_login_credentials(task.steam_account, task.steam_password):
            return False
        self._submit_login()
        if not self._wait_logged_in():
            return False
        if not self._verify_task_identity(task):
            self.app.ui_activation_issue(task.id, "账号省市/店名核心身份核对失败")
            return False
        self._goto(STEAM_ACTIVATE_URL)
        self.app.ui_activation_step(task.id, "登录完成", "已进入Steam产品激活页面")
        return True

    def fill_activation_code(self, code: str) -> None:
        self._goto(STEAM_ACTIVATE_URL)
        self._fill(
            [
                "input[name='product_key']",
                "input[type='text']",
            ],
            code,
            "激活码",
        )

    def _activation_product(self, task: Task) -> dict | None:
        return resolve_game_product(task.game_name)

    def _activation_result_text(self) -> str:
        selectors = [
            "[role='dialog']",
            "[aria-modal='true']",
            "[class*='modal']",
            "[class*='Modal']",
            "[class*='popup']",
            "[class*='Popup']",
            "[class*='success']",
            "[class*='Success']",
            "#error_display",
            ".error_display",
            ".newmodal_content",
            "#registerkey_form",
            "#game_area_purchase",
        ]
        parts = []
        for selector in selectors:
            locator = self.page.locator(selector).first
            try:
                if locator.count() > 0 and locator.is_visible(timeout=300):
                    text = (locator.inner_text(timeout=500) or "").strip()
                    if text:
                        parts.append(text)
            except Exception:
                continue
        if parts:
            return "\n".join(parts)
        try:
            return (self.page.locator("body").inner_text(timeout=1000) or "").strip()
        except Exception:
            return ""

    def _classify_activation_result(self, text: str) -> tuple[str, str]:
        normalized = re.sub(r"\s+", " ", text).strip()
        lowered = normalized.lower()
        phrase_groups = [
            (
                "激活成功",
                [
                    "activation successful",
                    "successfully activated",
                    "has been activated",
                    "激活成功",
                    "成功激活",
                    "已成功激活",
                    "已添加到您的 steam 帐户",
                    "已添加至您的 steam 帐户",
                    "已添加到你的 steam 帐户",
                ],
            ),
            (
                "已拥有",
                [
                    "already in your steam library",
                    "already registered to this account",
                    "already owns",
                    "already have",
                    "已经在您的帐户中",
                    "已经在你的账户中",
                    "已在您的帐户中",
                    "已经拥有",
                    "已拥有",
                ],
            ),
            (
                "已被使用",
                [
                    "duplicate product code",
                    "already been activated by another steam account",
                    "already been redeemed",
                    "already redeemed",
                    "already used",
                    "已由另一个 steam 帐户激活",
                    "已被另一个 steam 帐户激活",
                    "已被兑换",
                    "已被激活",
                    "已使用",
                ],
            ),
            ("无效", ["invalid product code", "invalid code", "not valid", "无效", "不正确", "无法识别"]),
            (
                "地区受限",
                [
                    "not available in your region",
                    "cannot be redeemed in your country",
                    "无法在您所在的地区",
                    "不能在您的国家",
                    "地区",
                ],
            ),
        ]
        for status, phrases in phrase_groups:
            if any(phrase in lowered for phrase in phrases):
                return status, normalized[:500]
        return "结果不明确", normalized[:500]

    def _wait_activation_result(self, timeout_seconds: int = 35) -> tuple[str, str]:
        deadline = time.time() + timeout_seconds
        last_text = ""
        while time.time() < deadline and not self._stopped():
            time.sleep(0.3)
            last_text = self._activation_result_text()
            status, message = self._classify_activation_result(last_text)
            if status != "结果不明确":
                return status, message
        return "结果不明确", re.sub(r"\s+", " ", last_text).strip()[:500]

    def _click_activation_submit(self) -> bool:
        checkbox = self._first_visible(
            [
                "#accept_ssa",
                "input[name='accept_ssa']",
                "input[type='checkbox'][name*='accept']",
                "input[type='checkbox']",
            ],
            timeout_ms=2500,
        )
        if checkbox is not None:
            try:
                if not checkbox.is_checked():
                    checkbox.check()
                    self._log("已勾选Steam订户协议确认")
            except Exception:
                try:
                    checkbox.click(force=True)
                except Exception:
                    pass
        clicked = self._click(
            [
                "#register_btn",
                "button[type='submit']",
                "input[type='submit']",
                "button:has-text('继续')",
                "button:has-text('下一步')",
                "button:has-text('Continue')",
                "button:has-text('激活')",
                "button:has-text('Activate')",
                "a:has-text('继续')",
                "a:has-text('Continue')",
            ],
            "确认激活按钮",
        )
        if clicked:
            return True
        try:
            self.page.locator("input[name='product_key'], input[type='text']").first.press("Enter")
            self._log("未找到确认激活按钮，已在激活码输入框内回车提交")
            return True
        except Exception:
            return False

    def activate_current_code(self, task: Task, code_index: int, code: str) -> bool:
        self._goto(STEAM_ACTIVATE_URL)
        if not self._fill(
            ["input[name='product_key']", "#product_key", "input[type='text']"],
            code,
            "激活码",
        ):
            self.app.ui_activation_issue(task.id, f"第{code_index + 1}个激活码输入框未找到")
            return False
        if not self._click_activation_submit():
            self.app.ui_activation_issue(task.id, f"第{code_index + 1}个激活码确认按钮未找到")
            return False
        status, message = self._wait_activation_result()
        path = self._capture_page(task, f"激活码激活-{status}-码{code_index + 1}-{code[-5:]}")
        self.app.ui_activation_code_result(task.id, code_index, code, status, message, path)
        return status in {"激活成功", "已拥有"}

    def check_activation_license(self, task: Task) -> bool:
        product = self._activation_product(task)
        if product is None:
            self.app.ui_activation_step(task.id, "资产检查", "当前任务未配置可核对的游戏产品标识")
            return False
        self._goto(STEAM_LICENSES_URL)
        time.sleep(2)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        owned = product_in_text(product, body_text)
        if owned:
            path = self._capture_page(task, "激活码激活-资产已确认")
            self.app.ui_activation_step(
                task.id,
                "资产确认",
                f"许可证与产品激活列表中已存在《{product['official_name']}》",
                path,
            )
            return True
        self.app.ui_activation_step(
            task.id,
            "资产检查",
            f"许可证与产品激活列表中未找到《{product['official_name']}》",
        )
        return False

    def finish_activation_and_logout(self, task: Task) -> bool:
        owned = self.check_activation_license(task)
        if not owned:
            self.app.ui_activation_issue(task.id, "激活后资产清单未确认目标游戏")
            return False
        if not self.logout_by_menu():
            self.app.ui_activation_issue(task.id, "激活后登出未确认")
            return False
        self.app.ui_activation_finished(task.id)
        return True

    def full_auto_activation(self, tasks: list[Task], skip_map: dict[int, set[int]] | None = None) -> None:
        skip_map = skip_map or {}
        self.app.ui_activation_batch_started(len(tasks))
        stopped = False
        for task_number, task in enumerate(tasks, start=1):
            if not self._wait_auto_gate():
                stopped = True
                break
            self.app.ui_activation_task_started(task, task_number, len(tasks))
            if not self.start_activation_account(task):
                self.app.ui_activation_batch_paused(task.id, "账号登录或身份核对失败")
                stopped = True
                break

            if self.check_activation_license(task):
                if not self.logout_by_menu():
                    self.app.ui_activation_batch_paused(task.id, "资产已存在，但账号登出未确认")
                    stopped = True
                    break
                self.app.ui_activation_finished(task.id)
                continue

            task_has_exception = False
            for code_index, code in enumerate(task.activation_codes):
                if code_index in skip_map.get(task.id, set()):
                    self._log(f"第{code_index + 1}个激活码已有终态结果，断点续跑时跳过")
                    continue
                if not self._wait_auto_gate():
                    stopped = True
                    task_has_exception = True
                    break
                self.app.ui_activation_code_started(task.id, code_index, code, len(task.activation_codes))
                success = self.activate_current_code(task, code_index, code)
                if not success:
                    task_has_exception = True
                    self.app.ui_activation_batch_paused(task.id, f"第{code_index + 1}个激活码结果未达成成功")
                    break
                if self.check_activation_license(task):
                    break

            if stopped or self._stopped():
                break
            if task_has_exception:
                break
            if not self.logout_by_menu():
                self.app.ui_activation_batch_paused(task.id, "当前账号登出未确认")
                stopped = True
                break
            self.app.ui_activation_finished(task.id)
            time.sleep(1)

        self.app.ui_activation_batch_finished(stopped or self._stopped())

    def _purchase_product(self, task: Task) -> dict:
        product = resolve_game_product(task.game_name)
        if product is None:
            raise RuntimeError(f"未配置游戏产品标识：{task.game_name}")
        expected_price = float(task.game_price or 0)
        if expected_price and abs(expected_price - product["price"]) > 0.01:
            raise RuntimeError(
                f"任务价格¥{expected_price:.2f}与Steam标准版配置"
                f"¥{product['price']:.2f}不一致"
            )
        return product

    def _verify_live_product_catalog(self, product: dict) -> dict:
        api_url = (
            "https://store.steampowered.com/api/appdetails"
            f"?appids={product['app_id']}&cc=CN&l=schinese"
        )
        response = self.context.request.get(api_url, timeout=20000)
        if not response.ok:
            raise RuntimeError(f"Steam产品目录请求失败：HTTP {response.status}")
        payload = response.json().get(str(product["app_id"]), {})
        if not payload.get("success"):
            raise RuntimeError("Steam产品目录未返回有效产品")
        data = payload.get("data") or {}
        live_name = str(data.get("name") or "").strip()
        live_price = float((data.get("price_overview") or {}).get("final", 0)) / 100
        package_options = [
            sub
            for group in data.get("package_groups") or []
            for sub in group.get("subs") or []
        ]
        standard = next(
            (
                sub
                for sub in package_options
                if int(sub.get("packageid", 0)) == product["package_id"]
            ),
            None,
        )
        if normalize_game_name(live_name) != normalize_game_name(product["official_name"]):
            raise RuntimeError(
                f"Steam正式名称变化：预期“{product['official_name']}”，实际“{live_name}”"
            )
        if abs(live_price - product["price"]) > 0.01:
            raise RuntimeError(
                f"Steam当前价格¥{live_price:.2f}与任务价格¥{product['price']:.2f}不一致"
            )
        if standard is None:
            raise RuntimeError(
                f"Steam目录未找到标准版Package {product['package_id']}"
            )
        standard_price = float(standard.get("price_in_cents_with_discount", 0)) / 100
        if abs(standard_price - product["price"]) > 0.01:
            raise RuntimeError(
                f"标准版Package价格¥{standard_price:.2f}与任务价格不一致"
            )
        return {
            "name": live_name,
            "price": live_price,
            "package_id": int(standard["packageid"]),
            "option_text": standard.get("option_text") or "",
        }

    def _steam_session_id(self) -> str:
        for cookie in self.context.cookies("https://store.steampowered.com"):
            if cookie.get("name") == "sessionid":
                return str(cookie.get("value") or "")
        try:
            return str(self.page.evaluate("() => window.g_sessionID || ''") or "")
        except Exception:
            return ""

    def start_purchase_account(self, task: Task) -> bool:
        if not self.ensure_logged_out_before_new_account():
            return False
        self._goto(STEAM_LOGIN_URL)
        if not self._fill_login_credentials(task.steam_account, task.steam_password):
            return False
        self._submit_login()
        if not self._wait_logged_in():
            return False
        if not self._verify_task_identity(task):
            self.app.ui_purchase_issue(task.id, "账号省市/店名核心身份核对失败")
            return False
        product = self._purchase_product(task)
        self.app.ui_purchase_step(
            task.id,
            "登录完成",
            f"目标：{product['official_name']}；App {product['app_id']}；"
            f"标准版Package {product['package_id']}",
        )
        return True

    def clear_purchase_cart(self, task: Task, require_confirmation: bool = False) -> bool:
        self._purchase_product(task)
        self._goto(STEAM_CART_URL)
        time.sleep(2)
        removed = 0

        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        cart_count_match = re.search(
            r"(?:购物车|Cart)\s*[\(\[]?\s*(\d+)",
            body_text,
            flags=re.IGNORECASE,
        )
        initial_count = int(cart_count_match.group(1)) if cart_count_match else 0
        if initial_count > 0 and require_confirmation:
            if not self.app.request_cart_clear_confirmation(task, initial_count):
                self.app.ui_purchase_issue(
                    task.id,
                    f"购物车已有{initial_count}项内容，用户未确认删除",
                )
                return False

        remove_all = self._first_visible(
            [
                "text=\"移除所有项目\"",
                "text=\"Remove all items\"",
            ],
            timeout_ms=2500,
        )
        if remove_all is not None:
            remove_all.click()
            removed = initial_count
            time.sleep(1)
            confirm = self._first_visible(
                [
                    "button:has-text('确定')",
                    "button:has-text('确认')",
                    "button:has-text('移除')",
                    "button:has-text('Remove')",
                ],
                timeout_ms=1200,
            )
            if confirm is not None:
                confirm.click()
                time.sleep(1)

        remove_selectors = [
            "[data-featuretarget='removeitem']",
            "[data-featuretarget='remove-item']",
            "button[aria-label*='移除']",
            "button:has-text('移除')",
            "a:has-text('移除')",
            "text=\"移除\"",
            "button:has-text('Remove')",
            "a:has-text('Remove')",
            "text=\"Remove\"",
            ".remove_link",
        ]
        while removed < 30 and not self._stopped():
            remove_button = self._first_visible(remove_selectors, timeout_ms=1800)
            if remove_button is None:
                break
            remove_button.click()
            removed += 1
            time.sleep(1)

        self._goto(STEAM_CART_URL)
        time.sleep(2)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        cart_count_match = re.search(
            r"(?:购物车|Cart)\s*[\(\[]?\s*(\d+)",
            body_text,
            flags=re.IGNORECASE,
        )
        remaining_count = int(cart_count_match.group(1)) if cart_count_match else 0
        empty_markers = (
            "您的购物车是空的",
            "您的购物车为空",
            "购物车中没有商品",
            "Your cart is empty",
        )
        cart_is_empty = remaining_count == 0 and (
            any(marker.lower() in body_text.lower() for marker in empty_markers)
            or self._first_visible(remove_selectors, timeout_ms=1000) is None
        )
        if not cart_is_empty:
            self.app.ui_purchase_issue(
                task.id,
                f"购物车仍有{remaining_count or '未知数量'}项内容，未继续后续购买步骤",
            )
            return False
        path = self._capture_page(task, "游戏购买-购物车已清空")
        self.app.ui_purchase_step(
            task.id,
            "购物车清空",
            f"已移除{max(removed, initial_count)}项购物车内容，当前计数为0",
            path,
        )
        return True

    def check_purchase_license(self, task: Task) -> bool:
        product = self._purchase_product(task)
        self._goto(STEAM_LICENSES_URL)
        time.sleep(2)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        owned = product_in_text(product, body_text)
        if owned:
            self.app.ui_purchase_already_deployed(
                task.id,
                f"资产列表已存在《{product['official_name']}》，已标记为完成部署",
            )
            return False
        self.app.ui_purchase_step(
            task.id,
            "资产检查",
            f"许可证与产品激活列表中未找到《{product['official_name']}》",
        )
        return True

    def open_standard_game_page(self, task: Task) -> bool:
        product = self._purchase_product(task)
        live = None
        catalog_error = ""
        try:
            live = self._verify_live_product_catalog(product)
        except Exception as exc:
            catalog_error = str(exc)
        self._goto(
            f"https://store.steampowered.com/app/{product['app_id']}/"
            "?cc=CN&l=schinese"
        )
        time.sleep(2)
        if live is None:
            body_text = (
                self.page.locator("body").inner_text(timeout=8000) or ""
            ).strip()
            page_html = self.page.content()
            normalized_body = normalize_game_name(body_text)
            if normalize_game_name(product["official_name"]) not in normalized_body:
                self.app.ui_purchase_issue(
                    task.id,
                    f"商店页未识别到《{product['official_name']}》；"
                    f"目录接口错误：{catalog_error}",
                )
                return False
            price_text = body_text.replace(",", "").replace(" ", "")
            if f"{product['price']:.2f}" not in price_text:
                self.app.ui_purchase_issue(
                    task.id,
                    f"商店页未识别到标准版价格¥{product['price']:.2f}；"
                    f"目录接口错误：{catalog_error}",
                )
                return False
            package_id = str(product["package_id"])
            if package_id not in page_html:
                self.app.ui_purchase_issue(
                    task.id,
                    f"商店页未识别到标准版Package {package_id}；"
                    f"目录接口错误：{catalog_error}",
                )
                return False
            verification_message = (
                f"目录接口暂时不可用，已改用商店页现场核对："
                f"{product['official_name']}；App {product['app_id']}；"
                f"标准版Package {product['package_id']}；¥{product['price']:.2f}"
            )
        else:
            verification_message = (
                f"官方目录已核对：{live['name']}；App {product['app_id']}；"
                f"标准版Package {live['package_id']}；¥{live['price']:.2f}"
            )
        self.app.ui_purchase_step(
            task.id,
            "标准版商店页",
            verification_message,
        )
        return True

    def add_standard_game_to_cart(self, task: Task) -> bool:
        product = self._purchase_product(task)
        package_id = product["package_id"]
        try:
            session_id = self._steam_session_id()
            if not session_id:
                raise RuntimeError("未取得Steam会话ID")
            self._goto(STEAM_CART_URL)
            self.page.evaluate(
                """({ sessionId, packageId }) => {
                    const form = document.createElement("form");
                    form.method = "POST";
                    form.action = "https://store.steampowered.com/cart/";
                    const fields = {
                        snr: "1_5_9__403",
                        originating_snr: "",
                        action: "add_to_cart",
                        sessionid: sessionId,
                        subid: String(packageId),
                    };
                    for (const [name, value] of Object.entries(fields)) {
                        const input = document.createElement("input");
                        input.type = "hidden";
                        input.name = name;
                        input.value = value;
                        form.appendChild(input);
                    }
                    document.body.appendChild(form);
                    form.submit();
                }""",
                {"sessionId": session_id, "packageId": package_id},
            )
            self.page.wait_for_load_state("domcontentloaded", timeout=20000)
            self.app.ui_purchase_step(
                task.id,
                "标准版提交",
                f"已在Steam页面内直接提交标准版Package {package_id}，"
                "无需经过商店详情页",
            )
        except Exception as exc:
            self.app.ui_purchase_issue(
                task.id,
                f"标准版Package {package_id}加入购物车失败：{exc}",
            )
            return False
        try:
            self.page.wait_for_url("**/cart/**", timeout=15000)
        except Exception:
            self._goto(STEAM_CART_URL)
        time.sleep(2)
        return self.verify_purchase_cart(task)

    def verify_purchase_cart(self, task: Task) -> bool:
        product = self._purchase_product(task)
        self._goto(STEAM_CART_URL)
        time.sleep(2)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        if normalize_game_name(product["official_name"]) not in normalize_game_name(body_text):
            self.app.ui_purchase_issue(
                task.id,
                f"购物车中未找到《{product['official_name']}》",
            )
            return False
        if f"{product['price']:.2f}" not in body_text.replace(",", ""):
            self.app.ui_purchase_issue(
                task.id,
                f"购物车价格与标准版¥{product['price']:.2f}不一致",
            )
            return False
        remove_candidates = self.page.locator(
            "[data-featuretarget='removeitem'], "
            "button[aria-label*='移除'], .remove_link"
        )
        visible_remove_count = 0
        for index in range(min(remove_candidates.count(), 30)):
            try:
                if remove_candidates.nth(index).is_visible(timeout=200):
                    visible_remove_count += 1
            except Exception:
                continue
        header_count_match = re.search(
            r"(?:购物车|Cart)\s*[\(\[]?\s*(\d+)",
            body_text,
            flags=re.IGNORECASE,
        )
        header_count = int(header_count_match.group(1)) if header_count_match else 0
        if header_count > 1:
            self.app.ui_purchase_issue(
                task.id,
                f"购物车标题显示{header_count}项内容，必须先清空后重试",
            )
            return False
        if visible_remove_count > 1:
            self.app.ui_purchase_issue(
                task.id,
                f"购物车检测到{visible_remove_count}项内容，必须先清空后重试",
            )
            return False
        self.app.ui_purchase_step(
            task.id,
            "购物车核对",
            f"购物车仅保留《{product['official_name']}》标准版，"
            f"价格¥{product['price']:.2f}",
        )
        return True

    def prepare_purchase_checkout(self, task: Task) -> bool:
        product = self._purchase_product(task)
        if not self.verify_purchase_cart(task):
            return False
        clicked_checkout = self._click(
            [
                "#checkout_btn",
                "#btn_checkout",
                "[id*='checkout']:has-text('跳转至支付')",
                "[class*='checkout']:has-text('跳转至支付')",
                "text=\"跳转至支付\"",
                "button:has-text('跳转至支付')",
                "a:has-text('跳转至支付')",
                "[id*='checkout']:has-text('Proceed to checkout')",
                "[class*='checkout']:has-text('Proceed to checkout')",
                "text=\"Proceed to checkout\"",
                "button:has-text('Proceed to checkout')",
                "a:has-text('Proceed to checkout')",
                "#btn_purchase_self",
                "button:has-text('为自己购买')",
                "a:has-text('为自己购买')",
                "button:has-text('Purchase for myself')",
                "a:has-text('Purchase for myself')",
            ],
            "跳转至支付",
        )
        if not clicked_checkout:
            try:
                clicked_checkout = bool(
                    self.page.evaluate(
                        """() => {
                            const labels = [
                                "跳转至支付",
                                "Proceed to checkout",
                                "为自己购买",
                                "Purchase for myself",
                            ];
                            const isVisible = (el) => {
                                const style = window.getComputedStyle(el);
                                const rect = el.getBoundingClientRect();
                                return style
                                    && style.display !== "none"
                                    && style.visibility !== "hidden"
                                    && rect.width > 20
                                    && rect.height > 10;
                            };
                            const nodes = Array.from(
                                document.querySelectorAll("button,a,div,span")
                            );
                            const candidates = nodes
                                .filter((el) => isVisible(el))
                                .filter((el) => {
                                    const text = (el.innerText || el.textContent || "")
                                        .replace(/\\s+/g, " ")
                                        .trim();
                                    return labels.some((label) => text === label || text.includes(label));
                                })
                                .sort((a, b) => {
                                    const ar = a.getBoundingClientRect();
                                    const br = b.getBoundingClientRect();
                                    return (ar.width * ar.height) - (br.width * br.height);
                                });
                            if (!candidates.length) {
                                return false;
                            }
                            candidates[0].click();
                            return true;
                        }"""
                    )
                )
                if clicked_checkout:
                    self._log("已通过页面文本点击跳转至支付")
            except Exception as exc:
                self._log(f"页面文本点击跳转至支付失败：{exc}")
        if not clicked_checkout:
            self.app.ui_purchase_issue(task.id, "未找到“跳转至支付”按钮")
            return False
        time.sleep(3)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        if not any(
            marker.lower() in body_text.lower()
            for marker in ["Steam 钱包", "Steam钱包", "Steam Wallet"]
        ):
            self.app.ui_purchase_issue(
                task.id,
                "结算页未确认使用Steam钱包，请人工检查付款方式",
            )
            return False
        self.app.ui_purchase_step(
            task.id,
            "结算准备",
            f"已进入《{product['official_name']}》结算页并检测到Steam钱包；"
            "尚未提交购买",
        )
        return True

    def confirm_purchase_checkout(self, task: Task) -> bool:
        product = self._purchase_product(task)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        if not any(
            marker.lower() in body_text.lower()
            for marker in ["Steam 钱包", "Steam钱包", "Steam Wallet"]
        ):
            self.app.ui_purchase_issue(
                task.id,
                "当前页面未确认Steam钱包付款，已阻止提交",
            )
            return False
        checkbox = self._first_visible(
            [
                "#accept_ssa",
                "input[type='checkbox'][name*='agree']",
                "input[type='checkbox']",
            ],
            timeout_ms=2500,
        )
        if checkbox is not None:
            try:
                if not checkbox.is_checked():
                    checkbox.check()
                    self._log("已勾选Steam订户协议确认")
            except Exception:
                pass
        if not self._click(
            [
                "#purchase_button",
                "button:has-text('购买')",
                "a:has-text('购买')",
                "button:has-text('Purchase')",
            ],
            "最终购买按钮",
        ):
            self.app.ui_purchase_issue(task.id, "未找到最终购买按钮")
            return False
        deadline = time.time() + 40
        success = False
        while time.time() < deadline and not self._stopped():
            time.sleep(1)
            result_text = (
                self.page.locator("body").inner_text(timeout=3000) or ""
            ).strip()
            if any(
                marker.lower() in result_text.lower()
                for marker in [
                    "感谢您的购买",
                    "谢谢惠顾",
                    "购买完成",
                    "购买成功",
                    "您的购物收据",
                    "确认代码",
                    "Thank you for your purchase",
                    "Purchase complete",
                ]
            ):
                success = True
                break
        if success:
            self.app.ui_purchase_success(
                task.id,
                f"《{product['official_name']}》购买成功",
            )
            return True
        else:
            self.app.ui_purchase_issue(
                task.id,
                "提交后未识别到明确成功文本，请检查保留现场",
            )
            return False

    def verify_purchase_asset_and_logout(self, task: Task) -> bool:
        product = self._purchase_product(task)
        self._goto(STEAM_LICENSES_URL)
        time.sleep(3)
        body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        owned = product_in_text(product, body_text)
        if not owned:
            self.app.ui_purchase_issue(
                task.id,
                f"购买后资产清单中未找到《{product['official_name']}》",
            )
            return False
        path = self._capture_page(task, "游戏购买-部署完成-资产已确认")
        self.app.ui_purchase_step(
            task.id,
            "部署确认",
            f"资产清单已确认存在《{product['official_name']}》",
            path,
        )
        if not self.logout_by_menu():
            self.app.ui_purchase_issue(task.id, "购买后登出未确认")
            return False
        self.app.ui_purchase_finished(task.id)
        return True

    def start_login_only(self, task: Task) -> None:
        if not self.ensure_logged_out_before_new_account():
            return
        self._goto(STEAM_LOGIN_URL)
        if not self._fill_login_credentials(task.steam_account, task.steam_password):
            return
        self._submit_login()
        self._wait_logged_in()

    @staticmethod
    def _normalize_friend_invite_link(raw_link: str) -> str:
        text = str(raw_link or "").strip()
        match = re.search(
            r"(?:https?://)?s\.team/p/[A-Za-z0-9]{2,8}-[A-Za-z0-9]{2,8}/[A-Za-z0-9]{4,32}",
            text,
            re.IGNORECASE,
        )
        if not match:
            return ""
        link = match.group(0).rstrip(").,;，；]}>\"'")
        if link.lower().startswith("s.team/"):
            link = f"https://{link}"
        return link

    def _friend_invite_links(self) -> list[str]:
        self._ensure_page()
        try:
            raw_links = self.page.evaluate(
                """
                () => {
                    const found = new Set();
                    const pattern = /(?:https?:\\/\\/)?s\\.team\\/p\\/[A-Za-z0-9]{2,8}-[A-Za-z0-9]{2,8}\\/[A-Za-z0-9]{4,32}/ig;
                    const scan = (value) => {
                        if (!value) return;
                        for (const match of String(value).matchAll(pattern)) {
                            found.add(match[0]);
                        }
                    };
                    scan(document.body ? document.body.innerText : "");
                    scan(document.body ? document.body.textContent : "");
                    for (const element of document.querySelectorAll('*')) {
                        scan(element.textContent);
                        scan(element.getAttribute('href'));
                        scan(element.getAttribute('value'));
                        scan(element.getAttribute('data-clipboard-text'));
                        scan(element.getAttribute('data-copy-text'));
                        scan(element.getAttribute('data-tooltip-text'));
                        scan(element.getAttribute('title'));
                        scan(element.getAttribute('aria-label'));
                        if ('value' in element) scan(element.value);
                        if ('href' in element) scan(element.href);
                    }
                    for (const script of document.querySelectorAll('script')) {
                        scan(script.textContent);
                    }
                    return Array.from(found);
                }
                """
            )
        except Exception as exc:
            self._log(f"读取好友邀请链接失败：{exc}")
            return []

        links: list[str] = []
        for raw_link in raw_links or []:
            link = self._normalize_friend_invite_link(str(raw_link))
            if not link:
                continue
            if link not in links:
                links.append(link)
        return links

    def _click_generate_friend_invite_control(self) -> bool:
        self._ensure_page()
        selectors = [
            "#generate_new_link",
            "#generate_new_invite",
            ".generate_new_link",
            "[onclick*='GenerateNewFriendInviteLink']",
            "[onclick*='GenerateFriendInviteLink']",
            "[onclick*='CreateFriendInviteLink']",
            "[onclick*='RegenerateFriendInviteLink']",
            "[onclick*='InviteLink']",
            "button:has-text('生成新链接')",
            "a:has-text('生成新链接')",
            "[role='button']:has-text('生成新链接')",
            "button:has-text('生成新的链接')",
            "a:has-text('生成新的链接')",
            "[role='button']:has-text('生成新的链接')",
            "button:has-text('生成新的邀请链接')",
            "a:has-text('生成新的邀请链接')",
            "[role='button']:has-text('生成新的邀请链接')",
            "button:has-text('重新生成')",
            "a:has-text('重新生成')",
            "[role='button']:has-text('重新生成')",
            "button:has-text('Generate a new link')",
            "a:has-text('Generate a new link')",
            "[role='button']:has-text('Generate a new link')",
            "button:has-text('Generate new link')",
            "a:has-text('Generate new link')",
            "[role='button']:has-text('Generate new link')",
        ]
        if self._click(selectors, "生成新的好友邀请链接"):
            return True
        try:
            clicked = self.page.evaluate(
                """
                () => {
                    const isVisible = (element) => {
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return style.visibility !== 'hidden'
                            && style.display !== 'none'
                            && rect.width > 0
                            && rect.height > 0;
                    };
                    const pattern = /GenerateNewFriendInviteLink|GenerateFriendInviteLink|CreateFriendInviteLink|RegenerateFriendInviteLink|generate.*(?:invite|link)|new.*(?:invite|link)|生成.*(?:邀请|链接)|重新生成/i;
                    const candidates = Array.from(document.querySelectorAll('button, a, input, [role="button"], span, div'))
                        .filter((element) => {
                            const values = [
                                element.innerText,
                                element.textContent,
                                element.value,
                                element.id,
                                element.className,
                                element.getAttribute('onclick'),
                                element.getAttribute('title'),
                                element.getAttribute('aria-label'),
                                element.getAttribute('data-tooltip-text'),
                            ].filter(Boolean).join(' ');
                            return pattern.test(values);
                        })
                        .map((element) => element.closest('button, a, input, [role="button"]') || element)
                        .filter((element, index, all) => all.indexOf(element) === index);
                    const target = candidates.find(isVisible) || candidates[0];
                    if (!target) return false;
                    target.click();
                    return true;
                }
                """
            )
            if clicked:
                self._log("已点击页面中的生成好友链接控件")
                return True
        except Exception as exc:
            self._log(f"点击生成好友链接控件失败：{exc}")
        return False

    def _generate_friend_invite_link_via_page_api(self) -> str:
        self._ensure_page()
        try:
            result = self.page.evaluate(
                """
                async () => {
                    const collectLinks = (value) => {
                        const links = [];
                        const pattern = /(?:https?:\\/\\/)?s\\.team\\/p\\/[A-Za-z0-9]{2,8}-[A-Za-z0-9]{2,8}\\/[A-Za-z0-9]{4,32}/ig;
                        if (!value) return links;
                        for (const match of String(value).matchAll(pattern)) {
                            links.push(match[0]);
                        }
                        return links;
                    };
                    const functionNames = [
                        'GenerateNewFriendInviteLink',
                        'GenerateFriendInviteLink',
                        'CreateFriendInviteLink',
                        'RegenerateFriendInviteLink',
                        'GenerateInviteLink',
                    ];
                    for (const functionName of functionNames) {
                        if (typeof window[functionName] === 'function') {
                            try {
                                const value = window[functionName]();
                                if (value && typeof value.then === 'function') {
                                    await value;
                                }
                                await new Promise((resolve) => setTimeout(resolve, 1200));
                                const links = collectLinks(document.body ? document.body.innerText : '')
                                    .concat(collectLinks(document.body ? document.body.textContent : ''));
                                if (links.length) {
                                    return { method: functionName, links };
                                }
                            } catch (error) {
                                // Try the next known Steam page function.
                            }
                        }
                    }
                    const cookieSession = (document.cookie.match(/(?:^|;\\s*)sessionid=([^;]+)/) || [])[1] || '';
                    const sessionId = window.g_sessionID || decodeURIComponent(cookieSession);
                    if (!sessionId) {
                        return { method: 'api', error: 'missing sessionid', links: [] };
                    }
                    const endpoints = [
                        'https://steamcommunity.com/actions/GenerateFriendInviteLink/',
                        'https://steamcommunity.com/actions/GenerateFriendInviteLink',
                        'https://steamcommunity.com/actions/GenerateNewFriendInviteLink/',
                        'https://steamcommunity.com/actions/GenerateNewFriendInviteLink',
                    ];
                    let lastResult = { method: 'api', error: 'not attempted', links: [] };
                    for (const endpoint of endpoints) {
                        const body = new URLSearchParams();
                        body.set('sessionid', sessionId);
                        body.set('sessionID', sessionId);
                        try {
                            const response = await fetch(endpoint, {
                                method: 'POST',
                                credentials: 'include',
                                headers: {
                                    'Accept': 'application/json, text/javascript, */*; q=0.01',
                                    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                                    'X-Requested-With': 'XMLHttpRequest',
                                },
                                body: body.toString(),
                            });
                            const text = await response.text();
                            const links = collectLinks(text);
                            lastResult = {
                                method: endpoint,
                                ok: response.ok,
                                status: response.status,
                                links,
                                text: text.slice(0, 500),
                            };
                            if (links.length) return lastResult;
                        } catch (error) {
                            lastResult = { method: endpoint, error: String(error), links: [] };
                        }
                    }
                    return lastResult;
                }
                """
            )
        except Exception as exc:
            self._log(f"调用Steam生成好友链接接口失败：{exc}")
            return ""

        links = result.get("links") if isinstance(result, dict) else []
        for raw_link in links or []:
            link = self._normalize_friend_invite_link(str(raw_link))
            if link:
                self._log("已通过Steam页面接口生成新的好友邀请链接")
                return link
        if isinstance(result, dict):
            self._log(
                "Steam生成好友链接接口未返回可识别链接"
                f"（method={result.get('method')} status={result.get('status')} error={result.get('error', '')}）"
            )
        return ""

    def _wait_for_new_friend_invite_link(
        self,
        existing_links: set[str],
        timeout_seconds: int = 45,
    ) -> str:
        deadline = time.time() + timeout_seconds
        while time.time() < deadline and not self._stopped():
            for link in self._friend_invite_links():
                if link not in existing_links:
                    return link
            time.sleep(0.5)
        return ""

    def _generate_new_friend_invite_link(self, existing_links: set[str]) -> str:
        if self._click_generate_friend_invite_control():
            link = self._wait_for_new_friend_invite_link(existing_links, timeout_seconds=20)
            if link:
                return link
        link = self._generate_friend_invite_link_via_page_api()
        if link and link not in existing_links:
            return link
        if link:
            existing_links.add(link)
        self.page.reload(wait_until="commit", timeout=45000)
        return self._wait_for_new_friend_invite_link(existing_links, timeout_seconds=20)

    def collect_friend_invite_links(self, task: Task) -> bool:
        if not self._goto_friend_claim_page(STEAM_FRIENDS_ADD_URL, "好友添加页"):
            self.app.ui_friend_claim_issue(task.id, "好友添加页无法打开，请检查网络后重试")
            return False
        friend_link_1 = self._wait_for_new_friend_invite_link(set())
        if not friend_link_1:
            self.app.ui_friend_claim_issue(
                task.id,
                "未读取到Steam快速邀请链接，请检查好友添加页面是否已加载",
            )
            return False

        # Store the first link before requesting a replacement link.
        self.app.ui_friend_claim_link_saved(task.id, 1, friend_link_1)
        friend_link_2 = self._generate_new_friend_invite_link({friend_link_1})
        if not friend_link_2:
            self.app.ui_friend_claim_issue(
                task.id,
                "生成新链接后未读取到第二条不同的好友邀请链接，已保留当前页面供人工处理",
            )
            return False

        self.app.ui_friend_claim_links_collected(task.id, friend_link_1, friend_link_2)
        self._log("已采集两条不同的Steam好友邀请链接")
        return True

    def _prepare_friend_claim_login(self) -> bool:
        """Go straight to login, logging out an existing Steam session only when needed."""
        if not self._goto_friend_claim_page(STEAM_LOGIN_URL, "Steam登录页"):
            return False
        session_marker = self._first_visible(
            ["#account_pulldown", "a[href*='logout']", "input[type='password']"],
            timeout_ms=15000,
        )
        if session_marker is not None and self._is_logged_in():
            self._log("检测到已有登录账号，先通过右上角菜单退出")
            if not self.logout_by_menu():
                return False
            return self._goto_friend_claim_page(STEAM_LOGIN_URL, "Steam登录页")
        return True

    def _login_friend_claim_task(self, task: Task) -> bool:
        if not self._prepare_friend_claim_login():
            self.app.ui_friend_claim_issue(task.id, "开始新账号前无法确认登出或登录页无法打开")
            return False
        if not self._fill_login_credentials(task.steam_account, task.steam_password):
            self.app.ui_friend_claim_issue(task.id, "登录表单填写失败")
            return False
        self._submit_login()
        if not self._wait_logged_in():
            self.app.ui_friend_claim_issue(task.id, "登录结果未确认")
            return False
        if not self._verify_task_identity(task):
            self.app.ui_friend_claim_issue(task.id, "账号省市/店名核心身份核对失败")
            return False
        return True

    def start_friend_claim_account(self, task: Task) -> bool:
        if not self._login_friend_claim_task(task):
            return False
        return self.collect_friend_invite_links(task)

    def _friend_claim_product(self, task: Task) -> dict | None:
        product = resolve_game_product(task.game_name)
        if product is None:
            self.app.ui_friend_game_claim_issue(task.id, f"未配置可核对的游戏产品标识：{task.game_name}")
        return product

    def _verify_friend_game_asset(self, task: Task, status_text: str) -> Path | None:
        product = self._friend_claim_product(task)
        if product is None:
            return None
        self._goto(STEAM_LICENSES_URL)
        if self._stopped():
            return None
        time.sleep(2)
        body_text = (self.page.locator("body").inner_text(timeout=7000) or "").strip()
        if not product_in_text(product, body_text):
            self._log(f"资产清单中未找到《{product['official_name']}》")
            return None
        return self._capture_page(task, status_text)

    def _click_notification_bell(self) -> bool:
        coordinate_selectors = [
            "[data-featuretarget='green-envelope']",
            "#header_notification_link",
            "#header_notification_area",
            "#header_notification_count",
            ".header_notification_btn",
        ]
        for selector in coordinate_selectors:
            try:
                locator = self.page.locator(selector).first
                if locator.count() <= 0 or not locator.is_visible(timeout=500):
                    continue
                box = locator.bounding_box()
                if not box:
                    continue
                x = box["x"] + box["width"] / 2
                y = box["y"] + box["height"] / 2
                self.page.mouse.move(x, y)
                time.sleep(0.15)
                self.page.mouse.down()
                time.sleep(0.08)
                self.page.mouse.up()
                self._log(f"已点击通知小铃铛坐标：{selector}")
                time.sleep(1.2)
                return True
            except Exception as exc:
                self._log(f"坐标点击通知小铃铛失败（{selector}）：{exc}")

        try:
            clicked = self.page.evaluate(
                """
                () => {
                    const visible = (element) => {
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return style.display !== 'none'
                            && style.visibility !== 'hidden'
                            && rect.width > 0
                            && rect.height > 0;
                    };
                    const candidates = [
                        document.querySelector('[data-featuretarget="green-envelope"]'),
                        document.querySelector('#header_notification_link'),
                        document.querySelector('#header_notification_area'),
                        document.querySelector('#header_notification_count'),
                        ...Array.from(document.querySelectorAll('a, button, [role="button"], div, span')).filter((element) => {
                            const values = [
                                element.id,
                                element.className,
                                element.href,
                                element.getAttribute('aria-label'),
                                element.getAttribute('title'),
                                element.getAttribute('data-tooltip-text'),
                                element.innerText,
                            ].filter(Boolean).join(' ');
                            return /notification|通知|bell|铃铛|green-envelope/i.test(values);
                        }),
                    ].filter(Boolean);
                    const target = candidates
                        .map((element) => element.closest('a, button, [role="button"]') || element)
                        .find(visible);
                    if (!target) return false;
                    target.click();
                    return true;
                }
                """
            )
            if clicked:
                self._log("已通过页面脚本点击通知小铃铛")
                return True
        except Exception as exc:
            self._log(f"脚本点击通知小铃铛失败：{exc}")

        clicked = self._click(
            [
                "#header_notification_link",
                "#header_notification_area",
                "#header_notification_count",
                ".header_notification_btn",
                "a[href*='/notifications']",
                "a[href*='notifications']",
                "button[aria-label*='通知']",
                "button[aria-label*='Notification']",
                "[aria-label*='通知']",
                "[aria-label*='Notification']",
                "[class*='notification'][class*='bell']",
                "[class*='NotificationBell']",
                "[class*='notification']",
                "[class*='NotificationBell']",
            ],
            "通知小铃铛",
        )
        if clicked:
            return True
        return False

    def _click_visible_text_candidate(
        self,
        pattern: str,
        label: str,
        exclude_pattern: str = "",
        container_selector: str = "",
    ) -> bool:
        self._ensure_page()
        try:
            result = self.page.evaluate(
                """
                ({ pattern, excludePattern, containerSelector }) => {
                    const re = new RegExp(pattern, 'i');
                    const exclude = excludePattern ? new RegExp(excludePattern, 'i') : null;
                    const root = containerSelector
                        ? (document.querySelector(containerSelector) || document)
                        : document;
                    const visible = (element) => {
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return style.display !== 'none'
                            && style.visibility !== 'hidden'
                            && rect.width > 0
                            && rect.height > 0;
                    };
                    const textOf = (element) => [
                        element.innerText,
                        element.textContent,
                        element.value,
                        element.getAttribute('title'),
                        element.getAttribute('aria-label'),
                    ].filter(Boolean).join(' ').replace(/\\s+/g, ' ').trim();
                    const candidates = Array.from(root.querySelectorAll('a, button, input, [role="button"], div, span'))
                        .map((element) => element.closest('a, button, input, [role="button"]') || element)
                        .filter((element, index, all) => all.indexOf(element) === index)
                        .filter((element) => visible(element))
                        .map((element) => ({ element, text: textOf(element) }))
                        .filter((item) => item.text && re.test(item.text) && !(exclude && exclude.test(item.text)));
                    const target = candidates
                        .sort((a, b) => a.text.length - b.text.length)[0];
                    if (!target) return null;
                    target.element.scrollIntoView({ block: 'center', inline: 'center' });
                    const rect = target.element.getBoundingClientRect();
                    return {
                        x: rect.left + rect.width / 2,
                        y: rect.top + rect.height / 2,
                        text: target.text,
                    };
                }
                """,
                {
                    "pattern": pattern,
                    "excludePattern": exclude_pattern,
                    "containerSelector": container_selector,
                },
            )
        except Exception as exc:
            self._log(f"查找{label}失败：{exc}")
            return False
        if not result:
            self._log(f"没有找到{label}")
            return False
        try:
            self.page.mouse.click(float(result["x"]), float(result["y"]))
        except Exception:
            try:
                self.page.evaluate(
                    """
                    ({ pattern, excludePattern, containerSelector }) => {
                        const re = new RegExp(pattern, 'i');
                        const exclude = excludePattern ? new RegExp(excludePattern, 'i') : null;
                        const root = containerSelector
                            ? (document.querySelector(containerSelector) || document)
                            : document;
                        const visible = (element) => {
                            const style = window.getComputedStyle(element);
                            const rect = element.getBoundingClientRect();
                            return style.display !== 'none'
                                && style.visibility !== 'hidden'
                                && rect.width > 0
                                && rect.height > 0;
                        };
                        const textOf = (element) => [
                            element.innerText,
                            element.textContent,
                            element.value,
                            element.getAttribute('title'),
                            element.getAttribute('aria-label'),
                        ].filter(Boolean).join(' ').replace(/\\s+/g, ' ').trim();
                        const target = Array.from(root.querySelectorAll('a, button, input, [role="button"], div, span'))
                            .map((element) => element.closest('a, button, input, [role="button"]') || element)
                            .filter((element, index, all) => all.indexOf(element) === index)
                            .filter((element) => visible(element))
                            .map((element) => ({ element, text: textOf(element) }))
                            .filter((item) => item.text && re.test(item.text) && !(exclude && exclude.test(item.text)))
                            .sort((a, b) => a.text.length - b.text.length)[0]?.element;
                        if (!target) return false;
                        target.click();
                        return true;
                    }
                    """,
                    {
                        "pattern": pattern,
                        "excludePattern": exclude_pattern,
                        "containerSelector": container_selector,
                    },
                )
            except Exception as exc:
                self._log(f"点击{label}失败：{exc}")
                return False
        self._log(f"已点击{label}：{result.get('text', '')[:80]}")
        return True

    def _click_gift_notification_candidate(self, product: dict) -> bool:
        self._ensure_page()
        try:
            result = self.page.evaluate(
                """
                () => {
                    const visible = (element) => {
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return style.display !== 'none'
                            && style.visibility !== 'hidden'
                            && rect.width > 0
                            && rect.height > 0;
                    };
                    const textOf = (element) => [
                        element.innerText,
                        element.textContent,
                        element.getAttribute('title'),
                        element.getAttribute('aria-label'),
                    ].filter(Boolean).join(' ').replace(/\\s+/g, ' ').trim();
                    const dropdowns = Array.from(document.querySelectorAll('body *'))
                        .filter(visible)
                        .map((element) => ({ element, text: textOf(element), rect: element.getBoundingClientRect() }))
                        .filter((item) => /通知|Notification/i.test(item.text)
                            && /新礼物|new\\s+gift/i.test(item.text)
                            && item.rect.width >= 240
                            && item.rect.height >= 80)
                        .sort((a, b) => (a.rect.width * a.rect.height) - (b.rect.width * b.rect.height));
                    const dropdown = dropdowns[0]?.element || document.querySelector('#header_notification_dropdown') || document;
                    const candidates = Array.from(dropdown.querySelectorAll('a, button, [role="button"], div, span'))
                        .map((element) => {
                            const clickable = element.closest('a, button, [role="button"]') || element;
                            return { element: clickable, text: textOf(clickable), rect: clickable.getBoundingClientRect() };
                        })
                        .filter((item, index, all) => all.findIndex((other) => other.element === item.element) === index)
                        .filter((item) => visible(item.element))
                        .filter((item) => /新\\s*礼物|new\\s+gift/i.test(item.text))
                        .filter((item) => !/礼物卡|库存|物品|商品|inventory|item|gift\\s*card/i.test(item.text));
                    const target = candidates
                        .sort((a, b) => {
                            const aRow = a.rect.width >= 200 && a.rect.height >= 25 ? 0 : 1;
                            const bRow = b.rect.width >= 200 && b.rect.height >= 25 ? 0 : 1;
                            if (aRow !== bRow) return aRow - bRow;
                            return a.text.length - b.text.length;
                        })[0];
                    if (target) {
                        target.element.scrollIntoView({ block: 'center', inline: 'center' });
                        const rect = target.element.getBoundingClientRect();
                        return {
                            found: true,
                            fallback: false,
                            x: rect.left + rect.width / 2,
                            y: rect.top + rect.height / 2,
                            text: target.text,
                        };
                    }
                    const visibleDropdown = dropdowns[0];
                    if (visibleDropdown) {
                        const rect = visibleDropdown.rect;
                        return {
                            found: true,
                            fallback: true,
                            x: rect.left + rect.width / 2,
                            y: rect.top + 78,
                            text: visibleDropdown.text.slice(0, 160),
                        };
                    }
                    return { found: false };
                }
                """
            )
        except Exception as exc:
            self._log(f"查找通知下拉中的新礼物失败：{exc}")
            return False
        clicked = bool(result and result.get("found"))
        if clicked:
            try:
                self.page.mouse.click(float(result["x"]), float(result["y"]))
                label = "通知下拉第一条新礼物" if result.get("fallback") else "通知下拉中的新礼物"
                self._log(f"已点击{label}：{str(result.get('text', ''))[:100]}")
            except Exception as exc:
                self._log(f"点击通知下拉中的新礼物失败：{exc}")
                return False
            try:
                self.page.wait_for_load_state("domcontentloaded", timeout=10000)
            except Exception:
                pass
            time.sleep(2)
        else:
            self._log("没有找到通知下拉中的新礼物")
        return clicked

    def _open_gift_notification_page(self, task: Task, product: dict) -> bool:
        candidate_urls = []
        profile_link = (task.profile_link or "").strip().rstrip("/")
        if profile_link:
            candidate_urls.append(f"{profile_link}/inventory/#pending_gifts")
        candidate_urls.append(STEAM_PENDING_GIFTS_URL)
        for url in candidate_urls:
            if self._stopped():
                return False
            if not self._goto_friend_claim_page(url, "待收礼物页面"):
                continue
            time.sleep(5)
            try:
                body_text = (self.page.locator("body").inner_text(timeout=8000) or "").strip()
            except Exception:
                body_text = ""
            if (
                "接收礼物" in body_text
                or "接受礼物" in body_text
                or "领取礼物" in body_text
                or "添加至我的库" in body_text
                or "添加到我的库" in body_text
                or "pending gift" in body_text.lower()
                or "accept gift" in body_text.lower()
            ):
                self._log(f"已打开待收礼物页面：{url}")
                return True
            self._log(f"待收礼物页面已打开但未识别到待处理礼物：{url}")
        return False

    def _current_gift_page_matches_product(self, product: dict) -> bool:
        try:
            body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
        except Exception:
            body_text = ""
        if product_in_text(product, body_text):
            return True
        lowered = body_text.lower()
        if "礼物" in body_text or "赠礼" in body_text or "gift" in lowered:
            self._log(f"当前礼物页面未识别到目标游戏《{product['official_name']}》，停止自动领取")
        return False

    def _click_positive_gift_claim_control(self) -> str:
        self._ensure_page()
        exact_steps: list[tuple[str, str, str]] = [
            (
                r"^\s*添加(至|到)我的库\s*$|^\s*添加(至|到)库\s*$|^\s*Add to my library\s*$",
                "添加至我的库按钮",
                "添加至我的库",
            ),
            (
                r"^\s*(接收|接受|领取|收下)礼物\s*$|^\s*Accept Gift\s*$|^\s*Redeem Gift\s*$",
                "接收礼物按钮",
                "接收礼物",
            ),
        ]
        for pattern, label, action in exact_steps:
            if self._click_visible_text_candidate(
                pattern,
                label,
                exclude_pattern=r"拒绝|拒收|取消|Decline|Reject|Cancel",
            ):
                return action
        selectors = [
            "button:has-text('添加到我的库')",
            "a:has-text('添加到我的库')",
            "button:has-text('添加至我的库')",
            "a:has-text('添加至我的库')",
            "button:has-text('添加到库')",
            "a:has-text('添加到库')",
            "button:has-text('Add to my library')",
            "a:has-text('Add to my library')",
            "button:has-text('接受礼物')",
            "a:has-text('接受礼物')",
            "button:has-text('接收礼物')",
            "a:has-text('接收礼物')",
            "button:has-text('领取礼物')",
            "a:has-text('领取礼物')",
            "button:has-text('收下礼物')",
            "a:has-text('收下礼物')",
            "button:has-text('Accept Gift')",
            "a:has-text('Accept Gift')",
            "button:has-text('Redeem Gift')",
            "a:has-text('Redeem Gift')",
            "input[value*='接受']",
            "input[value*='领取']",
            "[onclick*='AcceptGift']",
            "[onclick*='RedeemGift']",
        ]
        if self._click(selectors, "礼物领取确认按钮"):
            return "礼物领取确认"
        try:
            result = self.page.evaluate(
                """
                () => {
                    const addPattern = /添加.{0,8}(我的)?库|Add to.{0,12}Library/i;
                    const acceptPattern = /接受礼物|接收礼物|领取礼物|收下礼物|Accept Gift|Redeem Gift/i;
                    const negative = /拒绝|拒收|退回|取消|删除|Decline|Reject|Return|Cancel|Remove|Delete/i;
                    const visible = (element) => {
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return style.display !== 'none'
                            && style.visibility !== 'hidden'
                            && rect.width > 0
                            && rect.height > 0
                            && !element.disabled;
                    };
                    const textOf = (element) => [
                        element.innerText,
                        element.textContent,
                        element.value,
                        element.id,
                        element.className,
                        element.getAttribute('onclick'),
                        element.getAttribute('title'),
                        element.getAttribute('aria-label'),
                    ].filter(Boolean).join(' ');
                    const candidates = Array.from(document.querySelectorAll('button, a, input, [role="button"], span, div'))
                        .map((element) => element.closest('button, a, input, [role="button"]') || element)
                        .filter((element, index, all) => all.indexOf(element) === index)
                        .filter((element) => visible(element))
                        .map((element) => {
                            const text = textOf(element);
                            const action = addPattern.test(text) ? '添加至我的库' : (acceptPattern.test(text) ? '接收礼物' : '');
                            return { element, text, action };
                        });
                    const target = candidates
                        .filter((item) => item.action && !negative.test(item.text))
                        .sort((a, b) => {
                            const ap = a.action === '添加至我的库' ? 0 : 1;
                            const bp = b.action === '添加至我的库' ? 0 : 1;
                            return ap - bp;
                        })[0];
                    if (!target) return false;
                    target.element.scrollIntoView({ block: 'center', inline: 'center' });
                    target.element.click();
                    return { clicked: true, action: target.action };
                }
                """
            )
            if result and result.get("clicked"):
                action = str(result.get("action") or "礼物领取确认")
                self._log(f"已点击礼物领取确认控件：{action}")
                return action
        except Exception as exc:
            self._log(f"点击礼物领取控件失败：{exc}")
        return ""

    def _select_add_to_library_option(self) -> None:
        try:
            self.page.evaluate(
                """
                () => {
                    const pattern = /添加.{0,8}(我的)?库|add to.{0,12}library/i;
                    for (const label of Array.from(document.querySelectorAll('label'))) {
                        const text = [label.innerText, label.textContent].filter(Boolean).join(' ');
                        if (!pattern.test(text)) continue;
                        const target = label.control || label.querySelector('input');
                        if (target && !target.checked) target.click();
                        else label.click();
                        break;
                    }
                }
                """
            )
        except Exception:
            pass

    def _claim_gift_on_current_page(self, task: Task, product: dict) -> bool:
        clicked_any = False
        clicked_add_to_library = False
        success_markers = [
            "已添加到您的库",
            "已加入您的库",
            "已经在您的库",
            "添加到了您的 Steam 库",
            "added to your library",
            "already in your library",
            "successfully redeemed",
            "gift has been accepted",
            ]
        for _ in range(6):
            if self._stopped():
                return False
            try:
                body_text = (self.page.locator("body").inner_text(timeout=5000) or "").strip()
            except Exception:
                body_text = ""
            lowered = body_text.lower()
            if any(marker.lower() in lowered for marker in success_markers):
                self._log("页面显示礼物已领取")
                return True
            self._select_add_to_library_option()
            action = self._click_positive_gift_claim_control()
            if action:
                clicked_any = True
                if "库" in action or "library" in action.lower():
                    clicked_add_to_library = True
                try:
                    self.page.wait_for_load_state("domcontentloaded", timeout=12000)
                except Exception:
                    pass
                time.sleep(3)
                continue
            if clicked_add_to_library:
                return True
            if clicked_any:
                time.sleep(2)
                continue
            if "gift" in lowered or "礼物" in body_text or "赠礼" in body_text:
                self._log("礼物页面已打开，但未找到可点击的领取控件")
                return False
            time.sleep(1)
        return clicked_any

    def claim_friend_game_gift(self, task: Task) -> bool:
        product = self._friend_claim_product(task)
        if product is None:
            return False
        if not self._login_friend_claim_task(task):
            return False

        if not self._open_gift_notification_page(task, product):
            already_path = self._verify_friend_game_asset(task, "自动游戏领取-未找到通知但资产已存在")
            if already_path is not None:
                if not self.logout_by_menu():
                    self.app.ui_friend_game_claim_issue(task.id, "未找到通知但资产已存在，登出未确认")
                    return False
                self.app.ui_friend_game_claim_finished(
                    task.id,
                    f"未找到礼物通知，但资产清单已存在《{product['official_name']}》",
                    already_path,
                )
                return True
            self.app.ui_friend_game_claim_issue(task.id, "未找到礼物通知，且资产清单尚未出现目标游戏")
            return False
        self.app.ui_friend_game_claim_step(task.id, "通知", "已打开礼物通知或礼物领取页面")

        if not self._claim_gift_on_current_page(task, product):
            self.app.ui_friend_game_claim_issue(task.id, "礼物页面未能自动完成领取，请人工检查当前页面")
            return False

        path = self._verify_friend_game_asset(task, "自动游戏领取-资产已确认")
        if path is None:
            self.app.ui_friend_game_claim_issue(
                task.id,
                f"领取后资产清单中未找到《{product['official_name']}》",
            )
            return False
        if not self.logout_by_menu():
            self.app.ui_friend_game_claim_issue(task.id, "领取后登出未确认")
            return False
        self.app.ui_friend_game_claim_finished(
            task.id,
            f"已确认《{product['official_name']}》进入账号资产列表",
            path,
        )
        return True

    def full_auto_friend_game_claim(self, tasks: list[Task]) -> None:
        self.app.ui_friend_game_claim_batch_started(len(tasks))
        stopped = False
        for task_number, task in enumerate(tasks, start=1):
            if not self._wait_auto_gate():
                stopped = True
                break
            self.app.ui_friend_game_claim_batch_task_started(task, task_number, len(tasks))
            try:
                claimed = self.claim_friend_game_gift(task)
            except Exception as exc:
                self._log(f"自动游戏领取异常：{exc}")
                self.app.ui_friend_game_claim_issue(task.id, f"自动游戏领取异常：{exc}")
                claimed = False

            if self._stopped():
                stopped = True
                break

            if not claimed:
                try:
                    if self._is_logged_in():
                        self.logout_by_menu()
                except Exception as exc:
                    self._log(f"自动游戏领取异常后登出失败：{exc}")
                    stopped = True
                    break
            time.sleep(0.75)

        self.app.ui_friend_game_claim_batch_finished(stopped or self._stopped())

    def open_friend_claim_page(self, task: Task) -> bool:
        if not task.friend_link_1 or not task.friend_link_2:
            self.app.ui_friend_claim_issue(task.id, "尚未采集两条好友邀请链接，不能打开提货流程")
            return False
        url = str(task.pickup_url or "").strip()
        if not url:
            self.app.ui_friend_claim_issue(task.id, "任务缺少提货网址，无法打开提货页面")
            return False
        if not re.match(r"^https?://", url, re.IGNORECASE):
            url = f"https://{url}"
        self._goto(url)
        if self._stopped():
            return False
        self.app.ui_friend_claim_ready(task.id, "已打开提货页；请核对页面字段后完成本次提货")
        return True

    def capture_friend_claim_result(self, task: Task) -> bool:
        if self._stopped():
            return False
        path = self._capture_page(task, "提货结果")
        self.app.ui_friend_claim_result_captured(task.id, path)
        return True

    def finish_friend_claim_and_logout(self, task: Task) -> bool:
        if not self.logout_by_menu():
            self.app.ui_friend_claim_issue(task.id, "提货完成后账号登出未确认")
            return False
        self.app.ui_friend_claim_finished(task.id)
        return True

    def full_auto_friend_claim_collection(self, tasks: list[Task]) -> None:
        self.app.ui_friend_claim_batch_started(len(tasks))
        stopped = False
        for task_number, task in enumerate(tasks, start=1):
            if not self._wait_auto_gate():
                stopped = True
                break
            self.app.ui_friend_claim_batch_task_started(task, task_number, len(tasks))

            try:
                collected = self.start_friend_claim_account(task)
            except Exception as exc:
                self._log(f"好友链接采集异常：{exc}")
                self.app.ui_friend_claim_issue(task.id, f"好友链接采集异常：{exc}")
                collected = False

            if self._stopped():
                stopped = True
                break

            try:
                logged_out = not self._is_logged_in() or self.logout_by_menu()
            except Exception as exc:
                self._log(f"好友链接采集后登出异常：{exc}")
                logged_out = False

            if collected and logged_out:
                self.app.ui_friend_claim_collection_finished(task.id)
            elif not logged_out:
                self.app.ui_friend_claim_issue(task.id, "好友链接采集后账号登出未确认")

            if self._stopped():
                stopped = True
                break
            time.sleep(0.75)

        self.app.ui_friend_claim_batch_finished(stopped or self._stopped())

    def logout_by_menu(self) -> bool:
        account_menu = self._first_visible(
            ["#account_pulldown", ".pulldown.global_action_link"],
            timeout_ms=12000,
        )
        if account_menu is None:
            self._log("未找到右上角账号菜单，无法确认退出")
            return False
        try:
            account_menu.click(no_wait_after=True)
        except Exception:
            try:
                account_menu.evaluate("element => element.click()")
            except Exception:
                self._log("未能点击右上角账号菜单")
                return False
        self._log("已点击右上角账号菜单")
        time.sleep(0.5)
        logout = self._first_visible(
            [
                "a[href*='logout']",
                "a.popup_menu_item:has-text('退出账户')",
                "a.popup_menu_item:has-text('退出')",
                "text=Logout",
                "text=Sign out",
            ],
            timeout_ms=12000,
        )
        if logout is None:
            self._log("未能自动完成菜单退出，请人工点击右上角账号菜单退出")
            return False
        try:
            # Steam's logout item triggers a navigation; do not wait for every asset to finish.
            logout.evaluate("element => element.click()")
        except Exception:
            try:
                logout.click(no_wait_after=True)
            except Exception:
                self._log("未能点击退出账户")
                return False
        self._log("已点击退出账户")
        deadline = time.time() + 20
        while time.time() < deadline and not self._stopped():
            time.sleep(0.5)
            if not self._is_logged_in():
                try:
                    self.page.wait_for_load_state("domcontentloaded", timeout=10000)
                except Exception:
                    pass
                time.sleep(1)
                self._log("已确认退出当前账号")
                return True
        self._log("点击退出后仍检测到登录状态")
        return False


class SteamTaskAssistant:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.store = TaskStore(DB_PATH)
        self.current_task_id: int | None = None
        self.current_code_index = 0
        self.screenshot_root = SCREENSHOT_DIR
        self.emergency_stopped = False
        self.auto_pause_event = threading.Event()
        self.auto_stop_event = threading.Event()
        self.auto_running = False
        self.suppress_selection_event = False
        self.filter_mode_var = StringVar(value="兑换码兑换")
        self.auto_limit_var = IntVar(value=3)
        self.friend_claim_limit_var = IntVar(value=58)
        self.friend_claim_use_system_proxy_var = BooleanVar(
            value=os.environ.get("STEAM_ASSISTANT_USE_SYSTEM_PROXY", "").lower()
            in {"1", "true", "yes"}
        )
        self.auto_progress_var = StringVar(value="全自动：未启动")
        self.purchase_progress_var = StringVar(value="购买流程：未开始")
        self.activation_progress_var = StringVar(value="激活流程：未开始")
        self.playtime_progress_var = StringVar(value="游玩检测：未开始")
        self.friend_claim_progress_var = StringVar(value="好友码提货：未开始")
        self.current_var = StringVar(value="当前账号：无")
        self.next_var = StringVar(value="下一个账号：无")
        self.status_var = StringVar(value="就绪")
        self.browser_worker = BrowserWorker(self)
        self._build_ui()
        self.refresh_tasks()
        self.root.protocol("WM_DELETE_WINDOW", self.close_app)

    def _build_ui(self) -> None:
        self.root.title("Steam账号兑换/激活/购买/游玩检测/好友码提货助手")
        self.root.geometry("1360x820")
        self.root.minsize(1180, 720)

        toolbar = ttk.Frame(self.root, padding=8)
        toolbar.pack(side=TOP, fill=X)
        ttk.Button(toolbar, text="导入标准表格", command=self.import_excel).pack(side=LEFT, padx=4)
        ttk.Button(toolbar, text="选择截图目录", command=self.select_screenshot_dir).pack(side=LEFT, padx=4)
        ttk.Button(toolbar, text="导出结果CSV", command=self.export_csv).pack(side=LEFT, padx=4)
        ttk.Separator(toolbar, orient="vertical").pack(side=LEFT, fill=Y, padx=8)
        ttk.Label(toolbar, text="工作模式").pack(side=LEFT, padx=(0, 4))
        mode_box = ttk.Combobox(
            toolbar,
            textvariable=self.filter_mode_var,
            values=("兑换码兑换", "激活码激活", "游戏购买", "游玩时间检测", "好友码提货", "全部模式"),
            state="readonly",
            width=12,
        )
        mode_box.pack(side=LEFT, padx=4)
        mode_box.bind("<<ComboboxSelected>>", lambda _e: self.on_mode_changed())
        ttk.Separator(toolbar, orient="vertical").pack(side=LEFT, fill=Y, padx=8)
        ttk.Button(toolbar, text="急停", command=self.emergency_stop).pack(side=LEFT, padx=4)
        ttk.Button(toolbar, text="解除急停", command=self.release_emergency_stop).pack(side=LEFT, padx=4)

        main = ttk.PanedWindow(self.root, orient="horizontal")
        main.pack(fill=BOTH, expand=True)

        left = ttk.Frame(main, padding=8)
        main.add(left, weight=2)
        self.task_tree = ttk.Treeview(
            left,
            columns=("row", "status", "store_code", "store_name", "account", "mode", "game"),
            show="headings",
            height=26,
        )
        columns = {
            "row": ("序号", 54),
            "status": ("状态", 96),
            "store_code": ("店编", 96),
            "store_name": ("店名", 240),
            "account": ("Steam账号", 160),
            "mode": ("模式", 96),
            "game": ("游戏", 120),
        }
        for key, (title, width) in columns.items():
            self.task_tree.heading(key, text=title)
            self.task_tree.column(key, width=width, anchor="w")
        self.task_tree.pack(fill=BOTH, expand=True)
        self.task_tree.bind("<<TreeviewSelect>>", self.on_task_selected)

        right = ttk.Frame(main, padding=8)
        main.add(right, weight=3)

        cards = ttk.Frame(right)
        cards.pack(fill=X)
        current_box = ttk.LabelFrame(cards, text="当前账号", padding=8)
        current_box.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 8))
        next_box = ttk.LabelFrame(cards, text="下一个账号", padding=8)
        next_box.pack(side=RIGHT, fill=BOTH, expand=True)
        ttk.Label(current_box, textvariable=self.current_var, justify=LEFT, wraplength=520).pack(anchor="w", fill=X)
        ttk.Label(next_box, textvariable=self.next_var, justify=LEFT, wraplength=420).pack(anchor="w", fill=X)

        self.full_auto_box = ttk.LabelFrame(right, text="全自动兑换（监护运行）", padding=8)
        self.full_auto_box.pack(fill=X, pady=8)
        ttk.Button(self.full_auto_box, text="开始全自动", command=self.start_full_auto_redeem).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.full_auto_box, text="暂停", command=self.pause_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.full_auto_box, text="继续", command=self.resume_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.full_auto_box, text="停止并保留现场", command=self.stop_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(self.full_auto_box, text="本次账号数").pack(side=LEFT, padx=(16, 4))
        ttk.Spinbox(self.full_auto_box, from_=0, to=999, textvariable=self.auto_limit_var, width=5).pack(side=LEFT, padx=4)
        ttk.Label(self.full_auto_box, text="0=全部").pack(side=LEFT, padx=4)
        ttk.Label(self.full_auto_box, textvariable=self.auto_progress_var).pack(side=LEFT, padx=16)

        self.redeem_auto_box = ttk.LabelFrame(
            right,
            text="兑换/激活半自动操作",
            padding=8,
        )
        self.redeem_auto_box.pack(fill=X, pady=8)
        ttk.Button(self.redeem_auto_box, text="启动当前账号", command=self.start_current_task).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="填入当前兑换码/激活码", command=self.fill_current_code).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="确认兑换码", command=self.confirm_redeem_code).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="消费记录截图并返回兑换页", command=self.history_screenshot_back).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="退出当前账号", command=self.logout_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="完成当前账号", command=self.complete_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="标记异常", command=self.mark_exception).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.redeem_auto_box, text="跳过当前账号", command=self.skip_current_account).pack(side=LEFT, padx=4, pady=4)

        self.activation_full_auto_box = ttk.LabelFrame(right, text="全自动激活（监护运行）", padding=8)
        ttk.Button(self.activation_full_auto_box, text="开始全自动激活", command=self.start_full_auto_activation).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.activation_full_auto_box, text="暂停", command=self.pause_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.activation_full_auto_box, text="继续", command=self.resume_full_auto_activation).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.activation_full_auto_box, text="停止并保留现场", command=self.stop_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(self.activation_full_auto_box, text="本次账号数").pack(side=LEFT, padx=(16, 4))
        ttk.Spinbox(self.activation_full_auto_box, from_=0, to=999, textvariable=self.auto_limit_var, width=5).pack(side=LEFT, padx=4)
        ttk.Label(self.activation_full_auto_box, text="0=全部").pack(side=LEFT, padx=4)
        ttk.Label(self.activation_full_auto_box, textvariable=self.activation_progress_var).pack(side=LEFT, padx=16)

        self.activation_box = ttk.LabelFrame(right, text="激活码半自动流程", padding=8)
        activation_row_one = ttk.Frame(self.activation_box)
        activation_row_one.pack(fill=X)
        ttk.Button(activation_row_one, text="1 一键登录并打开激活页", command=self.start_current_task).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(activation_row_one, text="2 填入当前激活码", command=self.fill_current_code).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(activation_row_one, text="3 确认激活并截图", command=self.confirm_activation_code).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(activation_row_one, text="4 验证资产并登出", command=self.activation_finish_logout).pack(side=LEFT, padx=4, pady=4)
        activation_row_two = ttk.Frame(self.activation_box)
        activation_row_two.pack(fill=X)
        ttk.Button(activation_row_two, text="退出当前账号", command=self.logout_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(activation_row_two, text="完成当前账号", command=self.complete_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(activation_row_two, text="标记异常", command=self.mark_exception).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(activation_row_two, text="跳过当前账号", command=self.skip_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(activation_row_two, textvariable=self.activation_progress_var).pack(side=LEFT, padx=16)

        self.purchase_box = ttk.LabelFrame(right, text="游戏购买半自动流程", padding=8)
        purchase_row_one = ttk.Frame(self.purchase_box)
        purchase_row_one.pack(fill=X)
        ttk.Button(purchase_row_one, text="1 一键登录", command=self.purchase_login).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_one, text="2 清空购物车", command=self.purchase_clear_cart).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_one, text="3 检查数字资产", command=self.purchase_check_license).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_one, text="4 提交游戏到购物车", command=self.purchase_add_cart).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_one, text="5 跳转支付流程", command=self.purchase_prepare_checkout).pack(side=LEFT, padx=4, pady=4)
        purchase_row_two = ttk.Frame(self.purchase_box)
        purchase_row_two.pack(fill=X)
        ttk.Button(purchase_row_two, text="6 购买并截图", command=self.purchase_confirm_checkout).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_two, text="7 验证资产并登出", command=self.purchase_finish_logout).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_two, text="标记异常", command=self.mark_exception).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_two, text="跳过当前账号", command=self.skip_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(purchase_row_two, textvariable=self.purchase_progress_var).pack(side=LEFT, padx=16)
        purchase_row_three = ttk.Frame(self.purchase_box)
        purchase_row_three.pack(fill=X)
        ttk.Button(purchase_row_three, text="批量购买到确认点", command=self.start_full_auto_purchase).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_three, text="暂停", command=self.pause_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_three, text="继续", command=self.resume_full_auto_purchase).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(purchase_row_three, text="停止并保留现场", command=self.stop_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(purchase_row_three, text="本次账号数").pack(side=LEFT, padx=(16, 4))
        ttk.Spinbox(purchase_row_three, from_=0, to=999, textvariable=self.auto_limit_var, width=5).pack(side=LEFT, padx=4)
        self.purchase_box.pack(fill=X, pady=8)

        self.playtime_box = ttk.LabelFrame(right, text="游玩时间检测", padding=8)
        playtime_row_one = ttk.Frame(self.playtime_box)
        playtime_row_one.pack(fill=X)
        ttk.Button(playtime_row_one, text="导入/刷新索尼克检测任务", command=self.import_playtime_tasks).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(playtime_row_one, text="公开页检测当前账号", command=self.detect_current_playtime_public).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(playtime_row_one, text="登录检测当前账号", command=self.detect_current_playtime_login).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(playtime_row_one, text="批量登录检测", command=self.start_full_auto_playtime_login).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(playtime_row_one, text="批量公开检测", command=self.start_full_auto_playtime_public).pack(side=LEFT, padx=4, pady=4)
        playtime_row_two = ttk.Frame(self.playtime_box)
        playtime_row_two.pack(fill=X)
        ttk.Button(playtime_row_two, text="暂停", command=self.pause_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(playtime_row_two, text="继续", command=self.resume_full_auto_playtime).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(playtime_row_two, text="停止并保留现场", command=self.stop_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(playtime_row_two, text="本次账号数").pack(side=LEFT, padx=(16, 4))
        ttk.Spinbox(playtime_row_two, from_=0, to=999, textvariable=self.auto_limit_var, width=5).pack(side=LEFT, padx=4)
        ttk.Label(playtime_row_two, text="0=全部").pack(side=LEFT, padx=4)
        ttk.Label(playtime_row_two, textvariable=self.playtime_progress_var).pack(side=LEFT, padx=16)

        self.friend_claim_box = ttk.LabelFrame(right, text="好友码提货（半自动）", padding=8)
        friend_claim_row_one = ttk.Frame(self.friend_claim_box)
        friend_claim_row_one.pack(fill=X)
        ttk.Button(
            friend_claim_row_one,
            text="1 一键登录并采集两条好友链接",
            command=self.friend_claim_login_collect,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_claim_row_one,
            text="2 打开提货页",
            command=self.friend_claim_open_page,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_claim_row_one,
            text="3 提货结果截图",
            command=self.friend_claim_capture_result,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_claim_row_one,
            text="4 完成提货并登出",
            command=self.friend_claim_finish_logout,
        ).pack(side=LEFT, padx=4, pady=4)
        friend_claim_row_two = ttk.Frame(self.friend_claim_box)
        friend_claim_row_two.pack(fill=X)
        ttk.Button(friend_claim_row_two, text="复制好友链接1", command=lambda: self.copy_friend_link(1)).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(friend_claim_row_two, text="复制好友链接2", command=lambda: self.copy_friend_link(2)).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(friend_claim_row_two, text="复制提货码", command=self.copy_pickup_code).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(friend_claim_row_two, text="标记异常", command=self.mark_exception).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(friend_claim_row_two, text="跳过当前账号", command=self.skip_current_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(friend_claim_row_two, textvariable=self.friend_claim_progress_var).pack(side=LEFT, padx=16)

        self.friend_claim_full_auto_box = ttk.LabelFrame(
            right,
            text="58家好友链接全自动采集",
            padding=8,
        )
        friend_claim_auto_row_one = ttk.Frame(self.friend_claim_full_auto_box)
        friend_claim_auto_row_one.pack(fill=X)
        ttk.Button(
            friend_claim_auto_row_one,
            text="导入/刷新红色沙漠58家任务",
            command=self.import_friend_claim_tasks,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_claim_auto_row_one,
            text="开始全自动采集",
            command=self.start_full_auto_friend_claim,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(friend_claim_auto_row_one, text="暂停", command=self.pause_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_claim_auto_row_one,
            text="继续",
            command=self.resume_full_auto_friend_claim,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_claim_auto_row_one,
            text="停止并保留现场",
            command=self.stop_full_auto,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(friend_claim_auto_row_one, text="本次账号数").pack(side=LEFT, padx=(16, 4))
        ttk.Spinbox(
            friend_claim_auto_row_one,
            from_=0,
            to=999,
            textvariable=self.friend_claim_limit_var,
            width=5,
        ).pack(side=LEFT, padx=4)
        ttk.Label(friend_claim_auto_row_one, text="0=全部").pack(side=LEFT, padx=4)
        ttk.Checkbutton(
            friend_claim_auto_row_one,
            text="使用系统代理",
            variable=self.friend_claim_use_system_proxy_var,
        ).pack(side=LEFT, padx=(12, 4))
        friend_claim_auto_row_two = ttk.Frame(self.friend_claim_full_auto_box)
        friend_claim_auto_row_two.pack(fill=X)
        ttk.Label(
            friend_claim_auto_row_two,
            text="流程：登录 -> 两条快速邀请链接 -> 登出 -> 下一个账号；异常账号记录后继续。",
        ).pack(side=LEFT, padx=4, pady=2)
        ttk.Label(friend_claim_auto_row_two, textvariable=self.friend_claim_progress_var).pack(
            side=LEFT,
            padx=16,
        )
        friend_game_claim_row = ttk.Frame(self.friend_claim_full_auto_box)
        friend_game_claim_row.pack(fill=X)
        ttk.Button(
            friend_game_claim_row,
            text="当前账号自动游戏领取",
            command=self.friend_game_claim_current,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_game_claim_row,
            text="开始全自动游戏领取",
            command=self.start_full_auto_friend_game_claim,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(friend_game_claim_row, text="暂停", command=self.pause_full_auto).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_game_claim_row,
            text="继续游戏领取",
            command=self.resume_full_auto_friend_game_claim,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(
            friend_game_claim_row,
            text="停止并保留现场",
            command=self.stop_full_auto,
        ).pack(side=LEFT, padx=4, pady=4)
        ttk.Label(
            friend_game_claim_row,
            text="流程：登录 -> 通知礼物 -> 领取 -> 资产确认截图 -> 登出 -> 下一个账号。",
        ).pack(side=LEFT, padx=12, pady=2)

        self.manual_box = ttk.LabelFrame(right, text="应急手动复制（仅在半自动失败时使用）", padding=8)
        self.manual_box.pack(fill=X, pady=8)
        ttk.Button(self.manual_box, text="复制账号", command=self.copy_account).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.manual_box, text="复制密码", command=self.copy_password).pack(side=LEFT, padx=4, pady=4)
        ttk.Button(self.manual_box, text="复制当前码", command=self.copy_current_code).pack(side=LEFT, padx=4, pady=4)

        self.code_box = ttk.LabelFrame(right, text="当前任务号码", padding=8)
        self.code_box.pack(fill=BOTH, expand=True, pady=8)
        controls = ttk.Frame(self.code_box)
        controls.pack(fill=X)
        ttk.Button(controls, text="上一个码", command=lambda: self.shift_code(-1)).pack(side=LEFT, padx=4)
        ttk.Button(controls, text="下一个码", command=lambda: self.shift_code(1)).pack(side=LEFT, padx=4)
        self.code_tree = ttk.Treeview(self.code_box, columns=("idx", "kind", "code", "result"), show="headings", height=8)
        self.code_tree.heading("idx", text="序号")
        self.code_tree.heading("kind", text="类型")
        self.code_tree.heading("code", text="号码")
        self.code_tree.heading("result", text="自动结果")
        self.code_tree.column("idx", width=60, anchor="center")
        self.code_tree.column("kind", width=100, anchor="center")
        self.code_tree.column("code", width=320, anchor="w")
        self.code_tree.column("result", width=120, anchor="center")
        self.code_tree.pack(fill=BOTH, expand=True, pady=(6, 0))
        self.code_tree.bind("<<TreeviewSelect>>", self.on_code_selected)

        self.log_box = ttk.LabelFrame(right, text="日志", padding=8)
        self.log_box.pack(fill=BOTH, expand=True)
        self.log_tree = ttk.Treeview(self.log_box, columns=("time", "type", "message"), show="headings", height=8)
        self.log_tree.heading("time", text="时间")
        self.log_tree.heading("type", text="类型")
        self.log_tree.heading("message", text="内容")
        self.log_tree.column("time", width=150, anchor="w")
        self.log_tree.column("type", width=90, anchor="w")
        self.log_tree.column("message", width=820, anchor="w")
        self.log_tree.pack(fill=BOTH, expand=True)

        ttk.Label(self.root, textvariable=self.status_var, relief="sunken", anchor="w", padding=4).pack(side=BOTTOM, fill=X)
        self.update_mode_controls()

    def import_excel(self) -> None:
        path = filedialog.askopenfilename(title="选择标准任务表格", filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")])
        if not path:
            return
        try:
            count = self.store.import_excel(Path(path), replace=True)
        except Exception as exc:
            messagebox.showerror("导入失败", str(exc))
            return
        self.current_task_id = None
        self.refresh_tasks()
        self.status_var.set(f"已导入 {count} 条任务")

    def select_screenshot_dir(self) -> None:
        path = filedialog.askdirectory(title="选择截图保存目录")
        if path:
            self.screenshot_root = Path(path)
            self.screenshot_root.mkdir(parents=True, exist_ok=True)
            self.status_var.set(f"截图目录：{self.screenshot_root}")

    def export_csv(self) -> None:
        path = filedialog.asksaveasfilename(title="导出结果CSV", defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return
        self.store.export_csv(Path(path))
        self.status_var.set(f"已导出：{path}")

    def import_playtime_tasks(self) -> None:
        try:
            count = self.store.import_playtime_tasks(APP_DIR.parent, replace=True)
        except Exception as exc:
            messagebox.showerror("导入游玩检测任务失败", str(exc))
            return
        self.filter_mode_var.set("游玩时间检测")
        self.current_task_id = None
        self.update_mode_controls()
        self.refresh_tasks()
        self.status_var.set(f"已导入/刷新 {count} 个索尼克赛车交叉世界账号检测任务")

    def import_friend_claim_tasks(self) -> None:
        if self.auto_running:
            messagebox.showwarning("批次运行中", "请先停止当前批次，再刷新58家好友链接任务。")
            return
        try:
            count = self.store.import_friend_claim_tasks(APP_DIR.parent, replace=True)
        except Exception as exc:
            messagebox.showerror("导入好友链接任务失败", str(exc))
            return
        self.filter_mode_var.set("好友码提货")
        self.friend_claim_limit_var.set(count)
        self.current_task_id = None
        self.update_mode_controls()
        self.refresh_tasks()
        self.friend_claim_progress_var.set(f"好友码提货：已导入{count}家待采集账号")
        self.status_var.set(f"已导入/刷新 {count} 家红色沙漠部署账号的好友链接采集任务")

    def current_playtime_task(self) -> Task | None:
        task = self.current_task()
        if task is None:
            return None
        if task.mode != "游玩时间检测":
            messagebox.showwarning("模式不正确", "请先切换到“游玩时间检测”模式。")
            return None
        return task

    def detect_current_playtime_public(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_playtime_task()
        if task is None:
            return
        self.store.update_status(task.id, "处理中")
        self.refresh_tasks()
        self.browser_worker.submit("detect_playtime_public", task)

    def detect_current_playtime_login(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_playtime_task()
        if task is None:
            return
        self.store.update_status(task.id, "处理中")
        self.refresh_tasks()
        self.browser_worker.submit("detect_playtime_login", task)

    def start_full_auto_playtime_public(self) -> None:
        self.start_full_auto_playtime(use_login=False)

    def start_full_auto_playtime_login(self) -> None:
        self.start_full_auto_playtime(use_login=True)

    def start_full_auto_playtime(self, use_login: bool) -> None:
        if not self.ensure_not_stopped() or self.auto_running:
            return
        if self.filter_mode_var.get() != "游玩时间检测":
            messagebox.showwarning("模式不正确", "游玩检测只会在“游玩时间检测”模式中运行。")
            return
        limit = max(0, int(self.auto_limit_var.get() or 0))
        tasks = self.store.pending_playtime_tasks(self.current_task_id, limit)
        if not tasks:
            messagebox.showinfo("没有任务", "从当前账号开始没有待检测的游玩时间任务。")
            return
        mode_name = "登录检测" if use_login else "公开页检测"
        if not messagebox.askyesno(
            f"开始批量{mode_name}",
            f"将从当前账号开始连续检测 {len(tasks)} 个账号。\n"
            "程序只读取上次游玩的游戏和时间，不会执行购买、兑换或激活操作。\n\n是否开始？",
        ):
            return
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.browser_worker.submit("full_auto_playtime", tasks, use_login)

    def resume_full_auto_playtime(self) -> None:
        if self.emergency_stopped:
            return
        if self.auto_running:
            self.auto_pause_event.clear()
            self.playtime_progress_var.set("游玩检测：继续运行")
            self.log("游玩检测", "已继续运行")
            return
        if self.filter_mode_var.get() != "游玩时间检测":
            messagebox.showwarning("模式不正确", "继续检测只适用于“游玩时间检测”模式。")
            return
        self.start_full_auto_playtime(use_login=True)

    def start_full_auto_friend_claim(self) -> None:
        if not self.ensure_not_stopped() or self.auto_running:
            return
        if self.filter_mode_var.get() != "好友码提货":
            messagebox.showwarning("模式不正确", "全自动采集只会在“好友码提货”模式中运行。")
            return
        limit = max(0, int(self.friend_claim_limit_var.get() or 0))
        tasks = self.store.pending_friend_claim_tasks(self.current_task_id, limit)
        if not tasks:
            messagebox.showinfo("没有任务", "当前没有待采集两条好友链接的账号。")
            return
        if self.friend_claim_use_system_proxy_var.get():
            os.environ["STEAM_ASSISTANT_USE_SYSTEM_PROXY"] = "1"
        else:
            os.environ.pop("STEAM_ASSISTANT_USE_SYSTEM_PROXY", None)
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.browser_worker.submit("full_auto_friend_claim_collection", tasks)

    def resume_full_auto_friend_claim(self) -> None:
        if self.emergency_stopped:
            return
        if self.auto_running:
            self.auto_pause_event.clear()
            self.friend_claim_progress_var.set("好友码提货：批量继续运行")
            self.log("好友码提货", "已继续批量采集")
            return
        if self.filter_mode_var.get() != "好友码提货":
            messagebox.showwarning("模式不正确", "继续采集只适用于“好友码提货”模式。")
            return
        self.start_full_auto_friend_claim()

    def friend_game_claim_current(self) -> None:
        task = self.current_friend_claim_task()
        if task is None:
            return
        if resolve_game_product(task.game_name) is None:
            messagebox.showerror("游戏未配置", f"无法识别游戏：{task.game_name}")
            return
        self.store.update_friend_claim_state(task.id, "处理中", "正在自动游戏领取")
        self.friend_claim_progress_var.set("好友码提货：正在自动游戏领取")
        self.refresh_tasks()
        self.browser_worker.submit("claim_friend_game_gift", task)

    def start_full_auto_friend_game_claim(self) -> None:
        if not self.ensure_not_stopped() or self.auto_running:
            return
        if self.filter_mode_var.get() != "好友码提货":
            messagebox.showwarning("模式不正确", "自动游戏领取只会在“好友码提货”模式中运行。")
            return
        limit = max(0, int(self.friend_claim_limit_var.get() or 0))
        tasks = self.store.pending_friend_game_claim_tasks(self.current_task_id, limit)
        if not tasks:
            messagebox.showinfo("没有任务", "当前没有待执行自动游戏领取的账号。")
            return
        if not messagebox.askyesno(
            "开始自动游戏领取",
            f"将从当前账号开始连续处理 {len(tasks)} 个账号。\n\n"
            "程序会自动登录账号、点击通知中的礼物、执行领取、打开资产清单确认目标游戏并截图，然后登出进入下一个账号。\n\n"
            "如果没有找到礼物通知且资产清单也没有目标游戏，程序会停在当前账号并标记人工处理。\n\n是否开始？",
            default=messagebox.NO,
        ):
            return
        if self.friend_claim_use_system_proxy_var.get():
            os.environ["STEAM_ASSISTANT_USE_SYSTEM_PROXY"] = "1"
        else:
            os.environ.pop("STEAM_ASSISTANT_USE_SYSTEM_PROXY", None)
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.browser_worker.submit("full_auto_friend_game_claim", tasks)

    def resume_full_auto_friend_game_claim(self) -> None:
        if self.emergency_stopped:
            return
        if self.auto_running:
            self.auto_pause_event.clear()
            self.friend_claim_progress_var.set("好友码提货：自动游戏领取继续运行")
            self.log("自动游戏领取", "已继续运行")
            return
        if self.filter_mode_var.get() != "好友码提货":
            messagebox.showwarning("模式不正确", "继续游戏领取只适用于“好友码提货”模式。")
            return
        self.start_full_auto_friend_game_claim()

    def refresh_tasks(self) -> None:
        for item in self.task_tree.get_children():
            self.task_tree.delete(item)
        tasks = self.store.filtered_tasks(self.filter_mode_var.get())
        for task in tasks:
            game_display = task.last_played_game if task.mode == "游玩时间检测" and task.last_played_game else task.game_name
            if task.mode == "游玩时间检测" and task.last_played_days:
                game_display = f"{game_display}（{task.last_played_days}）" if game_display else task.last_played_days
            if task.mode == "好友码提货":
                game_display = task.pickup_status or "待采集"
            self.task_tree.insert(
                "",
                END,
                iid=str(task.id),
                values=(task.row_no, task.status, task.store_code, task.store_name, task.steam_account, task.mode, game_display),
            )
        visible = {task.id for task in tasks}
        if self.current_task_id not in visible:
            first = self.store.next_task_after(None, self.filter_mode_var.get())
            self.set_current(first.id if first else None)
        else:
            self.set_current(self.current_task_id)

    def on_mode_changed(self) -> None:
        self.current_task_id = None
        self.update_mode_controls()
        self.refresh_tasks()
        self.status_var.set(f"已切换工作模式：{self.filter_mode_var.get()}")

    def update_mode_controls(self) -> None:
        mode = self.filter_mode_var.get()
        self.full_auto_box.pack_forget()
        self.redeem_auto_box.pack_forget()
        self.activation_full_auto_box.pack_forget()
        self.activation_box.pack_forget()
        self.purchase_box.pack_forget()
        self.playtime_box.pack_forget()
        self.friend_claim_box.pack_forget()
        self.friend_claim_full_auto_box.pack_forget()
        self.code_box.pack_forget()

        if mode == "兑换码兑换":
            self.full_auto_box.pack(fill=X, pady=8, before=self.manual_box)
            self.redeem_auto_box.pack(fill=X, pady=8, before=self.manual_box)
            self.code_box.pack(fill=BOTH, expand=True, pady=8, before=self.log_box)
        elif mode == "激活码激活":
            self.activation_full_auto_box.pack(fill=X, pady=8, before=self.manual_box)
            self.activation_box.pack(fill=X, pady=8, before=self.manual_box)
            self.code_box.pack(fill=BOTH, expand=True, pady=8, before=self.log_box)
        elif mode == "游戏购买":
            self.purchase_box.pack(fill=X, pady=8, before=self.manual_box)
        elif mode == "游玩时间检测":
            self.playtime_box.pack(fill=X, pady=8, before=self.manual_box)
        elif mode == "好友码提货":
            self.friend_claim_full_auto_box.pack(fill=X, pady=8, before=self.manual_box)
            self.friend_claim_box.pack(fill=X, pady=8, before=self.manual_box)
            self.code_box.pack(fill=BOTH, expand=True, pady=8, before=self.log_box)
        else:
            self.status_var.set("请选择具体工作模式后再执行自动化操作")

    def on_task_selected(self, _event: object) -> None:
        if self.suppress_selection_event:
            return
        selection = self.task_tree.selection()
        if selection:
            self.set_current(int(selection[0]))

    def set_current(self, task_id: int | None) -> None:
        self.current_task_id = task_id
        self.current_code_index = 0
        task = self.current_task()
        if task is None:
            self.current_var.set("当前账号：无")
            self.next_var.set("下一个账号：无")
            self.fill_codes(None)
            self.fill_logs(None)
            return
        self.current_var.set(self.format_task(task, show_password=True))
        next_task = self.store.next_task_after(task.id, self.filter_mode_var.get())
        self.next_var.set(self.format_task(next_task, show_password=False) if next_task else "下一个账号：无")
        self.fill_codes(task)
        self.fill_logs(task.id)
        target = str(task.id)
        if target in self.task_tree.get_children() and self.task_tree.selection() != (target,):
            self.suppress_selection_event = True
            try:
                self.task_tree.selection_set(target)
                self.task_tree.focus(target)
            finally:
                self.suppress_selection_event = False

    def current_task(self) -> Task | None:
        return self.store.get_task(self.current_task_id)

    def format_task(self, task: Task | None, show_password: bool) -> str:
        if task is None:
            return "无"
        password = task.steam_password if show_password else "********"
        product = resolve_game_product(task.game_name)
        product_line = ""
        if task.mode == "游戏购买" and product is not None:
            product_line = (
                f"Steam正式名称：{product['official_name']}\n"
                f"App ID：{product['app_id']}    标准版Package：{product['package_id']}\n"
            )
        playtime_line = ""
        if task.mode == "游玩时间检测":
            playtime_line = (
                f"账号链接：{task.profile_link or '未匹配'}\n"
                f"上次游玩游戏：{task.last_played_game or '未检测'}\n"
                f"上次游玩时间：{task.last_played_time or '未检测'}\n"
                f"距离今天：{task.last_played_days or '未检测'}\n"
                f"检测来源：{task.last_played_source or '未检测'}\n"
            )
        friend_claim_line = ""
        if task.mode == "好友码提货":
            friend_claim_line = (
                f"提货码：{task.pickup_code or '未录入'}\n"
                f"提货网址：{task.pickup_url or '未录入'}\n"
                f"好友链接1：{'已采集' if task.friend_link_1 else '未采集'}\n"
                f"好友链接2：{'已采集' if task.friend_link_2 else '未采集'}\n"
                f"提货状态：{task.pickup_status or '待处理'}\n"
            )
        return (
            f"店编：{task.store_code}\n"
            f"店名：{task.store_name}\n"
            f"省份/城市：{task.province} / {task.city}\n"
            f"Steam账号：{task.steam_account}\n"
            f"Steam密码：{password}\n"
            f"模式：{task.mode}\n"
            f"购买游戏：{task.game_name} {task.game_price}\n"
            f"{product_line}"
            f"{playtime_line}"
            f"{friend_claim_line}"
            f"兑换码数量：{len(task.voucher_codes)}    激活码数量：{len(task.activation_codes)}\n"
            f"状态：{task.status}\n"
            f"备注：{task.note}"
        )

    def fill_codes(self, task: Task | None) -> None:
        for item in self.code_tree.get_children():
            self.code_tree.delete(item)
        if task is None:
            return
        results = self.store.voucher_results(task.id) if task.mode in {"兑换码兑换", "激活码激活"} else {}
        if task.mode == "兑换码兑换":
            rows = [("兑换码", code) for code in task.voucher_codes]
        elif task.mode == "激活码激活":
            rows = [("激活码", code) for code in task.activation_codes]
        elif task.mode == "游玩时间检测":
            rows = [("账号链接", task.profile_link or "未匹配账号链接")]
        elif task.mode == "好友码提货":
            rows = [
                ("提货码", task.pickup_code or "未录入"),
                ("好友链接1", task.friend_link_1 or "未采集"),
                ("好友链接2", task.friend_link_2 or "未采集"),
                ("提货网址", task.pickup_url or "未录入"),
            ]
        else:
            rows = [("购买游戏", task.game_name)] if task.game_name else []
        for idx, (kind, code) in enumerate(rows, start=1):
            result = results.get(idx - 1)
            if task.mode == "好友码提货":
                result_status = task.pickup_status or "待处理"
            else:
                result_status = result["status"] if result else "待处理"
            self.code_tree.insert("", END, iid=str(idx - 1), values=(idx, kind, code, result_status))
        if rows:
            self.code_tree.selection_set("0")
            self.code_tree.focus("0")

    def fill_logs(self, task_id: int | None) -> None:
        for item in self.log_tree.get_children():
            self.log_tree.delete(item)
        for event in reversed(self.store.events(task_id)):
            self.log_tree.insert("", END, values=(event["event_time"], event["event_type"], event["message"]))

    def current_codes(self) -> list[str]:
        task = self.current_task()
        if task is None:
            return []
        if task.mode == "兑换码兑换":
            return task.voucher_codes
        if task.mode == "激活码激活":
            return task.activation_codes
        if task.mode == "好友码提货":
            return [
                task.pickup_code,
                task.friend_link_1,
                task.friend_link_2,
                task.pickup_url,
            ]
        return [task.game_name] if task.game_name else []

    def on_code_selected(self, _event: object) -> None:
        selection = self.code_tree.selection()
        if selection:
            self.current_code_index = int(selection[0])

    def shift_code(self, delta: int) -> None:
        codes = self.current_codes()
        if not codes:
            return
        self.current_code_index = max(0, min(len(codes) - 1, self.current_code_index + delta))
        iid = str(self.current_code_index)
        self.code_tree.selection_set(iid)
        self.code_tree.focus(iid)

    def ensure_not_stopped(self) -> bool:
        if self.emergency_stopped:
            messagebox.showwarning("急停已启用", "当前处于急停状态。请先解除急停再继续。")
            return False
        return True

    def emergency_stop(self) -> None:
        self.emergency_stopped = True
        self.auto_stop_event.set()
        self.log("急停", "用户触发急停，已阻止后续自动化动作")

    def release_emergency_stop(self) -> None:
        self.emergency_stopped = False
        self.auto_stop_event.clear()
        self.log("急停", "用户解除急停")

    def start_full_auto_redeem(self) -> None:
        if not self.ensure_not_stopped() or self.auto_running:
            return
        if self.filter_mode_var.get() != "兑换码兑换":
            messagebox.showwarning("模式不正确", "全自动兑换只会在“兑换码兑换”模式中运行。")
            return
        limit = max(0, int(self.auto_limit_var.get() or 0))
        tasks = self.store.pending_redeem_tasks(self.current_task_id, limit)
        if not tasks:
            messagebox.showinfo("没有任务", "从当前账号开始没有待处理的兑换任务。")
            return
        if not messagebox.askyesno(
            "开始全自动兑换",
            f"将从当前账号开始连续处理 {len(tasks)} 个账号。\n"
            "登录身份不符、结果不明确或登出失败时会自动暂停。\n\n是否开始？",
        ):
            return
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.browser_worker.submit("full_auto_redeem", tasks, self.auto_skip_map(tasks))

    def auto_skip_map(self, tasks: list[Task]) -> dict[int, set[int]]:
        terminal_statuses = {"兑换成功", "已使用", "激活成功", "已拥有", "已被使用", "无效", "地区受限"}
        return {
            task.id: {
                code_index
                for code_index, row in self.store.voucher_results(task.id).items()
                if row["status"] in terminal_statuses
            }
            for task in tasks
        }

    def start_full_auto_activation(self) -> None:
        if not self.ensure_not_stopped() or self.auto_running:
            return
        if self.filter_mode_var.get() != "激活码激活":
            messagebox.showwarning("模式不正确", "全自动激活只会在“激活码激活”模式中运行。")
            return
        limit = max(0, int(self.auto_limit_var.get() or 0))
        tasks = self.store.pending_activation_tasks(self.current_task_id, limit)
        if not tasks:
            messagebox.showinfo("没有任务", "从当前账号开始没有待处理的激活任务。")
            return
        if not messagebox.askyesno(
            "开始全自动激活",
            f"将从当前账号开始连续处理 {len(tasks)} 个账号。\n"
            "程序会自动登录、核对身份、提交激活码、截图成功结果、验证资产并登出。\n\n"
            "激活码会真实消耗；结果不明确、无效、已被其他账号使用或登出失败时会自动暂停。\n\n是否开始？",
            default=messagebox.NO,
        ):
            return
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.browser_worker.submit("full_auto_activation", tasks, self.auto_skip_map(tasks))

    def pause_full_auto(self) -> None:
        if self.auto_running:
            self.auto_pause_event.set()
            if self.filter_mode_var.get() == "游戏购买":
                self.purchase_progress_var.set("购买流程：批量已暂停")
                self.log("批量购买", "已请求暂停，将在当前安全步骤停下")
            elif self.filter_mode_var.get() == "激活码激活":
                self.activation_progress_var.set("激活流程：批量已暂停")
                self.log("全自动激活", "已请求暂停，将在当前安全步骤停下")
            elif self.filter_mode_var.get() == "游玩时间检测":
                self.playtime_progress_var.set("游玩检测：批量已暂停")
                self.log("游玩检测", "已请求暂停，将在当前安全步骤停下")
            elif self.filter_mode_var.get() == "好友码提货":
                self.friend_claim_progress_var.set("好友码提货：批量已暂停")
                self.log("好友码提货", "已请求暂停，将在当前账号安全步骤停下")
            else:
                self.auto_progress_var.set("全自动：已暂停")
                self.log("全自动", "已请求暂停，将在当前安全步骤停下")

    def resume_full_auto(self) -> None:
        if self.emergency_stopped:
            return
        if self.auto_running:
            self.auto_pause_event.clear()
            self.auto_progress_var.set("全自动：继续运行")
            self.log("全自动", "已继续运行")
            return
        if self.filter_mode_var.get() != "兑换码兑换":
            messagebox.showwarning("模式不正确", "断点继续只适用于“兑换码兑换”模式。")
            return
        tasks = self.store.pending_redeem_tasks(self.current_task_id, 1)
        if not tasks:
            messagebox.showinfo("没有任务", "从当前账号开始没有可继续的兑换任务。")
            return
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.auto_progress_var.set("全自动：从断点重新启动")
        self.log("全自动", "批次已结束，现从数据库断点重新启动")
        self.browser_worker.submit("full_auto_redeem", tasks, self.auto_skip_map(tasks))

    def stop_full_auto(self) -> None:
        if self.auto_running:
            self.auto_stop_event.set()
            self.auto_pause_event.clear()
            if self.filter_mode_var.get() == "游戏购买":
                self.purchase_progress_var.set("购买流程：批量正在停止")
                self.log("批量购买", "已请求停止，浏览器现场将保留")
            elif self.filter_mode_var.get() == "激活码激活":
                self.activation_progress_var.set("激活流程：批量正在停止")
                self.log("全自动激活", "已请求停止，浏览器现场将保留")
            elif self.filter_mode_var.get() == "游玩时间检测":
                self.playtime_progress_var.set("游玩检测：批量正在停止")
                self.log("游玩检测", "已请求停止，浏览器现场将保留")
            elif self.filter_mode_var.get() == "好友码提货":
                self.friend_claim_progress_var.set("好友码提货：批量正在停止")
                self.log("好友码提货", "已请求停止，浏览器现场将保留")
            else:
                self.auto_progress_var.set("全自动：正在停止")
                self.log("全自动", "已请求停止，浏览器现场将保留")

    def close_app(self) -> None:
        self.auto_stop_event.set()
        self.auto_pause_event.clear()
        self.browser_worker.shutdown()
        self.root.after(700, self.root.destroy)

    def copy_account(self) -> None:
        task = self.current_task()
        if task:
            copy_to_clipboard(self.root, task.steam_account)
            self.log("应急复制", "已复制账号")

    def copy_password(self) -> None:
        task = self.current_task()
        if task:
            copy_to_clipboard(self.root, task.steam_password)
            self.log("应急复制", "已复制密码")

    def copy_current_code(self) -> None:
        codes = self.current_codes()
        if not codes:
            return
        if not codes[self.current_code_index]:
            messagebox.showwarning("没有可复制内容", "当前项目尚未录入或采集。")
            return
        copy_to_clipboard(self.root, codes[self.current_code_index])
        self.log("应急复制", f"已复制当前码：第 {self.current_code_index + 1} 个")

    def copy_friend_link(self, number: int) -> None:
        task = self.current_task()
        if task is None or task.mode != "好友码提货":
            messagebox.showwarning("模式不正确", "请先切换到“好友码提货”模式。")
            return
        link = task.friend_link_1 if number == 1 else task.friend_link_2
        if not link:
            messagebox.showwarning("尚未采集", f"好友链接{number}尚未采集。")
            return
        copy_to_clipboard(self.root, link)
        self.log("好友码提货", f"已复制好友链接{number}")

    def copy_pickup_code(self) -> None:
        task = self.current_task()
        if task is None or task.mode != "好友码提货":
            messagebox.showwarning("模式不正确", "请先切换到“好友码提货”模式。")
            return
        if not task.pickup_code:
            messagebox.showwarning("未录入提货码", "请在标准任务表的“提货码”列填写后重新导入。")
            return
        copy_to_clipboard(self.root, task.pickup_code)
        self.log("好友码提货", "已复制提货码")

    def current_friend_claim_task(
        self,
        require_pickup_code: bool = False,
        require_pickup_url: bool = False,
    ) -> Task | None:
        if not self.ensure_not_stopped():
            return None
        task = self.current_task()
        if task is None:
            return None
        if task.mode != "好友码提货":
            messagebox.showwarning("模式不正确", "当前任务不是好友码提货任务。")
            return None
        if require_pickup_code and not task.pickup_code:
            messagebox.showwarning("缺少提货码", "请在标准任务表的“提货码”列填写后重新导入。")
            return None
        if require_pickup_url and not task.pickup_url:
            messagebox.showwarning("缺少提货网址", "请在标准任务表的“提货网址”列填写后重新导入。")
            return None
        return task

    def friend_claim_login_collect(self) -> None:
        task = self.current_friend_claim_task()
        if task is None:
            return
        self.store.update_friend_claim_state(task.id, "处理中", "正在登录并采集好友链接")
        self.friend_claim_progress_var.set("好友码提货：正在登录并采集好友链接")
        self.refresh_tasks()
        self.browser_worker.submit("start_friend_claim_account", task)

    def friend_claim_open_page(self) -> None:
        task = self.current_friend_claim_task(require_pickup_code=True, require_pickup_url=True)
        if task is None:
            return
        self.store.update_friend_claim_state(task.id, "处理中", "正在打开提货页")
        self.friend_claim_progress_var.set("好友码提货：正在打开提货页")
        self.refresh_tasks()
        self.browser_worker.submit("open_friend_claim_page", task)

    def friend_claim_capture_result(self) -> None:
        task = self.current_friend_claim_task()
        if task is None:
            return
        self.browser_worker.submit("capture_friend_claim_result", task)

    def friend_claim_finish_logout(self) -> None:
        task = self.current_friend_claim_task()
        if task is None:
            return
        self.browser_worker.submit("finish_friend_claim_and_logout", task)

    def current_purchase_task(self) -> Task | None:
        if not self.ensure_not_stopped():
            return None
        task = self.current_task()
        if task is None:
            return None
        if task.mode != "游戏购买":
            messagebox.showwarning("模式不正确", "当前任务不是游戏购买任务。")
            return None
        product = resolve_game_product(task.game_name)
        if product is None:
            messagebox.showerror("游戏未配置", f"无法识别游戏：{task.game_name}")
            return None
        return task

    def request_cart_clear_confirmation(self, task: Task, item_count: int) -> bool:
        event = threading.Event()
        result = {"confirmed": False}

        def ask() -> None:
            result["confirmed"] = messagebox.askyesno(
                "确认清空购物车",
                f"即将删除当前Steam购物车内的 {item_count} 项内容：\n\n"
                f"店编：{task.store_code}\n"
                f"店名：{task.store_name}\n"
                f"账号：{task.steam_account}\n\n"
                "点击“是”后程序将清空购物车并继续批量购买流程，是否继续？",
                default=messagebox.NO,
            )
            if result["confirmed"]:
                self.store.add_event(
                    task.id,
                    "清空确认",
                    f"用户确认删除购物车内{item_count}项内容",
                )
            event.set()

        self.root.after(0, ask)
        while not event.wait(0.2):
            if self.emergency_stopped or self.auto_stop_event.is_set():
                return False
        return bool(result["confirmed"])

    def request_purchase_confirmation(self, task: Task) -> bool:
        product = resolve_game_product(task.game_name)
        if product is None:
            return False
        event = threading.Event()
        result = {"confirmed": False}

        def ask() -> None:
            result["confirmed"] = messagebox.askyesno(
                "确认本次购买",
                f"即将使用Steam钱包完成真实购买：\n\n"
                f"店编：{task.store_code}\n"
                f"店名：{task.store_name}\n"
                f"账号：{task.steam_account}\n"
                f"游戏：{product['official_name']} 标准版\n"
                f"价格：¥{product['price']:.2f}\n\n"
                "点击“是”后程序将提交最终购买，是否继续？",
                default=messagebox.NO,
            )
            if result["confirmed"]:
                self.store.add_event(
                    task.id,
                    "支付确认",
                    "用户确认使用Steam钱包提交最终购买",
                )
            event.set()

        self.root.after(0, ask)
        while not event.wait(0.2):
            if self.emergency_stopped or self.auto_stop_event.is_set():
                return False
        return bool(result["confirmed"])

    def submit_purchase_command(self, command: str) -> None:
        task = self.current_purchase_task()
        if task is None:
            return
        self.store.update_status(task.id, "处理中")
        self.refresh_tasks()
        self.browser_worker.submit(command, task)

    def purchase_login(self) -> None:
        self.submit_purchase_command("start_purchase_account")

    def purchase_clear_cart(self) -> None:
        self.submit_purchase_command("clear_purchase_cart")

    def purchase_check_license(self) -> None:
        self.submit_purchase_command("check_purchase_license")

    def purchase_open_standard(self) -> None:
        self.submit_purchase_command("open_standard_game_page")

    def purchase_add_cart(self) -> None:
        self.submit_purchase_command("add_standard_game_to_cart")

    def purchase_verify_cart(self) -> None:
        self.submit_purchase_command("verify_purchase_cart")

    def purchase_prepare_checkout(self) -> None:
        self.submit_purchase_command("prepare_purchase_checkout")

    def purchase_confirm_checkout(self) -> None:
        task = self.current_purchase_task()
        if task is None:
            return
        product = resolve_game_product(task.game_name)
        if product is None:
            return
        confirmed = messagebox.askyesno(
            "确认支付",
            f"即将使用Steam钱包完成真实购买：\n\n"
            f"店编：{task.store_code}\n"
            f"账号：{task.steam_account}\n"
            f"游戏：{product['official_name']} 标准版\n"
            f"价格：¥{product['price']:.2f}\n\n"
            "点击“是”后程序将提交最终购买，是否继续？",
        )
        if not confirmed:
            self.log("游戏购买", "用户取消最终支付")
            return
        self.store.add_event(task.id, "支付确认", "用户确认使用Steam钱包提交最终购买")
        self.browser_worker.submit("confirm_purchase_checkout", task)

    def purchase_finish_logout(self) -> None:
        self.submit_purchase_command("verify_purchase_asset_and_logout")

    def start_full_auto_purchase(self) -> None:
        if not self.ensure_not_stopped() or self.auto_running:
            return
        if self.filter_mode_var.get() != "游戏购买":
            messagebox.showwarning("模式不正确", "批量购买只会在“游戏购买”模式中运行。")
            return
        limit = max(0, int(self.auto_limit_var.get() or 0))
        tasks = self.store.pending_purchase_tasks(self.current_task_id, limit)
        if not tasks:
            messagebox.showinfo("没有任务", "从当前账号开始没有待处理的购买任务。")
            return
        if not messagebox.askyesno(
            "开始批量购买",
            f"将从当前账号开始处理 {len(tasks)} 个购买任务。\n\n"
            "程序会自动登录、清空购物车、检查资产、加入购物车并进入结算页；"
            "这些准备步骤不会再弹出中间确认。\n\n"
            "到达每个账号的最终扣款页时，会显示该店编、店名、账号、游戏和价格，"
            "并只在这一笔真实购买前确认一次。\n\n"
            "此处仅确认启动批量准备流程，是否开始？",
            default=messagebox.NO,
        ):
            return
        self.auto_pause_event.clear()
        self.auto_stop_event.clear()
        self.auto_running = True
        self.browser_worker.submit("full_auto_purchase", tasks)

    def resume_full_auto_purchase(self) -> None:
        if self.emergency_stopped:
            return
        if self.auto_running:
            self.auto_pause_event.clear()
            self.purchase_progress_var.set("购买流程：批量继续运行")
            self.log("批量购买", "已继续运行")
            return
        if self.filter_mode_var.get() != "游戏购买":
            messagebox.showwarning("模式不正确", "批量购买继续只适用于“游戏购买”模式。")
            return
        self.start_full_auto_purchase()

    def resume_full_auto_activation(self) -> None:
        if self.emergency_stopped:
            return
        if self.auto_running:
            self.auto_pause_event.clear()
            self.activation_progress_var.set("激活流程：批量继续运行")
            self.log("全自动激活", "已继续运行")
            return
        if self.filter_mode_var.get() != "激活码激活":
            messagebox.showwarning("模式不正确", "批量激活继续只适用于“激活码激活”模式。")
            return
        self.start_full_auto_activation()

    def start_current_task(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_task()
        if task is None:
            return
        self.store.update_status(task.id, "处理中")
        self.refresh_tasks()
        if task.mode == "兑换码兑换":
            self.browser_worker.submit("start_redeem_account", task)
        elif task.mode == "激活码激活":
            self.browser_worker.submit("start_activation_account", task)
        elif task.mode == "游戏购买":
            self.browser_worker.submit("start_purchase_account", task)
        elif task.mode == "游玩时间检测":
            self.browser_worker.submit("detect_playtime_login", task)
        elif task.mode == "好友码提货":
            self.browser_worker.submit("start_friend_claim_account", task)
        self.log("任务", "已启动当前账号半自动流程")

    def fill_current_code(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_task()
        codes = self.current_codes()
        if task is None or not codes:
            return
        code = codes[self.current_code_index]
        if task.mode == "激活码激活":
            self.browser_worker.submit("fill_activation_code", code)
        elif task.mode == "兑换码兑换":
            self.browser_worker.submit("fill_redeem_code", code)

    def confirm_redeem_code(self) -> None:
        if self.ensure_not_stopped():
            self.browser_worker.submit("confirm_redeem_code")

    def confirm_activation_code(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_task()
        codes = self.current_codes()
        if task is None or task.mode != "激活码激活" or not codes:
            return
        code = codes[self.current_code_index]
        self.store.update_status(task.id, "处理中")
        self.refresh_tasks()
        self.browser_worker.submit("activate_current_code", task, self.current_code_index, code)

    def activation_finish_logout(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_task()
        if task is None or task.mode != "激活码激活":
            return
        self.browser_worker.submit("finish_activation_and_logout", task)

    def history_screenshot_back(self) -> None:
        if not self.ensure_not_stopped():
            return
        task = self.current_task()
        if task:
            self.browser_worker.submit("capture_history_then_redeem", task)

    def logout_current_account(self) -> None:
        if self.ensure_not_stopped():
            self.browser_worker.submit("logout_by_menu")

    def complete_current_account(self) -> None:
        task = self.current_task()
        if task:
            self.store.update_status(task.id, "成功")
            self.log("状态", "当前账号已标记成功")
            self.move_to_next()

    def mark_exception(self) -> None:
        task = self.current_task()
        if task:
            self.store.update_status(task.id, "需要人工处理")
            self.log("状态", "当前账号已标记需要人工处理")
            self.refresh_tasks()

    def skip_current_account(self) -> None:
        task = self.current_task()
        if task:
            self.store.update_status(task.id, "跳过")
            self.log("状态", "当前账号已跳过")
            self.move_to_next()

    def move_to_next(self) -> None:
        task = self.current_task()
        next_task = self.store.next_task_after(task.id if task else None, self.filter_mode_var.get())
        self.set_current(next_task.id if next_task else None)

    def make_screenshot_path(self, task: Task, status_text: str) -> Path:
        folder = self.screenshot_root / "-".join(
            [clean_filename_part(task.store_code), clean_filename_part(task.store_name), clean_filename_part(task.steam_account)]
        )
        folder.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        filename = "-".join(
            [
                clean_filename_part(task.store_code),
                clean_filename_part(task.store_name),
                clean_filename_part(task.steam_account),
                clean_filename_part(status_text),
                timestamp,
            ]
        ) + ".png"
        return folder / filename

    def ui_log(self, event_type: str, message: str) -> None:
        self.root.after(0, lambda: self.log(event_type, message))

    def ui_error(self, message: str) -> None:
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        with AUTOMATION_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {message}\n")
        def update() -> None:
            if self.auto_running:
                self.auto_running = False
                self.auto_pause_event.clear()
                self.auto_stop_event.clear()
                if self.filter_mode_var.get() == "游戏购买":
                    self.purchase_progress_var.set("购买流程：自动化异常，已停止")
                elif self.filter_mode_var.get() == "激活码激活":
                    self.activation_progress_var.set("激活流程：自动化异常，已停止")
                elif self.filter_mode_var.get() == "游玩时间检测":
                    self.playtime_progress_var.set("游玩检测：自动化异常，已停止")
                elif self.filter_mode_var.get() == "好友码提货":
                    self.friend_claim_progress_var.set("好友码提货：自动化异常，已停止")
                else:
                    self.auto_progress_var.set("全自动：自动化异常，已停止")
            messagebox.showerror("自动化错误", message)

        self.root.after(0, update)

    def ui_screenshot_saved(self, task_id: int, path: Path, status_text: str) -> None:
        def update() -> None:
            self.store.update_screenshot_dir(task_id, path.parent)
            self.store.add_event(task_id, "截图", f"{status_text}：{path}", str(path))
            self.fill_logs(task_id)
            self.status_var.set(f"已截图：{path}")

        self.root.after(0, update)

    def ui_purchase_step(
        self,
        task_id: int,
        step: str,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.add_event(
                task_id,
                "游戏购买",
                f"{step}：{message}",
                str(screenshot_path or ""),
            )
            self.purchase_progress_var.set(f"购买流程：{step}")
            self.status_var.set(message)
            if self.current_task_id == task_id:
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_purchase_issue(
        self,
        task_id: int,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.update_status(task_id, "需要人工处理")
            self.store.add_event(
                task_id,
                "购买暂停",
                message,
                str(screenshot_path or ""),
            )
            self.purchase_progress_var.set(f"购买流程：已暂停，{message}")
            self.status_var.set(message)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_success(
        self,
        task_id: int,
        message: str,
    ) -> None:
        def update() -> None:
            self.store.update_status(task_id, "处理中")
            self.store.add_event(task_id, "购买成功", message)
            self.purchase_progress_var.set("购买流程：购买成功，等待登出")
            self.status_var.set(message)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_already_deployed(
        self,
        task_id: int,
        message: str,
    ) -> None:
        def update() -> None:
            self.store.update_status(task_id, "成功")
            self.store.add_event(
                task_id,
                "游戏购买",
                message,
            )
            self.purchase_progress_var.set("购买流程：资产已存在，完成部署")
            self.status_var.set(message)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_finished(self, task_id: int) -> None:
        def update() -> None:
            self.store.update_status(task_id, "成功")
            self.store.add_event(task_id, "游戏购买", "购买完成并确认登出")
            self.purchase_progress_var.set("购买流程：完成并已登出")
            if self.current_task_id == task_id:
                next_task = self.store.next_task_after(task_id, "游戏购买")
                self.set_current(next_task.id if next_task else None)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_batch_started(self, task_count: int) -> None:
        def update() -> None:
            self.auto_running = True
            self.purchase_progress_var.set(f"购买流程：批量准备处理 {task_count} 个账号")
            self.status_var.set("批量购买已启动")

        self.root.after(0, update)

    def ui_purchase_batch_task_started(
        self,
        task: Task,
        task_number: int,
        task_count: int,
    ) -> None:
        def update() -> None:
            self.set_current(task.id)
            self.store.update_status(task.id, "处理中")
            self.store.add_event(
                task.id,
                "批量购买",
                f"开始账号 {task_number}/{task_count}",
            )
            self.purchase_progress_var.set(
                f"购买流程：批量 {task_number}/{task_count}，{task.store_code}"
            )
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_batch_paused(self, task_id: int, reason: str) -> None:
        self.auto_pause_event.set()

        def update() -> None:
            self.store.update_status(task_id, "需要人工处理")
            self.store.add_event(task_id, "批量购买暂停", reason)
            self.purchase_progress_var.set(f"购买流程：批量已暂停，{reason}")
            self.status_var.set(reason)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_batch_already_deployed(self, task_id: int) -> None:
        def update() -> None:
            self.store.add_event(task_id, "批量购买", "资产已存在，跳过购买并已登出")
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_purchase_batch_finished(self, stopped: bool) -> None:
        def update() -> None:
            self.auto_running = False
            self.auto_pause_event.clear()
            self.auto_stop_event.clear()
            text = "购买流程：批量已停止" if stopped else "购买流程：批量完成"
            self.purchase_progress_var.set(text)
            self.status_var.set(text)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_playtime_batch_started(self, task_count: int, use_login: bool) -> None:
        def update() -> None:
            self.auto_running = True
            mode_name = "登录检测" if use_login else "公开页检测"
            self.playtime_progress_var.set(f"游玩检测：{mode_name}准备处理 {task_count} 个账号")
            self.status_var.set("批量游玩检测已启动")

        self.root.after(0, update)

    def ui_playtime_task_started(self, task: Task, task_number: int, task_count: int) -> None:
        def update() -> None:
            self.set_current(task.id)
            self.store.update_status(task.id, "处理中")
            self.store.add_event(task.id, "游玩检测", f"开始账号 {task_number}/{task_count}")
            self.playtime_progress_var.set(
                f"游玩检测：账号 {task_number}/{task_count}，{task.store_code}"
            )
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_playtime_result(
        self,
        task_id: int,
        status: str,
        game: str,
        played_time: str,
        days: str,
        source: str,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.update_playtime_result(
                task_id,
                status,
                game,
                played_time,
                days,
                source,
                message,
                str(screenshot_path or ""),
            )
            self.store.add_event(task_id, "游玩检测", message, str(screenshot_path or ""))
            self.playtime_progress_var.set(f"游玩检测：{message}")
            self.status_var.set(message)
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_playtime_batch_finished(self, stopped: bool) -> None:
        def update() -> None:
            self.auto_running = False
            self.auto_pause_event.clear()
            self.auto_stop_event.clear()
            text = "游玩检测：批量已停止" if stopped else "游玩检测：批量完成"
            self.playtime_progress_var.set(text)
            self.status_var.set(text)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_friend_claim_batch_started(self, task_count: int) -> None:
        def update() -> None:
            self.auto_running = True
            self.friend_claim_progress_var.set(f"好友码提货：准备处理 {task_count} 家账号")
            self.status_var.set("58家好友链接全自动采集已启动")

        self.root.after(0, update)

    def ui_friend_claim_batch_task_started(
        self,
        task: Task,
        task_number: int,
        task_count: int,
    ) -> None:
        def update() -> None:
            self.set_current(task.id)
            self.store.update_friend_claim_state(
                task.id,
                "处理中",
                f"正在采集好友链接（{task_number}/{task_count}）",
            )
            self.store.add_event(
                task.id,
                "好友码提货批量",
                f"开始账号 {task_number}/{task_count}",
            )
            self.friend_claim_progress_var.set(
                f"好友码提货：{task_number}/{task_count}，{task.store_code}"
            )
            self.status_var.set(f"正在采集 {task.store_code} 的两条好友链接")
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_friend_claim_collection_finished(self, task_id: int) -> None:
        def update() -> None:
            self.store.update_friend_claim_state(task_id, "成功", "好友链接已采集并已登出")
            self.store.add_event(task_id, "好友码提货批量", "两条好友链接已采集并确认登出")
            self.friend_claim_progress_var.set("好友码提货：当前账号采集完成并已登出")
            self.status_var.set("当前账号的两条好友链接已采集并已登出")
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_claim_batch_finished(self, stopped: bool) -> None:
        def update() -> None:
            self.auto_running = False
            self.auto_pause_event.clear()
            self.auto_stop_event.clear()
            text = "好友码提货：批量已停止" if stopped else "好友码提货：批量采集完成"
            self.friend_claim_progress_var.set(text)
            self.status_var.set(text)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_friend_claim_links_collected(
        self,
        task_id: int,
        friend_link_1: str,
        friend_link_2: str,
    ) -> None:
        def update() -> None:
            self.store.update_friend_claim_links(task_id, friend_link_1, friend_link_2)
            self.store.update_friend_claim_state(task_id, "处理中", "好友链接已采集")
            self.store.add_event(task_id, "好友码提货", "已采集两条不同的Steam好友邀请链接")
            self.friend_claim_progress_var.set("好友码提货：好友链接已采集")
            self.status_var.set("已采集两条好友邀请链接，可打开提货页")
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_claim_link_saved(
        self,
        task_id: int,
        link_number: int,
        friend_link: str,
    ) -> None:
        def update() -> None:
            self.store.update_friend_claim_link(task_id, link_number, friend_link)
            self.store.add_event(
                task_id,
                "好友码提货",
                f"已写入好友链接{link_number}",
            )
            self.friend_claim_progress_var.set(f"好友码提货：已写入好友链接{link_number}")
            self.status_var.set(f"好友链接{link_number}已写入任务表")
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_claim_ready(self, task_id: int, message: str) -> None:
        def update() -> None:
            self.store.update_friend_claim_state(task_id, "处理中", "提货页已打开")
            self.store.add_event(task_id, "好友码提货", message)
            self.friend_claim_progress_var.set("好友码提货：提货页已打开")
            self.status_var.set(message)
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_claim_result_captured(self, task_id: int, path: Path) -> None:
        def update() -> None:
            self.store.update_friend_claim_state(task_id, "处理中", "提货结果已截图", str(path))
            self.store.add_event(task_id, "好友码提货", "已保存提货结果截图", str(path))
            self.friend_claim_progress_var.set("好友码提货：提货结果已截图")
            self.status_var.set(f"已保存提货结果截图：{path}")
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_claim_issue(self, task_id: int, message: str) -> None:
        def update() -> None:
            self.store.update_friend_claim_state(task_id, "需要人工处理", "需要人工处理")
            self.store.add_event(task_id, "好友码提货暂停", message)
            self.friend_claim_progress_var.set(f"好友码提货：已暂停，{message}")
            self.status_var.set(message)
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_claim_finished(self, task_id: int) -> None:
        def update() -> None:
            self.store.update_friend_claim_state(task_id, "成功", "提货完成并已登出")
            self.store.add_event(task_id, "好友码提货", "提货完成并确认登出")
            self.friend_claim_progress_var.set("好友码提货：完成并已登出")
            self.status_var.set("好友码提货完成并已登出")
            if self.current_task_id == task_id:
                next_task = self.store.next_task_after(task_id, "好友码提货")
                self.set_current(next_task.id if next_task else None)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_friend_game_claim_batch_started(self, task_count: int) -> None:
        def update() -> None:
            self.auto_running = True
            self.friend_claim_progress_var.set(f"好友码提货：自动游戏领取准备处理 {task_count} 家账号")
            self.status_var.set("自动游戏领取已启动")

        self.root.after(0, update)

    def ui_friend_game_claim_batch_task_started(
        self,
        task: Task,
        task_number: int,
        task_count: int,
    ) -> None:
        def update() -> None:
            self.set_current(task.id)
            self.store.update_friend_claim_state(
                task.id,
                "处理中",
                f"正在自动游戏领取（{task_number}/{task_count}）",
            )
            self.store.add_event(
                task.id,
                "自动游戏领取",
                f"开始账号 {task_number}/{task_count}",
            )
            self.friend_claim_progress_var.set(
                f"好友码提货：自动游戏领取 {task_number}/{task_count}，{task.store_code}"
            )
            self.status_var.set(f"正在领取 {task.store_code} 的礼物游戏")
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_friend_game_claim_step(
        self,
        task_id: int,
        step: str,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.add_event(
                task_id,
                "自动游戏领取",
                f"{step}：{message}",
                str(screenshot_path or ""),
            )
            self.friend_claim_progress_var.set(f"好友码提货：自动游戏领取-{step}")
            self.status_var.set(message)
            if self.current_task_id == task_id:
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_game_claim_finished(
        self,
        task_id: int,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.update_friend_claim_state(
                task_id,
                "成功",
                "自动游戏领取完成",
                str(screenshot_path or ""),
            )
            self.store.add_event(
                task_id,
                "自动游戏领取",
                message,
                str(screenshot_path or ""),
            )
            self.friend_claim_progress_var.set("好友码提货：自动游戏领取完成")
            self.status_var.set(message)
            if self.current_task_id == task_id:
                next_task = self.store.next_task_after(task_id, "好友码提货")
                self.set_current(next_task.id if next_task else None)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_friend_game_claim_issue(self, task_id: int, message: str) -> None:
        self.auto_pause_event.set()

        def update() -> None:
            self.store.update_friend_claim_state(task_id, "需要人工处理", f"自动游戏领取暂停：{message}")
            self.store.add_event(task_id, "自动游戏领取暂停", message)
            self.friend_claim_progress_var.set(f"好友码提货：自动游戏领取暂停，{message}")
            self.status_var.set(message)
            self.refresh_tasks()
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_friend_game_claim_batch_finished(self, stopped: bool) -> None:
        def update() -> None:
            self.auto_running = False
            self.auto_pause_event.clear()
            self.auto_stop_event.clear()
            text = "好友码提货：自动游戏领取已停止" if stopped else "好友码提货：自动游戏领取完成"
            self.friend_claim_progress_var.set(text)
            self.status_var.set(text)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_activation_step(
        self,
        task_id: int,
        step: str,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.add_event(
                task_id,
                "激活码激活",
                f"{step}：{message}",
                str(screenshot_path or ""),
            )
            self.activation_progress_var.set(f"激活流程：{step}")
            self.status_var.set(message)
            if self.current_task_id == task_id:
                self.fill_logs(task_id)

        self.root.after(0, update)

    def ui_activation_issue(
        self,
        task_id: int,
        message: str,
        screenshot_path: Path | None = None,
    ) -> None:
        def update() -> None:
            self.store.update_status(task_id, "需要人工处理")
            self.store.add_event(
                task_id,
                "激活暂停",
                message,
                str(screenshot_path or ""),
            )
            self.activation_progress_var.set(f"激活流程：已暂停，{message}")
            self.status_var.set(message)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_activation_code_started(self, task_id: int, code_index: int, code: str, code_count: int) -> None:
        def update() -> None:
            self.current_code_index = code_index
            self.store.upsert_voucher_result(task_id, code_index, code, "激活中")
            self.store.add_event(task_id, "全自动激活", f"开始激活码 {code_index + 1}/{code_count}")
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                iid = str(code_index)
                if iid in self.code_tree.get_children():
                    self.code_tree.selection_set(iid)
                    self.code_tree.focus(iid)
            self.activation_progress_var.set(f"激活流程：激活码 {code_index + 1}/{code_count}")

        self.root.after(0, update)

    def ui_activation_code_result(
        self,
        task_id: int,
        code_index: int,
        code: str,
        status: str,
        message: str,
        screenshot_path: Path,
    ) -> None:
        def update() -> None:
            self.store.upsert_voucher_result(
                task_id,
                code_index,
                code,
                status,
                message,
                str(screenshot_path),
            )
            self.store.add_event(
                task_id,
                "激活结果",
                f"第{code_index + 1}个：{status}；{message[:180]}",
                str(screenshot_path),
            )
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)
            self.activation_progress_var.set(f"激活流程：第{code_index + 1}个码 {status}")

        self.root.after(0, update)

    def ui_activation_task_started(self, task: Task, task_number: int, task_count: int) -> None:
        def update() -> None:
            self.set_current(task.id)
            self.store.update_status(task.id, "处理中")
            self.store.add_event(task.id, "全自动激活", f"开始账号 {task_number}/{task_count}")
            self.activation_progress_var.set(
                f"激活流程：账号 {task_number}/{task_count}，{task.store_code}"
            )
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_activation_batch_started(self, task_count: int) -> None:
        def update() -> None:
            self.auto_running = True
            self.activation_progress_var.set(f"激活流程：准备处理 {task_count} 个账号")
            self.status_var.set("全自动激活已启动")

        self.root.after(0, update)

    def ui_activation_batch_paused(self, task_id: int, reason: str) -> None:
        self.auto_pause_event.set()

        def update() -> None:
            self.store.update_status(task_id, "需要人工处理")
            self.store.add_event(task_id, "全自动激活暂停", reason)
            self.activation_progress_var.set(f"激活流程：批量已暂停，{reason}")
            self.status_var.set(reason)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_activation_finished(self, task_id: int) -> None:
        def update() -> None:
            self.store.update_status(task_id, "成功")
            self.store.add_event(task_id, "激活码激活", "激活完成并确认登出")
            self.activation_progress_var.set("激活流程：完成并已登出")
            if self.current_task_id == task_id:
                next_task = self.store.next_task_after(task_id, "激活码激活")
                self.set_current(next_task.id if next_task else None)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_activation_batch_finished(self, stopped: bool) -> None:
        def update() -> None:
            self.auto_running = False
            self.auto_pause_event.clear()
            self.auto_stop_event.clear()
            text = "激活流程：批量已停止" if stopped else "激活流程：批量完成"
            self.activation_progress_var.set(text)
            self.status_var.set(text)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_auto_batch_started(self, task_count: int) -> None:
        def update() -> None:
            self.auto_running = True
            self.auto_progress_var.set(f"全自动：准备处理 {task_count} 个账号")
            self.status_var.set("全自动兑换已启动")

        self.root.after(0, update)

    def ui_auto_task_started(self, task: Task, task_number: int, task_count: int) -> None:
        def update() -> None:
            self.set_current(task.id)
            self.store.update_status(task.id, "处理中")
            self.store.add_event(task.id, "全自动", f"开始账号 {task_number}/{task_count}")
            self.auto_progress_var.set(
                f"全自动：账号 {task_number}/{task_count}，{task.store_code}"
            )
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_auto_code_started(self, task_id: int, code_index: int, code: str, code_count: int) -> None:
        def update() -> None:
            self.current_code_index = code_index
            self.store.upsert_voucher_result(task_id, code_index, code, "兑换中")
            self.store.add_event(task_id, "全自动", f"开始兑换码 {code_index + 1}/{code_count}")
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                iid = str(code_index)
                if iid in self.code_tree.get_children():
                    self.code_tree.selection_set(iid)
                    self.code_tree.focus(iid)
            self.auto_progress_var.set(f"全自动：兑换码 {code_index + 1}/{code_count}")

        self.root.after(0, update)

    def ui_auto_code_result(
        self,
        task_id: int,
        code_index: int,
        code: str,
        status: str,
        message: str,
        screenshot_path: Path,
    ) -> None:
        def update() -> None:
            self.store.upsert_voucher_result(
                task_id,
                code_index,
                code,
                status,
                message,
                str(screenshot_path),
            )
            self.store.add_event(
                task_id,
                "兑换结果",
                f"第{code_index + 1}个：{status}；{message[:180]}",
                str(screenshot_path),
            )
            if self.current_task_id == task_id:
                self.fill_codes(self.current_task())
                self.fill_logs(task_id)
            self.auto_progress_var.set(f"全自动：第{code_index + 1}个码 {status}")

        self.root.after(0, update)

    def ui_auto_paused(self, task_id: int, reason: str) -> None:
        self.auto_pause_event.set()

        def update() -> None:
            self.store.update_status(task_id, "需要人工处理")
            self.store.add_event(task_id, "全自动暂停", reason)
            self.auto_progress_var.set(f"全自动：已暂停，{reason}")
            self.status_var.set(reason)
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_auto_task_finished(self, task_id: int, status: str) -> None:
        def update() -> None:
            self.store.update_status(task_id, status)
            self.store.add_event(task_id, "全自动", f"账号处理完成：{status}")
            self.refresh_tasks()

        self.root.after(0, update)

    def ui_auto_batch_finished(self, stopped: bool) -> None:
        def update() -> None:
            self.auto_running = False
            self.auto_pause_event.clear()
            self.auto_stop_event.clear()
            text = "全自动：已停止" if stopped else "全自动：批次完成"
            self.auto_progress_var.set(text)
            self.status_var.set(text)
            self.refresh_tasks()

        self.root.after(0, update)

    def log(self, event_type: str, message: str) -> None:
        task = self.current_task()
        if task:
            self.store.add_event(task.id, event_type, message)
            self.fill_logs(task.id)
        self.status_var.set(message)


ASSASSINS_CREED_PRODUCT = {
    "official_name": "刺客信条：黑旗 记忆重置",
    "license_names": [
        "刺客信条",
        "刺客信条：黑旗",
        "刺客信条：黑旗 记忆重置",
        "Assassin's Creed",
        "Assassin's Creed Black Flag",
        "Assassin's Creed Black Flag Resynced",
    ],
}

GAME_CATALOG.update(
    {
        "刺客信条": ASSASSINS_CREED_PRODUCT,
        "刺客信条黑旗": ASSASSINS_CREED_PRODUCT,
        "刺客信条：黑旗 记忆重置": ASSASSINS_CREED_PRODUCT,
        "Assassin's Creed Black Flag Resynced": ASSASSINS_CREED_PRODUCT,
    }
)


def _fast_open_gift_notification_page(self: SteamBrowserController, task: Task, product: dict) -> bool:
    candidate_urls: list[str] = []
    profile_link = (task.profile_link or "").strip().rstrip("/")
    if profile_link:
        candidate_urls.append(f"{profile_link}/inventory/#pending_gifts")
    candidate_urls.append(STEAM_PENDING_GIFTS_URL)

    seen: set[str] = set()
    for url in candidate_urls:
        if self._stopped():
            return False
        if url in seen:
            continue
        seen.add(url)
        if not self._goto_friend_claim_page(url, "待收礼物页面"):
            continue
        try:
            self.page.wait_for_load_state("domcontentloaded", timeout=6000)
        except Exception:
            pass
        time.sleep(0.8)
        _raise_if_steam_e502(self, "待收礼物页面")
        self._log(f"已直达待收礼物页面：{url}")
        return True
    return False


def _fast_click_positive_gift_claim_control(self: SteamBrowserController) -> str:
    self._ensure_page()
    try:
        target = self.page.evaluate(
            """
            () => {
                const addPattern = /添加.{0,10}(我的)?库|添加至我的库|添加到我的库|Add to.{0,12}Library/i;
                const acceptPattern = /接受礼物|接收礼物|领取礼物|收下礼物|Accept Gift|Redeem Gift/i;
                const negative = /拒绝|拒收|取消|退回|删除|Decline|Reject|Cancel|Return|Remove|Delete/i;
                const visible = (element) => {
                    const style = window.getComputedStyle(element);
                    const rect = element.getBoundingClientRect();
                    return style.display !== 'none'
                        && style.visibility !== 'hidden'
                        && rect.width > 3
                        && rect.height > 3
                        && !element.disabled;
                };
                const textOf = (element) => [
                    element.innerText,
                    element.textContent,
                    element.value,
                    element.id,
                    element.className,
                    element.getAttribute('onclick'),
                    element.getAttribute('title'),
                    element.getAttribute('aria-label'),
                    element.getAttribute('data-tooltip-text'),
                ].filter(Boolean).join(' ');
                const nodes = Array.from(document.querySelectorAll(
                    'button, a, input, [role="button"], .btn_green_steamui, .btnv6_green_white_innerfade, .btn_green_white_innerfade, span, div'
                ));
                const candidates = [];
                for (const node of nodes) {
                    const element = node.closest('button, a, input, [role="button"], .btn_green_steamui, .btnv6_green_white_innerfade, .btn_green_white_innerfade') || node;
                    if (candidates.some((item) => item.element === element) || !visible(element)) continue;
                    const text = textOf(element);
                    if (!text || negative.test(text)) continue;
                    const action = addPattern.test(text) ? '添加至我的库' : (acceptPattern.test(text) ? '接受礼物' : '');
                    if (!action) continue;
                    candidates.push({ element, text, action });
                }
                candidates.sort((a, b) => {
                    const ap = a.action === '添加至我的库' ? 0 : 1;
                    const bp = b.action === '添加至我的库' ? 0 : 1;
                    return ap - bp;
                });
                const target = candidates[0];
                if (!target) return null;
                target.element.scrollIntoView({ block: 'center', inline: 'center' });
                const rect = target.element.getBoundingClientRect();
                return {
                    action: target.action,
                    text: target.text.slice(0, 80),
                    x: rect.left + rect.width / 2,
                    y: rect.top + rect.height / 2,
                };
            }
            """
        )
        if not target:
            return ""
        self.page.mouse.move(float(target["x"]), float(target["y"]))
        time.sleep(0.18)
        self.page.mouse.down()
        time.sleep(0.12)
        self.page.mouse.up()
        action = str(target.get("action") or "礼物领取确认")
        self._log(f"已点击礼物控件：{action}")
        return action
    except Exception as exc:
        self._log(f"快速点击礼物控件失败：{exc}")
        return ""


def _fast_page_text(self: SteamBrowserController) -> str:
    try:
        return str(self.page.evaluate("() => document.body ? document.body.innerText : ''") or "")
    except Exception:
        return ""


class SteamE502Error(RuntimeError):
    pass


def _text_has_steam_e502(text: str) -> bool:
    lowered = str(text or "").lower()
    return (
        "e502" in lowered
        or ("something went wrong" in lowered and "unable to service your request" in lowered)
        or "we were unable to service your request" in lowered
    )


def _steam_e502_visible(self: SteamBrowserController) -> bool:
    return _text_has_steam_e502(_fast_page_text(self))


def _wait_for_steam_e502_recovery(
    self: SteamBrowserController,
    context: str,
    delays: tuple[int, ...] = (20, 45, 90, 180),
) -> bool:
    if not _steam_e502_visible(self):
        return True
    for attempt, delay in enumerate(delays, start=1):
        if self._stopped():
            return False
        self._log(f"{context} 出现 Steam E502 L2，冷却 {delay} 秒后重试（{attempt}/{len(delays)}）")
        time.sleep(delay)
        try:
            self.page.reload(wait_until="domcontentloaded", timeout=45000)
        except Exception as exc:
            self._log(f"{context} E502 重载失败，继续检测：{exc}")
        try:
            self.page.wait_for_load_state("domcontentloaded", timeout=8000)
        except Exception:
            pass
        time.sleep(1.5)
        if not _steam_e502_visible(self):
            self._log(f"{context} E502 已恢复")
            return True
    return False


def _raise_if_steam_e502(self: SteamBrowserController, context: str) -> None:
    if _steam_e502_visible(self) and not _wait_for_steam_e502_recovery(self, context):
        raise SteamE502Error(f"{context} 持续返回 Steam E502 L2")


def _gift_claim_page_state(self: SteamBrowserController) -> dict:
    try:
        return dict(
            self.page.evaluate(
                """
                () => {
                    const addPattern = /添加.{0,10}(我的)?库|添加至我的库|添加到我的库|Add to.{0,12}Library/i;
                    const acceptPattern = /接受礼物|接收礼物|领取礼物|收下礼物|Accept Gift|Redeem Gift/i;
                    const successPattern = /已添加到您的库|已加入您的库|已经在您的库|添加到了您的 Steam 库|added to your library|already in your library|successfully redeemed|gift has been accepted/i;
                    const negative = /拒绝|拒收|取消|退回|删除|Decline|Reject|Cancel|Return|Remove|Delete/i;
                    const visible = (element) => {
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return style.display !== 'none'
                            && style.visibility !== 'hidden'
                            && rect.width > 3
                            && rect.height > 3
                            && !element.disabled;
                    };
                    const textOf = (element) => [
                        element.innerText,
                        element.textContent,
                        element.value,
                        element.id,
                        element.className,
                        element.getAttribute('onclick'),
                        element.getAttribute('title'),
                        element.getAttribute('aria-label'),
                    ].filter(Boolean).join(' ');
                    const bodyText = document.body ? document.body.innerText || '' : '';
                    let addVisible = false;
                    let acceptVisible = false;
                    for (const node of Array.from(document.querySelectorAll('button, a, input, [role="button"], .btn_green_steamui, .btnv6_green_white_innerfade, .btn_green_white_innerfade, span, div'))) {
                        const element = node.closest('button, a, input, [role="button"], .btn_green_steamui, .btnv6_green_white_innerfade, .btn_green_white_innerfade') || node;
                        if (!visible(element)) continue;
                        const text = textOf(element);
                        if (!text || negative.test(text)) continue;
                        addVisible = addVisible || addPattern.test(text);
                        acceptVisible = acceptVisible || acceptPattern.test(text);
                    }
                    return {
                        addVisible,
                        acceptVisible,
                        successVisible: successPattern.test(bodyText),
                    };
                }
                """
            )
            or {}
        )
    except Exception:
        return {}


def _fast_claim_gift_on_current_page(self: SteamBrowserController, task: Task, product: dict) -> bool:
    clicked_any = False
    success_markers = [
        "已添加到您的库",
        "已加入您的库",
        "已经在您的库",
        "添加到了您的 Steam 库",
        "added to your library",
        "already in your library",
        "successfully redeemed",
        "gift has been accepted",
    ]
    deadline = time.time() + 38

    while time.time() < deadline:
        if self._stopped():
            return False
        _raise_if_steam_e502(self, "礼物领取页面")
        state = _gift_claim_page_state(self)
        if state.get("successVisible"):
            self._log("页面显示礼物已领取")
            return True
        self._select_add_to_library_option()
        action = self._click_positive_gift_claim_control()
        if action:
            clicked_any = True
            if "库" in action or "library" in action.lower():
                add_deadline = time.time() + 9
                retried_add_click = False
                while time.time() < add_deadline:
                    if self._stopped():
                        return False
                    time.sleep(0.8)
                    state = _gift_claim_page_state(self)
                    if state.get("successVisible"):
                        self._log("页面显示礼物已加入库")
                        return True
                    if not state.get("addVisible"):
                        time.sleep(2.5)
                        self._log("添加至我的库按钮已消失，转入最终资产校验")
                        return True
                    if not retried_add_click:
                        retried_add_click = True
                        self._log("添加至我的库按钮仍可见，补点一次")
                        self._click_positive_gift_claim_control()
                self._log("添加至我的库后等待完成，转入最终资产校验")
                return True

            accept_deadline = time.time() + 8
            while time.time() < accept_deadline:
                if self._stopped():
                    return False
                time.sleep(0.6)
                state = _gift_claim_page_state(self)
                if state.get("successVisible") or state.get("addVisible"):
                    break
            continue

        body_text = _fast_page_text(self)
        lowered = body_text.lower()
        if any(marker.lower() in lowered for marker in success_markers):
            self._log("页面显示礼物已领取")
            return True
        time.sleep(0.55 if clicked_any else 0.8)

    if clicked_any:
        self._log("已点击过领取控件，转入最终资产校验")
    else:
        self._log("待收礼物页未找到可点击的领取控件")
    return clicked_any


def _fast_verify_friend_game_asset(self: SteamBrowserController, task: Task, status_text: str) -> Path | None:
    product = self._friend_claim_product(task)
    if product is None:
        return None
    opened = False
    for attempt in range(3):
        try:
            self._goto(STEAM_LICENSES_URL)
            opened = True
            break
        except Exception as exc:
            message = str(exc)
            if "ERR_HTTP_RESPONSE_CODE_FAILURE" in message or "502" in message:
                delay = (20, 45, 90)[attempt]
                self._log(f"资产清单页面疑似 Steam E502，冷却 {delay} 秒后重试 {attempt + 1}/3：{exc}")
                time.sleep(delay)
                continue
            self._log(f"资产清单页面打开失败，准备重试 {attempt + 1}/3：{exc}")
            time.sleep(2.5)
    if not opened:
        raise SteamE502Error("资产清单页面持续无法打开，疑似 Steam E502 L2")
    if self._stopped():
        return None
    try:
        self.page.wait_for_load_state("domcontentloaded", timeout=8000)
    except Exception:
        pass
    _raise_if_steam_e502(self, "资产清单页面")
    deadline = time.time() + 18
    while time.time() < deadline:
        body_text = _fast_page_text(self)
        if _text_has_steam_e502(body_text):
            _raise_if_steam_e502(self, "资产清单页面")
        if product_in_text(product, body_text):
            official_name = str(product["official_name"])
            label = status_text if official_name in status_text else f"{status_text}-{official_name}"
            return self._capture_page(task, label)
        time.sleep(0.8)
    self._log(f"资产清单中未找到《{product['official_name']}》")
    return None


def _fast_claim_friend_game_gift(self: SteamBrowserController, task: Task) -> bool:
    product = self._friend_claim_product(task)
    if product is None:
        return False
    if not self._login_friend_claim_task(task):
        return False

    if not self._open_gift_notification_page(task, product):
        already_path = self._verify_friend_game_asset(task, "自动游戏领取-资产已确认")
        if already_path is not None:
            if not self.logout_by_menu():
                self.app.ui_friend_game_claim_issue(task.id, "资产已存在，但登出未确认")
                return False
            self.app.ui_friend_game_claim_finished(
                task.id,
                f"未打开礼物页，但资产清单已存在《{product['official_name']}》",
                already_path,
            )
            return True
        self.app.ui_friend_game_claim_issue(task.id, "未能打开待收礼物页，且资产清单尚未出现目标游戏")
        return False

    self.app.ui_friend_game_claim_step(task.id, "待收礼物页", "已直达待收礼物页面")
    if not self._claim_gift_on_current_page(task, product):
        existing_path = self._verify_friend_game_asset(task, "自动游戏领取-资产已确认")
        if existing_path is not None:
            if not self.logout_by_menu():
                self.app.ui_friend_game_claim_issue(task.id, "资产已存在，但登出未确认")
                return False
            self.app.ui_friend_game_claim_finished(
                task.id,
                f"礼物控件未再次出现，但资产清单已存在《{product['official_name']}》",
                existing_path,
            )
            return True
        self.app.ui_friend_game_claim_issue(task.id, "礼物页面未能自动完成领取，请人工检查当前页面")
        return False

    path = self._verify_friend_game_asset(task, "自动游戏领取-资产已确认")
    if path is None:
        self.app.ui_friend_game_claim_issue(task.id, f"领取后资产清单中未找到《{product['official_name']}》")
        return False
    if not self.logout_by_menu():
        self.app.ui_friend_game_claim_issue(task.id, "领取后登出未确认")
        return False
    self.app.ui_friend_game_claim_finished(
        task.id,
        f"已确认《{product['official_name']}》进入账号资产列表",
        path,
    )
    return True


SteamBrowserController._open_gift_notification_page = _fast_open_gift_notification_page
SteamBrowserController._click_positive_gift_claim_control = _fast_click_positive_gift_claim_control
SteamBrowserController._claim_gift_on_current_page = _fast_claim_gift_on_current_page
SteamBrowserController._verify_friend_game_asset = _fast_verify_friend_game_asset
SteamBrowserController.claim_friend_game_gift = _fast_claim_friend_game_gift


def _ui_friend_game_claim_e502(self: SteamTaskAssistant, task_id: int, message: str) -> None:
    self.auto_pause_event.set()

    def update() -> None:
        self.store.update_friend_claim_state(
            task_id,
            "需要人工处理",
            "Steam E502 L2-待稍后重试",
            "",
        )
        self.store.add_event(task_id, "Steam E502 L2", message)
        self.friend_claim_progress_var.set("好友码提货：Steam E502 L2，已暂停等待稍后重试")
        self.status_var.set(message)
        self.refresh_tasks()
        if self.current_task_id == task_id:
            self.fill_codes(self.current_task())
            self.fill_logs(task_id)

    self.root.after(0, update)


def _e502_safe_claim_friend_game_gift(self: SteamBrowserController, task: Task) -> bool:
    try:
        return _fast_claim_friend_game_gift(self, task)
    except SteamE502Error as exc:
        message = f"{exc}；这通常是 Steam 服务端限流/暂时不可用，已冷却重试但仍失败。请稍后单独重试该账号。"
        self._log(message)
        self.app.ui_friend_game_claim_e502(task.id, message)
        return False


SteamTaskAssistant.ui_friend_game_claim_e502 = _ui_friend_game_claim_e502
SteamBrowserController.claim_friend_game_gift = _e502_safe_claim_friend_game_gift


def main() -> None:
    ensure_dirs()
    root = Tk()
    try:
        ttk.Style().theme_use("clam")
    except Exception:
        pass
    SteamTaskAssistant(root)
    root.mainloop()


if __name__ == "__main__":
    main()
